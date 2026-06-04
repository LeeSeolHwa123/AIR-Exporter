import os
import json
import re
import time
import ctypes
import shutil
import threading
import traceback
from ctypes import wintypes
from datetime import date, datetime, timedelta
from pathlib import Path
from tkinter import messagebox

from pywinauto import Desktop
from pywinauto.keyboard import send_keys

try:
    import customtkinter as ctk
except Exception:
    ctk = None


# =========================================================
# 0. 기본 설정
# =========================================================

CODE_VERSION = "AIR_DAILY_EXPORT_MASTER_GLOBAL_LATEST_V11_FAST_DEPARTMENT_CHANGE"
GUI_VERSION = "QEX AIR Exporter v1 (2026-06-04)"

STOP_REQUESTED = False
LOG_LISTENERS = []


class UserStopRequested(Exception):
    pass

try:
    ctypes.windll.user32.SetProcessDPIAware()
except Exception:
    pass


# =========================================================
# 1. 사용자 설정
# =========================================================

AIR_SHORTCUT_PATH = ""

CURRENT_WINDOWS_USER = Path.home().name

DEFAULT_EXPORT_DIR = rf"C:\Users\{CURRENT_WINDOWS_USER}\OneDrive - kochind.com\QIC_QEX - 98. AIR Inspection Monitoring\AIR Raw Data"
DEFAULT_MASTER_EXCEL_PATH = rf"C:\Users\{CURRENT_WINDOWS_USER}\Downloads\History_Master.xlsx"

EXPORT_DIR = DEFAULT_EXPORT_DIR

# 이번 실행에서 내보낸 Daily 파일만 명확히 구분하기 위한 실행 단위 폴더
# 예: C:/Users/seolhl/Downloads/AIR_Daily_Export/20260601_1
RUN_DATE = time.strftime("%Y%m%d")
RUN_EXPORT_DIR = None

DEPARTMENT_JOBS = [
    {
        "department": "일반사출",
        "combo_index": 3,
    },
    {
        "department": "스탬핑",
        "combo_index": 1,
    },
    {
        "department": "도금",
        "combo_index": 2,
    },
]

DAILY_START_TIME = "08:30" #20:30으로 설정하면 휴일 전날 아침에 20:30으로 조회했을 경우 휴일 이후에 조회 시, 휴일 전날 아침~20:30까지의 데이터 누락 위험이 있음 
                           #안전하게 무조건 08:30으로 설정하되, 이러면 중복이 발생하므로 모든 업데이트 이후 중복 삭제
DAILY_END_TIME = "20:30"

SEARCH_WAIT_SECONDS = 0.2
SAVE_DIALOG_WAIT_SECONDS = 0.7

# 속도 최적화: 조회/저장 완료 팝업은 길게 전체 창 스캔하지 않고
# foreground 작은 팝업을 먼저 즉시 Enter 처리합니다.
FAST_POPUP_TIMEOUT_SECONDS = 1.2
POST_EXPORT_DIALOG_WAIT_SECONDS = 0.5
FILE_CHECK_INTERVAL_SECONDS = 0.2

HANDLE_SAVE_DIALOG = True

# 반복 조회 중 같은 AIR 화면 컨트롤을 매번 전체 탐색하지 않기 위한 캐시
CONTROL_CACHE = {}

# =========================================================
# 1-1. 마스터 엑셀 업데이트 설정
# =========================================================

# AIR에서 부서별 Daily 파일을 내보낸 뒤, 아래 마스터 파일에 자동 삽입합니다.
UPDATE_MASTER_AFTER_EXPORT = True

MASTER_EXCEL_PATH = DEFAULT_MASTER_EXCEL_PATH
MASTER_SHEET_NAME = "List"

# 마스터 파일이 열려 있으면 저장에 실패할 수 있으므로, 업데이트 전에 Excel에서 닫아두는 것을 권장합니다.
CREATE_MASTER_BACKUP = True

# 중복 기준: A열 순번 제외, B~마지막열 값이 모두 같으면 중복으로 판단합니다.
DUPLICATE_COMPARE_START_COL = 2

# 마스터 구조 기준: A열=부서별 순번, B열=부서명
MASTER_INDEX_COL = 1
MASTER_DEPARTMENT_COL = 2

# AIR 조회기간 자동 산정 기준
# 마스터 파일을 먼저 열어서 전체 부서 중 '최근검사' 열의 가장 최근 날짜를 From으로 사용합니다.
MASTER_RECENT_INSPECTION_HEADER = "최근검사"
USE_MASTER_LATEST_RECENT_DATE_FOR_FROM = True
QUERY_START_FALLBACK_DAYS = 1

# True: 신규 Daily 데이터는 마스터 전체 마지막 행 아래에 추가합니다.
# False: 기존처럼 해당 부서의 마지막 행 바로 아래에 삽입합니다.
APPEND_NEW_ROWS_TO_BOTTOM = True


def get_settings_file_path() -> Path:
    app_data = os.getenv("APPDATA")
    if app_data:
        base_dir = Path(app_data)
    else:
        base_dir = Path.home()
    return base_dir / "QEXAirExporter" / "settings.json"


def load_user_path_settings() -> dict:
    settings_path = get_settings_file_path()

    if not settings_path.exists():
        return {}

    try:
        with settings_path.open("r", encoding="utf-8") as fp:
            data = json.load(fp)

        if not isinstance(data, dict):
            return {}

        return data
    except Exception as e:
        print(f"[WARN] 사용자 설정 파일 로드 실패: {settings_path} / {e}", flush=True)
        return {}


def remap_path_to_current_windows_user(path: str) -> str:
    """
    저장된 경로의 사용자명(C:\\Users\\<name>)만 현재 사용자로 치환합니다.
    나머지 경로는 그대로 유지합니다.
    """
    value = str(path or "").strip()
    if not value:
        return value

    match = re.match(r"^([A-Za-z]:\\Users\\)([^\\]+)(\\.*)$", value, re.IGNORECASE)
    if not match:
        return value

    return f"{match.group(1)}{CURRENT_WINDOWS_USER}{match.group(3)}"


def save_user_path_settings(export_dir: str, master_excel_path: str):
    settings_path = get_settings_file_path()
    settings_path.parent.mkdir(parents=True, exist_ok=True)

    data = {
        "export_dir": str(export_dir or "").strip(),
        "master_excel_path": str(master_excel_path or "").strip(),
    }

    with settings_path.open("w", encoding="utf-8") as fp:
        json.dump(data, fp, ensure_ascii=False, indent=2)


def apply_saved_path_settings():
    global EXPORT_DIR
    global MASTER_EXCEL_PATH

    settings = load_user_path_settings()

    export_dir = remap_path_to_current_windows_user(settings.get("export_dir", ""))
    master_excel_path = remap_path_to_current_windows_user(settings.get("master_excel_path", ""))

    if export_dir:
        EXPORT_DIR = export_dir

    if master_excel_path:
        MASTER_EXCEL_PATH = master_excel_path


apply_saved_path_settings()


# =========================================================
# 2. AIR 컨트롤 설정
# =========================================================

AIR_TITLE_PATTERNS = [
    r".*CCS AIR System.*",
    r".*OPC AIR System.*",
    r".*검사기록조회.*",
]

INSPECTION_BUTTON_AUTO_ID = "button6"
INSPECTION_BUTTON_TITLE = "검사이력"

INSPECTION_WINDOW_KEYWORD = "검사기록조회"

INSPECTION_FORM_AUTO_ID = "frmSetupList_1"
INSPECTION_FORM_TITLE = "검사기록조회"

PLANT_COMBO_AUTO_ID = "cboPlant"
DEPARTMENT_COMBO_AUTO_ID = "cboCostCenter"

START_DATE_PICKER_AUTO_ID = "dFrom"
END_DATE_PICKER_AUTO_ID = "dTo"

START_TIME_COMBO_AUTO_ID = "cboStart"
END_TIME_COMBO_AUTO_ID = "cboEnd"

CONDITION_PANEL_AUTO_ID = "panel1"
PERIOD_GROUP_AUTO_ID = "gbPeroid"

INSPECTION_SEARCH_BUTTON_AUTO_ID = "button1"

EXCEL_BUTTON_AUTO_ID = "cmdExcel"

FIELD_AUTO_IDS = {
    "제품번호": "txtPNo",
    "기계번호": "txtmcno",
    "Lot No": "txtLotNo",
    "Barcode": "txtBarcode",
    "툴번호": "txtToolNo",
    "PO No": "txtPOno",
    "판정": "txtDisposition",
}


# =========================================================
# 3. Win32 ComboBox 메시지 상수
# =========================================================
CB_GETCOUNT = 0x0146
CB_GETCURSEL = 0x0147
CB_GETLBTEXT = 0x0148
CB_GETLBTEXTLEN = 0x0149
CB_SETCURSEL = 0x014E

WM_COMMAND = 0x0111
CBN_SELCHANGE = 1
CBN_SELENDOK = 9

WPARAM_T = ctypes.c_size_t
LPARAM_T = ctypes.c_ssize_t
LRESULT_T = ctypes.c_ssize_t

_SendMessageW = ctypes.windll.user32.SendMessageW
_SendMessageW.argtypes = [wintypes.HWND, wintypes.UINT, WPARAM_T, LPARAM_T]
_SendMessageW.restype = LRESULT_T


# =========================================================
# 4. 공통 유틸
# =========================================================

def log(msg: str):
    message = str(msg)
    print(message, flush=True)

    for listener in list(LOG_LISTENERS):
        try:
            listener(message)
        except Exception:
            pass


def register_log_listener(listener):
    if listener not in LOG_LISTENERS:
        LOG_LISTENERS.append(listener)


def unregister_log_listener(listener):
    if listener in LOG_LISTENERS:
        LOG_LISTENERS.remove(listener)


def request_stop():
    global STOP_REQUESTED
    STOP_REQUESTED = True


def ensure_not_stopped():
    if STOP_REQUESTED:
        raise UserStopRequested("사용자 중지 요청")


def normalize_text(text: str) -> str:
    return str(text).strip().replace(" ", "")


def set_clipboard_text(text: str):
    import tkinter as tk

    root = tk.Tk()
    root.withdraw()
    root.clipboard_clear()
    root.clipboard_append(str(text))
    root.update()
    root.destroy()


def parse_date_ymd(value: str):
    value = str(value).strip()

    if "-" in value:
        parts = value.split("-")
    elif "/" in value:
        parts = value.split("/")
    elif "." in value:
        parts = value.split(".")
    else:
        value = value.replace(" ", "")
        if len(value) != 8:
            raise ValueError(f"날짜 형식 오류: {value}")
        return value[0:4], value[4:6], value[6:8]

    if len(parts) != 3:
        raise ValueError(f"날짜 형식 오류: {value}")

    return parts[0].zfill(4), parts[1].zfill(2), parts[2].zfill(2)


def create_run_export_dir() -> str:
    """
    오늘 날짜 기준 실행 폴더를 생성합니다.

    예:
    C:/Users/seolhl/Downloads/AIR_Daily_Export/20260601_1
    C:/Users/seolhl/Downloads/AIR_Daily_Export/20260601_2

    같은 날짜 폴더가 이미 있으면 뒤의 실행횟수를 1씩 증가시킵니다.
    """
    base_dir = Path(EXPORT_DIR)
    base_dir.mkdir(parents=True, exist_ok=True)

    run_no = 1

    while True:
        candidate = base_dir / f"{RUN_DATE}_{run_no}"

        if not candidate.exists():
            candidate.mkdir(parents=True, exist_ok=False)

            log("[RUN-FOLDER] 이번 실행 저장 폴더 생성 완료")
            log(f" - 기준 폴더: {base_dir}")
            log(f" - 실행 날짜: {RUN_DATE}")
            log(f" - 실행 횟수: {run_no}")
            log(f" - 저장 폴더: {candidate}")

            return str(candidate)

        run_no += 1


def make_export_path(department: str) -> str:
    """
    이번 실행 폴더 안에 부서별 Daily 엑셀 저장 경로를 생성합니다.

    예:
    C:/Users/seolhl/Downloads/AIR_Daily_Export/20260601_1/AIR_검사이력_일반사출_20260601_171000.xlsx
    """
    global RUN_EXPORT_DIR

    if RUN_EXPORT_DIR is None:
        RUN_EXPORT_DIR = create_run_export_dir()

    Path(RUN_EXPORT_DIR).mkdir(parents=True, exist_ok=True)

    safe_dept = re.sub(r'[\\/:*?"<>|]', "_", department)
    filename = f"AIR_검사이력_{safe_dept}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    export_path = os.path.join(RUN_EXPORT_DIR, filename)

    log("[EXPORT-PATH] 부서별 저장 경로 생성")
    log(f" - 부서명: {department}")
    log(f" - 저장 경로: {export_path}")

    return export_path


def wait_until_file_exists(path: str, timeout: int = 30, interval: float = None) -> bool:
    """
    파일 생성 확인.
    기존 1초 단위 polling은 저장 후 대기 시간이 길어져서 0.2초 단위로 줄입니다.
    """
    if interval is None:
        interval = FILE_CHECK_INTERVAL_SECONDS

    start = time.time()

    while time.time() - start < timeout:
        if os.path.exists(path) and os.path.getsize(path) > 0:
            return True
        time.sleep(interval)

    return False


def normalize_excel_file_path(path: str) -> str:
    """
    Excel COM의 Workbook.FullName과 비교하기 위한 경로 정규화.
    대소문자/슬래시 차이 때문에 같은 파일을 못 찾는 문제를 줄입니다.
    """
    return os.path.normcase(os.path.abspath(str(path)))


def extract_export_paths(export_items):
    """
    saved_exports 구조에서 실제 파일 경로만 추출합니다.

    허용 형태:
    1. ["C:\\...\\AIR_검사이력_일반사출.xlsx", ...]
    2. [{"department": "일반사출", "path": "C:\\...xlsx"}, ...]
    """
    paths = []

    for item in export_items:
        if isinstance(item, dict):
            path = (
                item.get("path")
                or item.get("file_path")
                or item.get("saved_path")
                or item.get("export_path")
            )
        else:
            path = item

        if not path:
            continue

        path = str(path)

        if os.path.exists(path):
            paths.append(path)
        else:
            log(f"[WARN] 내보낸 파일 경로가 존재하지 않아 닫기 대상에서 제외: {path}")

    return paths


def is_file_unlocked(path: str) -> bool:
    """
    파일이 Excel 등에 의해 잠겨 있는지 확인합니다.
    """
    try:
        with open(path, "r+b"):
            return True
    except PermissionError:
        return False
    except OSError:
        return False


def wait_until_export_files_unlocked(export_items, timeout: int = 10) -> bool:
    """
    내보낸 엑셀 파일들이 닫혀서 잠금 해제될 때까지 대기합니다.
    """
    paths = extract_export_paths(export_items)

    if not paths:
        return True

    start = time.time()

    while time.time() - start < timeout:
        locked = []

        for path in paths:
            if not is_file_unlocked(path):
                locked.append(path)

        if not locked:
            log("[OK] 내보낸 엑셀 파일 잠금 해제 확인 완료")
            return True

        log("[INFO] 아직 열려 있는 내보낸 엑셀 파일 대기 중:")
        for path in locked:
            log(f" - {path}")

        time.sleep(0.5)

    log("[WARN] 일부 내보낸 엑셀 파일 잠금이 해제되지 않았습니다.")
    return False


def close_only_exported_excel_workbooks(export_items, timeout: int = 20) -> bool:
    """
    Excel에서 열려 있는 워크북 중 이번 자동화로 내보낸 파일만 닫습니다.

    개선점:
    1. Excel COM Workbook.FullName 전체 경로 비교
    2. 경로 비교 실패 시 파일명 기준 fallback 비교
    3. 자동으로 열린 Excel이 늦게 COM에 등록될 수 있으므로 timeout 동안 반복 시도
    4. 닫기 실패 시 False 반환
    """
    paths = extract_export_paths(export_items)

    if not paths:
        log("[INFO] 닫을 내보낸 엑셀 파일이 없습니다.")
        return True

    target_paths = {
        normalize_excel_file_path(path): path
        for path in paths
    }

    target_file_names = {
        os.path.basename(path).lower(): path
        for path in paths
    }

    log("\n[EXCEL-CLOSE] 마스터 업데이트 전, 자동으로 열린 내보낸 엑셀 파일만 닫기 시작")
    for path in paths:
        log(f" - 닫기 대상: {path}")

    try:
        import win32com.client
    except ImportError:
        log("[WARN] pywin32가 설치되어 있지 않아 Excel 창 자동 닫기를 건너뜁니다.")
        log("       설치 명령: pip install pywin32")
        return False

    start = time.time()
    total_closed_count = 0

    while time.time() - start < timeout:
        try:
            excel = win32com.client.GetActiveObject("Excel.Application")
        except Exception:
            log("[INFO] 실행 중인 Excel COM 인스턴스를 찾지 못했습니다.")
            return wait_until_export_files_unlocked(paths, timeout=2)

        try:
            workbooks = excel.Workbooks
            workbook_count = workbooks.Count

            log(f"[EXCEL-CLOSE] 현재 Excel COM Workbook 수: {workbook_count}")

            closed_this_round = 0

            for i in range(workbook_count, 0, -1):
                wb = workbooks.Item(i)

                try:
                    wb_fullname = str(wb.FullName)
                except Exception:
                    continue

                normalized_wb_path = normalize_excel_file_path(wb_fullname)
                wb_file_name = os.path.basename(wb_fullname).lower()

                log(f"[EXCEL-CHECK] 열려 있는 Workbook: {wb_fullname}")

                matched = False

                # 1차: 전체 경로 정확 비교
                if normalized_wb_path in target_paths:
                    matched = True

                # 2차: 파일명 fallback
                # AIR export 파일명은 초 단위 timestamp가 들어가므로 파일명만으로도 충돌 가능성이 낮음
                elif wb_file_name in target_file_names:
                    matched = True

                if matched:
                    log(f"[EXCEL-CLOSE] 내보낸 파일 닫기: {wb_fullname}")
                    wb.Close(SaveChanges=False)
                    closed_this_round += 1
                    total_closed_count += 1
                    time.sleep(0.3)

            if closed_this_round > 0:
                log(f"[EXCEL-CLOSE] 이번 반복에서 닫은 파일 수: {closed_this_round}")

        except Exception as e:
            log(f"[WARN] Excel Workbook 닫기 중 오류: {e}")

        # 닫힌 뒤 실제 파일 잠금이 풀렸는지 확인
        if wait_until_export_files_unlocked(paths, timeout=2):
            log(f"[EXCEL-CLOSE] 닫은 내보낸 엑셀 파일 총 수: {total_closed_count}")
            return True

        log("[INFO] Excel 자동 실행/잠금 해제를 기다린 뒤 닫기 재시도")
        time.sleep(0.5)

    log(f"[EXCEL-CLOSE] 닫은 내보낸 엑셀 파일 총 수: {total_closed_count}")
    log("[WARN] timeout 동안 내보낸 엑셀 파일 잠금이 해제되지 않았습니다.")
    return False


def patch_openpyxl_print_page_setup_collated():
    """
    AIR가 만든 xlsx에 pageSetup collated 속성이 들어 있으면 openpyxl에서 아래 오류가 날 수 있습니다.

    PrintPageSetup.__init__() got an unexpected keyword argument 'collated'

    load_workbook() 전에 이 함수를 실행해서 openpyxl이 모르는 collated 속성만 무시하게 합니다.
    """
    try:
        from openpyxl.worksheet.page import PrintPageSetup

        original_init = PrintPageSetup.__init__

        if getattr(original_init, "_air_collated_patch", False):
            return

        def patched_init(self, *args, **kwargs):
            kwargs.pop("collated", None)
            return original_init(self, *args, **kwargs)

        patched_init._air_collated_patch = True
        PrintPageSetup.__init__ = patched_init

        log("[OK] openpyxl PrintPageSetup collated 무시 패치 적용 완료")

    except Exception as e:
        log(f"[WARN] openpyxl collated 패치 적용 실패: {e}")


def snapshot_excel_files():
    """
    이번 실행 폴더 안의 엑셀 파일 상태만 기록합니다.
    과거 실행 폴더나 Downloads 전체는 보지 않습니다.
    """
    global RUN_EXPORT_DIR

    result = {}

    if RUN_EXPORT_DIR is None:
        return result

    folder = Path(RUN_EXPORT_DIR)

    if not folder.exists():
        return result

    for ext in ("*.xlsx", "*.xls"):
        for file in folder.glob(ext):
            try:
                result[str(file.resolve())] = file.stat().st_mtime
            except Exception:
                pass

    return result


def wait_for_new_excel_file(before_snapshot: dict, timeout: int = 30):
    """
    이번 실행 폴더 안에서 새로 생성된 엑셀 파일만 찾습니다.
    과거 Daily 파일이나 Downloads의 다른 엑셀 파일은 무시합니다.
    """
    global RUN_EXPORT_DIR

    if RUN_EXPORT_DIR is None:
        return None

    folder = Path(RUN_EXPORT_DIR)

    if not folder.exists():
        return None

    start = time.time()

    while time.time() - start < timeout:
        candidates = []

        for ext in ("*.xlsx", "*.xls"):
            for file in folder.glob(ext):
                try:
                    full = str(file.resolve())
                    mtime = file.stat().st_mtime
                    size = file.stat().st_size

                    if size <= 0:
                        continue

                    if full not in before_snapshot:
                        candidates.append(file)
                    elif mtime > before_snapshot[full] + 0.5:
                        candidates.append(file)
                except Exception:
                    pass

        if candidates:
            candidates.sort(key=lambda f: f.stat().st_mtime, reverse=True)
            return str(candidates[0].resolve())

        time.sleep(FILE_CHECK_INTERVAL_SECONDS)

    return None


def copy_or_rename_export_file(src: str, dst: str):
    src_path = Path(src)
    dst_path = Path(dst)

    if src_path.resolve() == dst_path.resolve():
        return str(dst_path)

    dst_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        shutil.copy2(src_path, dst_path)
        return str(dst_path)
    except Exception:
        return str(src_path)


def is_real_visible_control(ctrl) -> bool:
    try:
        rect = ctrl.rectangle()

        if rect.width() <= 0 or rect.height() <= 0:
            return False

        if rect.left < -30000 or rect.top < -30000:
            return False

        if not ctrl.is_visible():
            return False

        if not ctrl.is_enabled():
            return False

        return True

    except Exception:
        return False


def is_inside_screen(x: int, y: int) -> bool:
    try:
        SM_XVIRTUALSCREEN = 76
        SM_YVIRTUALSCREEN = 77
        SM_CXVIRTUALSCREEN = 78
        SM_CYVIRTUALSCREEN = 79

        left = ctypes.windll.user32.GetSystemMetrics(SM_XVIRTUALSCREEN)
        top = ctypes.windll.user32.GetSystemMetrics(SM_YVIRTUALSCREEN)
        width = ctypes.windll.user32.GetSystemMetrics(SM_CXVIRTUALSCREEN)
        height = ctypes.windll.user32.GetSystemMetrics(SM_CYVIRTUALSCREEN)

        return left <= x <= left + width and top <= y <= top + height
    except Exception:
        return True


def real_mouse_click(x: int, y: int):
    ctypes.windll.user32.SetCursorPos(x, y)
    time.sleep(0.05)

    MOUSEEVENTF_LEFTDOWN = 0x0002
    MOUSEEVENTF_LEFTUP = 0x0004

    ctypes.windll.user32.mouse_event(MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    time.sleep(0.05)
    ctypes.windll.user32.mouse_event(MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    time.sleep(0.15)


def click_control_by_real_mouse(ctrl, label: str):
    rect = ctrl.rectangle()
    x = int((rect.left + rect.right) / 2)
    y = int((rect.top + rect.bottom) / 2)

    log(f"[INFO] {label} 실제 클릭: x={x}, y={y}, rect={rect}")

    if not is_inside_screen(x, y):
        raise RuntimeError(f"{label} 클릭 좌표가 화면 범위 밖입니다: ({x}, {y})")

    try:
        top = ctrl.top_level_parent()
        top.set_focus()
        time.sleep(0.1)
    except Exception:
        pass

    real_mouse_click(x, y)


def get_foreground_window_info():
    hwnd = ctypes.windll.user32.GetForegroundWindow()

    length = ctypes.windll.user32.GetWindowTextLengthW(hwnd)
    buff = ctypes.create_unicode_buffer(length + 1)
    ctypes.windll.user32.GetWindowTextW(hwnd, buff, length + 1)

    return hwnd, buff.value


# =========================================================
# 5. AIR 창 찾기 / 진입
# =========================================================

def launch_air_if_needed():
    if not AIR_SHORTCUT_PATH:
        log("[INFO] AIR 실행 생략: 현재 실행 중인 AIR에 연결합니다.")
        return

    if not os.path.exists(AIR_SHORTCUT_PATH):
        raise FileNotFoundError(f"AIR 바로가기를 찾지 못했습니다: {AIR_SHORTCUT_PATH}")

    os.startfile(AIR_SHORTCUT_PATH)
    log("[OK] AIR 실행 요청 완료")
    time.sleep(8)


def find_air_window(timeout=30):
    start = time.time()

    while time.time() - start < timeout:
        for w in Desktop(backend="uia").windows():
            try:
                title = w.window_text()

                if not title.strip():
                    continue

                for pattern in AIR_TITLE_PATTERNS:
                    if re.search(pattern, title):
                        log(f"[OK] AIR 창 발견: {title}")
                        return w

            except Exception:
                pass

        time.sleep(1)

    raise RuntimeError("AIR 창을 찾지 못했습니다. AIR가 실행 중인지 확인하세요.")


def bring_window_to_front(win):
    try:
        win.restore()
    except Exception:
        pass

    try:
        win.set_focus()
    except Exception:
        pass

    try:
        handle = int(win.handle)
        ctypes.windll.user32.ShowWindow(handle, 9)
        ctypes.windll.user32.SetForegroundWindow(handle)
    except Exception:
        pass

    time.sleep(0.2)


def refresh_air_window():
    win = find_air_window(timeout=15)
    bring_window_to_front(win)
    return win


def print_current_screen_summary(win):
    log("\n===== 현재 화면 요약 =====")
    log(f"창 제목: {win.window_text()}")
    log(f"창 좌표: {win.rectangle()}")

    log("\n[Button 목록]")
    for i, btn in enumerate(win.descendants(control_type="Button")):
        try:
            log(
                f"Button {i}: "
                f"title='{btn.window_text()}', "
                f"auto_id='{btn.automation_id()}', "
                f"rect={btn.rectangle()}"
            )
        except Exception:
            pass

    log("\n[주요 컨트롤 목록]")
    for i, ctrl in enumerate(win.descendants()):
        try:
            ctype = ctrl.element_info.control_type
            auto_id = ctrl.automation_id()
            title = ctrl.window_text()

            if ctype in ["Edit", "ComboBox", "Custom", "Pane", "Window", "Button"] or auto_id:
                log(
                    f"{i}: type='{ctype}', "
                    f"title='{title}', "
                    f"auto_id='{auto_id}', "
                    f"rect={ctrl.rectangle()}"
                )
        except Exception:
            pass

    log("=========================\n")


def find_child_by_auto_id(parent, auto_id: str):
    candidates = []

    for ctrl in parent.descendants():
        try:
            if ctrl.automation_id() == auto_id and is_real_visible_control(ctrl):
                candidates.append(ctrl)
        except Exception:
            pass

    if not candidates:
        return None

    candidates.sort(key=lambda c: (c.rectangle().top, c.rectangle().left))
    return candidates[0]


def get_child_by_auto_id(parent, auto_id: str, label: str):
    ctrl = find_child_by_auto_id(parent, auto_id)

    if ctrl is None:
        raise RuntimeError(f"{label} 컨트롤을 찾지 못했습니다. auto_id={auto_id}")

    return ctrl


def find_button_by_auto_id(parent, auto_id: str):
    candidates = []

    for btn in parent.descendants(control_type="Button"):
        try:
            if btn.automation_id() == auto_id and is_real_visible_control(btn):
                candidates.append(btn)
        except Exception:
            pass

    if not candidates:
        return None

    candidates.sort(key=lambda c: (c.rectangle().top, c.rectangle().left))
    return candidates[0]


def find_button_by_title(parent, title: str):
    candidates = []

    for btn in parent.descendants(control_type="Button"):
        try:
            if btn.window_text().strip() == title and is_real_visible_control(btn):
                candidates.append(btn)
        except Exception:
            pass

    if not candidates:
        return None

    candidates.sort(key=lambda c: (c.rectangle().top, c.rectangle().left))
    return candidates[0]


def is_inspection_history_screen(win) -> bool:
    try:
        title = win.window_text()
    except Exception:
        title = ""

    result = INSPECTION_WINDOW_KEYWORD in title

    log(f"[CHECK] 검사이력 화면 여부: {result}")
    log(f" - 현재 창 제목: {title}")

    return result


def wait_until_inspection_history_screen(timeout=2):
    start = time.time()

    while time.time() - start < timeout:
        try:
            new_win = refresh_air_window()
            title = new_win.window_text()

            if INSPECTION_WINDOW_KEYWORD in title:
                log(f"[OK] 검사이력 화면 진입 확인: {title}")
                return new_win
        except Exception:
            pass

        time.sleep(0.1)

    return None


def click_inspection_history_menu(win):
    log("[STEP] 검사이력 메뉴 버튼 클릭 시도")

    btn = find_button_by_auto_id(win, INSPECTION_BUTTON_AUTO_ID)

    if btn is None:
        btn = find_button_by_title(win, INSPECTION_BUTTON_TITLE)

    if btn is None:
        print_current_screen_summary(win)
        raise RuntimeError("검사이력 메뉴 버튼을 찾지 못했습니다.")

    log("[OK] 검사이력 버튼 발견")
    log(f"title: {btn.window_text()}")
    log(f"automation_id: {btn.automation_id()}")
    log(f"rectangle: {btn.rectangle()}")

    bring_window_to_front(win)

    try:
        log("[TRY] 검사이력 메뉴 invoke")
        btn.invoke()
        time.sleep(0.2)

        new_win = wait_until_inspection_history_screen(timeout=6)
        if new_win is not None:
            log("[OK] 검사이력 메뉴 invoke 성공")
            return new_win
    except Exception as e:
        log(f"[WARN] 검사이력 메뉴 invoke 실패: {e}")

    click_control_by_real_mouse(btn, "검사이력 메뉴")

    new_win = wait_until_inspection_history_screen(timeout=6)
    if new_win is not None:
        log("[OK] 검사이력 메뉴 실제 클릭 성공")
        return new_win

    print_current_screen_summary(win)
    raise RuntimeError("검사이력 메뉴 진입 실패")


def enter_inspection_history_strict(win):
    log("[STEP] 검사이력 화면 진입 시작")

    if is_inspection_history_screen(win):
        log("[OK] 이미 검사기록조회 화면입니다.")
        return win

    win = click_inspection_history_menu(win)

    if is_inspection_history_screen(win):
        log("[OK] 검사이력 화면 확인 완료")
        return win

    print_current_screen_summary(win)
    raise RuntimeError("검사이력 버튼 클릭 후에도 검사기록조회 화면이 확인되지 않았습니다.")


def get_inspection_form(win):
    """
    검사기록조회 폼(frmSetupList_1)을 찾습니다.

    속도 개선:
    - 한 번 찾은 form은 CONTROL_CACHE에 저장합니다.
    - 다음 부서 변경 때는 전체 descendants() 탐색을 하지 않고 캐시를 먼저 사용합니다.
    """
    cached = CONTROL_CACHE.get("inspection_form")

    if cached is not None:
        try:
            if cached.automation_id() == INSPECTION_FORM_AUTO_ID and is_real_visible_control(cached):
                return cached
        except Exception:
            CONTROL_CACHE.pop("inspection_form", None)

    for ctrl in win.descendants():
        try:
            if ctrl.automation_id() == INSPECTION_FORM_AUTO_ID:
                CONTROL_CACHE["inspection_form"] = ctrl
                log("[OK] 검사기록조회 폼 발견 및 캐시 저장: auto_id=frmSetupList_1")
                return ctrl
        except Exception:
            pass

    for ctrl in win.descendants():
        try:
            title = ctrl.window_text().strip()
            ctype = ctrl.element_info.control_type

            if title == INSPECTION_FORM_TITLE and ctype in ["Window", "Pane", "Custom"]:
                CONTROL_CACHE["inspection_form"] = ctrl
                log("[OK] 검사기록조회 폼 발견 및 캐시 저장: title=검사기록조회")
                return ctrl
        except Exception:
            pass

    print_current_screen_summary(win)
    raise RuntimeError("검사기록조회 폼(frmSetupList_1)을 찾지 못했습니다.")


def get_period_group(form):
    candidates = []

    for ctrl in form.descendants():
        try:
            if ctrl.automation_id() == PERIOD_GROUP_AUTO_ID and is_real_visible_control(ctrl):
                candidates.append(ctrl)
        except Exception:
            pass

    if not candidates:
        raise RuntimeError(f"기간설정 그룹박스({PERIOD_GROUP_AUTO_ID})를 찾지 못했습니다.")

    candidates.sort(key=lambda c: (c.rectangle().top, c.rectangle().left))
    group = candidates[0]

    log("[OK] 기간설정 그룹박스 발견")
    log(f" - title: {group.window_text()}")
    log(f" - auto_id: {group.automation_id()}")
    log(f" - rect: {group.rectangle()}")

    return group


def get_period_child_by_auto_id(form, auto_id: str, label: str):
    period_group = get_period_group(form)
    candidates = []

    for ctrl in period_group.descendants():
        try:
            if ctrl.automation_id() == auto_id and is_real_visible_control(ctrl):
                candidates.append(ctrl)
        except Exception:
            pass

    if not candidates:
        raise RuntimeError(f"{label} 컨트롤을 기간설정 그룹박스 안에서 찾지 못했습니다. auto_id={auto_id}")

    candidates.sort(key=lambda c: (c.rectangle().top, c.rectangle().left))
    ctrl = candidates[0]

    log(f"[OK] {label} 기간설정 내부 컨트롤 발견")
    log(f" - title: {ctrl.window_text()}")
    log(f" - auto_id: {ctrl.automation_id()}")
    log(f" - class: {ctrl.class_name()}")
    log(f" - rect: {ctrl.rectangle()}")

    return ctrl


def get_condition_panel(form):
    cached = CONTROL_CACHE.get("condition_panel")
    if cached is not None and is_real_visible_control(cached):
        return cached

    candidates = []

    for ctrl in form.descendants():
        try:
            if ctrl.automation_id() == CONDITION_PANEL_AUTO_ID and is_real_visible_control(ctrl):
                candidates.append(ctrl)
        except Exception:
            pass

    if not candidates:
        raise RuntimeError(f"조건 영역 패널({CONDITION_PANEL_AUTO_ID})을 찾지 못했습니다.")

    candidates.sort(key=lambda c: (c.rectangle().top, c.rectangle().left))
    panel = candidates[0]
    CONTROL_CACHE["condition_panel"] = panel

    log("[OK] 조건 영역 panel1 발견")
    log(f" - rect: {panel.rectangle()}")

    return panel


# =========================================================
# 6. ComboBox 제어
# =========================================================

def send_message(hwnd, msg, wparam=0, lparam=0):
    return _SendMessageW(
        wintypes.HWND(int(hwnd)),
        wintypes.UINT(msg),
        WPARAM_T(int(wparam)),
        LPARAM_T(int(lparam)),
    )


def get_control_hwnd(ctrl) -> int:
    try:
        hwnd = int(ctrl.handle)
        if hwnd:
            return hwnd
    except Exception:
        pass

    try:
        hwnd = int(ctrl.element_info.handle)
        if hwnd:
            return hwnd
    except Exception:
        pass

    return 0


def get_combobox_items_by_win32(combo):
    hwnd = get_control_hwnd(combo)
    if not hwnd:
        return []

    count = send_message(hwnd, CB_GETCOUNT, 0, 0)

    if count < 0 or count > 1000:
        return []

    items = []

    for i in range(int(count)):
        text_len = send_message(hwnd, CB_GETLBTEXTLEN, i, 0)

        if text_len < 0 or text_len > 1000:
            items.append("")
            continue

        buf = ctypes.create_unicode_buffer(int(text_len) + 1)
        send_message(hwnd, CB_GETLBTEXT, i, ctypes.addressof(buf))
        items.append(buf.value)

    return items


def get_combobox_cur_sel(combo) -> int:
    hwnd = get_control_hwnd(combo)
    if not hwnd:
        return -1

    return int(send_message(hwnd, CB_GETCURSEL, 0, 0))


def notify_combobox_selection_changed(combo):
    hwnd = get_control_hwnd(combo)

    if not hwnd:
        return

    user32 = ctypes.windll.user32
    parent_hwnd = user32.GetParent(wintypes.HWND(hwnd))
    ctrl_id = user32.GetDlgCtrlID(wintypes.HWND(hwnd))

    if not parent_hwnd:
        return

    for notify_code in (CBN_SELCHANGE, CBN_SELENDOK):
        wparam = (notify_code << 16) | (ctrl_id & 0xFFFF)
        send_message(parent_hwnd, WM_COMMAND, wparam, hwnd)


def set_combobox_index_win32(combo, index: int, label: str) -> bool:
    hwnd = get_control_hwnd(combo)
    if not hwnd:
        return False

    log(f"[INFO] {label} CB_SETCURSEL 인덱스 선택 시도: index={index}")

    result = send_message(hwnd, CB_SETCURSEL, index, 0)
    
    notify_combobox_selection_changed(combo)
    

    cur_sel = get_combobox_cur_sel(combo)

    log(f"[INFO] {label} 선택 후 cur_sel={cur_sel}, target_index={index}, result={result}")

    return cur_sel == index


def select_combobox_index_by_keyboard(combo, index: int, label: str):
    rect = combo.rectangle()
    x = rect.right - 12
    y = int((rect.top + rect.bottom) / 2)

    log(f"[INFO] {label} 드롭다운 실제 클릭 후 키보드 인덱스 선택")
    log(f" - target_index={index}")
    log(f" - click point: ({x}, {y}), rect={rect}")

    real_mouse_click(x, y)
    

    send_keys("{HOME}")
    

    for _ in range(index):
        send_keys("{DOWN}")
        

    send_keys("{ENTER}")
    time.sleep(0.3)

def get_department_combo_fast(form):
    """
    부서명 ComboBox(cboCostCenter)를 빠르게 찾습니다.

    속도 개선:
    - 처음 1회만 form.descendants(control_type="ComboBox")로 탐색합니다.
    - 이후에는 CONTROL_CACHE["department_combo"]를 재사용합니다.
    - 기존처럼 ComboBox 항목 전체를 매번 읽지 않습니다.
    """
    cached = CONTROL_CACHE.get("department_combo")

    if cached is not None:
        try:
            if cached.automation_id() == DEPARTMENT_COMBO_AUTO_ID and is_real_visible_control(cached):
                return cached
        except Exception:
            CONTROL_CACHE.pop("department_combo", None)

    candidates = []

    for ctrl in form.descendants(control_type="ComboBox"):
        try:
            if ctrl.automation_id() == DEPARTMENT_COMBO_AUTO_ID and is_real_visible_control(ctrl):
                candidates.append(ctrl)
        except Exception:
            pass

    if not candidates:
        raise RuntimeError("부서명 ComboBox(cboCostCenter)를 찾지 못했습니다.")

    candidates.sort(key=lambda c: (c.rectangle().top, c.rectangle().left))
    combo = candidates[0]

    CONTROL_CACHE["department_combo"] = combo

    log("[OK] 부서명 ComboBox 발견 및 캐시 저장")
    log(f" - auto_id: {combo.automation_id()}")
    log(f" - rect: {combo.rectangle()}")

    return combo

def safe_select_department_by_index(form, department: str, target_index: int):
    """
    부서명 ComboBox를 인덱스로 빠르게 선택합니다.

    기존 방식:
    - 매번 모든 ComboBox 탐색
    - 매번 ComboBox item 전체 읽기
    - item 텍스트와 department 비교

    개선 방식:
    - 부서 ComboBox 캐시 사용
    - 현재 선택 index가 같으면 바로 skip
    - DEPARTMENT_JOBS의 combo_index를 신뢰하고 바로 CB_SETCURSEL 실행
    """
    combo = get_department_combo_fast(form)

    cur_sel = get_combobox_cur_sel(combo)

    if cur_sel == target_index:
        log(f"[SKIP] 부서명 이미 선택됨: {department} -> index {target_index}")
        return

    if set_combobox_index_win32(combo, target_index, f"부서명 {department}"):
        log(f"[OK] 부서명 선택 완료: {department} -> index {target_index}")
        return

    log(f"[WARN] Win32 인덱스 선택 실패, 키보드 방식으로 재시도: {department}")
    select_combobox_index_by_keyboard(combo, target_index, f"부서명 {department}")
    log(f"[OK] 부서명 키보드 선택 완료: {department} -> index {target_index}")


def safe_select_combobox_by_text(combo, value: str, label: str):
    value = str(value).strip()

    if not value:
        log(f"[SKIP] {label}: 값 없음")
        return

    items = get_combobox_items_by_win32(combo)

    log(f"[INFO] {label} ComboBox 항목: {items}")

    target_index = None

    for i, item in enumerate(items):
        if normalize_text(item) == normalize_text(value):
            target_index = i
            break

    if target_index is not None:
        if set_combobox_index_win32(combo, target_index, label):
            log(f"[OK] {label} 선택 완료: {value}")
            return

        select_combobox_index_by_keyboard(combo, target_index, label)
        log(f"[OK] {label} 키보드 선택 완료: {value}")
        return

    try:
        combo.set_focus()
        time.sleep(0.1)
        combo.click_input()
        time.sleep(0.1)

        send_keys("^a")
        time.sleep(0.05)
        send_keys("{BACKSPACE}")
        time.sleep(0.05)

        set_clipboard_text(value)
        send_keys("^v")
        time.sleep(0.1)
        send_keys("{ENTER}")
        time.sleep(0.2)

        log(f"[OK] {label} 텍스트 입력 방식 완료: {value}")
    except Exception as e:
        raise RuntimeError(f"{label} ComboBox 선택 실패: {e}")


def safe_select_period_time(form, auto_id: str, value: str, label: str):
    combo = get_period_child_by_auto_id(form, auto_id, label)
    safe_select_combobox_by_text(combo, value, label)


# =========================================================
# 7. 날짜 / 조건 입력
# =========================================================

def safe_set_text(control, value: str, field_name: str):
    if value is None or str(value).strip() == "":
        log(f"[SKIP] {field_name}: 입력값 없음")
        return

    value = str(value).strip()

    try:
        control.set_focus()
        time.sleep(0.1)

        try:
            control.set_edit_text(value)
            log(f"[OK] {field_name} set_edit_text 입력 완료: {value}")
            return
        except Exception as e:
            log(f"[WARN] {field_name} set_edit_text 실패, 클립보드 입력으로 재시도: {e}")

        control.click_input()
        time.sleep(0.1)

        send_keys("^a")
        time.sleep(0.05)
        send_keys("{BACKSPACE}")
        time.sleep(0.05)

        set_clipboard_text(value)
        send_keys("^v")
        time.sleep(0.1)
        send_keys("{TAB}")
        time.sleep(0.1)

        log(f"[OK] {field_name} 클립보드 입력 완료: {value}")

    except Exception as e:
        raise RuntimeError(f"{field_name} 입력 실패: {e}")


def safe_set_datetime_picker(form, auto_id: str, value: str, label: str):
    if value is None or str(value).strip() == "":
        log(f"[SKIP] {label}: 입력값 없음")
        return

    value = str(value).strip()
    year, month, day = parse_date_ymd(value)

    picker = get_period_child_by_auto_id(form, auto_id, label)

    log(f"[INFO] {label} DateTimePicker 입력 시도: {value}")
    log(f"[INFO] year={year}, month={month}, day={day}")

    rect = picker.rectangle()

    x = rect.left + 20
    y = int((rect.top + rect.bottom) / 2)

    log(f"[INFO] {label} 실제 클릭 좌표: x={x}, y={y}, rect={rect}")

    real_mouse_click(x, y)
    time.sleep(0.15)

    send_keys("{HOME}")
    time.sleep(0.05)

    send_keys(year)
    time.sleep(0.05)

    send_keys("{RIGHT}")
    time.sleep(0.05)
    send_keys(month)
    time.sleep(0.05)

    send_keys("{RIGHT}")
    time.sleep(0.05)
    send_keys(day)
    time.sleep(0.05)

    send_keys("{TAB}")
    time.sleep(0.15)

    log(f"[OK] {label} 입력 완료: {value}")


def input_conditions(
    win,
    department: str,
    department_index: int,
    start_date: str,
    end_date: str,
    start_time: str,
    end_time: str,
    product_no: str = "",
    machine_no: str = "",
    lot_no: str = "",
    barcode: str = "",
    tool_no: str = "",
    po_no: str = "",
    disposition: str = "",
    plant: str = "",
    set_period: bool = True,
):
    log("[STEP] 검사이력 조건 입력 시작")

    form = get_inspection_form(win)

    text_values = {
        "제품번호": product_no,
        "기계번호": machine_no,
        "Lot No": lot_no,
        "Barcode": barcode,
        "툴번호": tool_no,
        "PO No": po_no,
        "판정": disposition,
    }

    for label, value in text_values.items():
        auto_id = FIELD_AUTO_IDS[label]
        if value:
            ctrl = get_child_by_auto_id(form, auto_id, label)
            safe_set_text(ctrl, value, label)

    if plant:
        plant_combo = get_child_by_auto_id(form, PLANT_COMBO_AUTO_ID, "플랜트")
        safe_select_combobox_by_text(plant_combo, plant, "플랜트")

    # 부서명은 매번 변경해야 함
    safe_select_department_by_index(form, department, department_index)

    # 기간/시간은 첫 번째 부서에서만 설정
    if set_period:
        log("[STEP] 첫 번째 부서이므로 기간/시간 조건을 설정합니다.")

        safe_set_datetime_picker(form, START_DATE_PICKER_AUTO_ID, start_date, "시작일")
        safe_set_datetime_picker(form, END_DATE_PICKER_AUTO_ID, end_date, "종료일")

        safe_select_period_time(form, START_TIME_COMBO_AUTO_ID, start_time, "시작시간")
        safe_select_period_time(form, END_TIME_COMBO_AUTO_ID, end_time, "종료시간")

    else:
        log("[SKIP] 기간/시간 조건은 이미 설정되어 있으므로 생략합니다.")

    log("[OK] 검사이력 조건 입력 완료")
    return form


# =========================================================
# 8. 조회 버튼 / 엑셀 버튼
# =========================================================

def get_search_icon_button(form):
    cached = CONTROL_CACHE.get("search_button")
    if cached is not None and is_real_visible_control(cached):
        return cached

    panel = get_condition_panel(form)

    candidates = []

    for btn in panel.descendants(control_type="Button"):
        try:
            if btn.automation_id() == INSPECTION_SEARCH_BUTTON_AUTO_ID and is_real_visible_control(btn):
                candidates.append(btn)
        except Exception:
            pass

    if not candidates:
        raise RuntimeError("panel1 내부에서 기간설정 옆 조회 버튼(button1)을 찾지 못했습니다.")

    candidates.sort(key=lambda c: (c.rectangle().top, c.rectangle().left))
    btn = candidates[0]
    CONTROL_CACHE["search_button"] = btn

    log("[OK] 기간설정 옆 조회 아이콘 발견")
    log(f" - title: {btn.window_text()}")
    log(f" - auto_id: {btn.automation_id()}")
    log(f" - rect: {btn.rectangle()}")

    return btn


def get_excel_button(form):
    cached = CONTROL_CACHE.get("excel_button")
    if cached is not None and is_real_visible_control(cached):
        return cached

    panel = get_condition_panel(form)

    candidates = []

    for btn in panel.descendants(control_type="Button"):
        try:
            if btn.automation_id() == EXCEL_BUTTON_AUTO_ID and is_real_visible_control(btn):
                candidates.append(btn)
        except Exception:
            pass

    if not candidates:
        raise RuntimeError("panel1 내부에서 엑셀 내보내기 버튼(cmdExcel)을 찾지 못했습니다.")

    candidates.sort(key=lambda c: (c.rectangle().top, c.rectangle().left))
    btn = candidates[0]
    CONTROL_CACHE["excel_button"] = btn

    log("[OK] 기간설정 옆 엑셀 내보내기 버튼 발견")
    log(f" - title: {btn.window_text()}")
    log(f" - auto_id: {btn.automation_id()}")
    log(f" - rect: {btn.rectangle()}")

    return btn


def click_search(win, form=None):
    """
    조회 클릭 속도 개선 버전.
    - 매번 refresh_air_window()로 AIR를 다시 찾지 않음
    - 이미 잡아둔 form을 재사용
    - 조회 완료 팝업은 foreground 팝업 우선 Enter 처리
    """
    if form is None:
        form = get_inspection_form(win)

    search_button = get_search_icon_button(form)
    click_control_by_real_mouse(search_button, "기간설정 옆 조회 아이콘")

    # 고정 대기 최소화
    log(f"[INFO] 조회 후 최소 대기: {SEARCH_WAIT_SECONDS}초")
    time.sleep(SEARCH_WAIT_SECONDS)

    # 조회 완료/알림 팝업이 뜨면 즉시 닫음
    handle_air_message_popup_fast(
        label="조회 완료",
        keywords=["조회", "검색", "완료", "조회가 완료", "조회되었습니다"],
        timeout=FAST_POPUP_TIMEOUT_SECONDS,
    )


# =========================================================
# 9. 저장 대화상자 처리 - 기존 성공 방식 복원 + 부서별 저장경로 적용
# =========================================================

def debug_dump_windows_for_save():
    """
    저장창 감지 실패 시 Python이 실제로 보고 있는 창 목록을 출력합니다.
    """
    log("\n===== 저장창 디버그: 현재 감지 가능한 창 목록 =====")

    for backend in ["uia", "win32"]:
        log(f"\n[{backend.upper()} windows]")
        try:
            for w in Desktop(backend=backend).windows():
                try:
                    title = w.window_text().strip()
                    rect = w.rectangle()
                    handle = w.handle

                    if not title:
                        continue

                    try:
                        class_name = w.class_name()
                    except Exception:
                        class_name = ""

                    log(f"hwnd={handle}, class='{class_name}', title='{title}', rect={rect}")
                except Exception:
                    pass
        except Exception as e:
            log(f"[WARN] {backend} 창 목록 출력 실패: {e}")

    log("=================================================\n")


def is_air_main_window_title(title: str) -> bool:
    title = str(title)

    return (
        "CCS AIR System" in title
        or "OPC AIR System" in title
        or "검사기록조회" in title
    )


def find_save_dialog(timeout=8):
    """
    Windows '다른 이름으로 저장' 창을 찾습니다.

    핵심:
    1. UIA와 Win32 backend를 모두 탐색합니다.
    2. AIR 메인창은 저장창 후보에서 제외합니다.
    3. 저장창을 못 찾았다고 해서 foreground 창을 저장창으로 간주하지 않습니다.
       이 부분이 기존 오류의 핵심 원인이었습니다.
    """
    title_patterns = [
        r".*다른 이름으로 저장.*",
        r".*Save As.*",
        r".*Save as.*",
        r".*파일 저장.*",
        r".*저장.*",
    ]

    start = time.time()

    while time.time() - start < timeout:
        for backend in ["uia", "win32"]:
            try:
                for w in Desktop(backend=backend).windows():
                    try:
                        title = w.window_text().strip()

                        if not title:
                            continue

                        if is_air_main_window_title(title):
                            continue

                        for pattern in title_patterns:
                            if re.search(pattern, title, re.IGNORECASE):
                                log(f"[OK] 저장 대화상자 감지: backend={backend}, title='{title}'")
                                log(f" - rect: {w.rectangle()}")
                                return w

                    except Exception:
                        pass
            except Exception:
                pass

        time.sleep(0.3)

    log("[WARN] 저장 대화상자를 찾지 못했습니다.")
    debug_dump_windows_for_save()
    return None


def bring_dialog_to_front(dialog):
    """
    저장 대화상자를 전면으로 가져옵니다.
    AIR 메인창이 아니라, 이미 찾은 저장 dialog 객체에 대해서만 실행합니다.
    """
    try:
        dialog.restore()
    except Exception:
        pass

    try:
        dialog.set_focus()
    except Exception:
        pass

    try:
        hwnd = int(dialog.handle)
        ctypes.windll.user32.ShowWindow(hwnd, 9)
        ctypes.windll.user32.SetForegroundWindow(hwnd)
    except Exception:
        pass

    time.sleep(0.3)


def find_filename_edit_in_save_dialog(dialog):
    """
    Windows Save As 창에서 파일 이름 입력 Edit를 찾습니다.

    우선순위:
    1. auto_id == '1001'인 Edit
    2. 아래쪽에 있는 넓은 Edit
    3. Win32 class_name='Edit' fallback
    """
    candidates = []

    # 1차: UIA 방식
    try:
        for edit in dialog.descendants(control_type="Edit"):
            try:
                if is_real_visible_control(edit):
                    candidates.append(edit)
            except Exception:
                pass
    except Exception:
        pass

    # 2차: Win32 방식
    if not candidates:
        try:
            for edit in dialog.descendants(class_name="Edit"):
                try:
                    if is_real_visible_control(edit):
                        candidates.append(edit)
                except Exception:
                    pass
        except Exception:
            pass

    if not candidates:
        log("[WARN] 저장창 안에서 Edit 컨트롤을 찾지 못했습니다.")
        return None

    log("\n===== 저장창 Edit 후보 =====")
    for i, edit in enumerate(candidates):
        try:
            log(
                f"[{i}] title='{edit.window_text()}', "
                f"auto_id='{edit.automation_id()}', "
                f"class='{edit.class_name()}', "
                f"rect={edit.rectangle()}"
            )
        except Exception:
            pass
    log("===========================\n")

    # 표준 Windows Save As 파일명 입력칸은 auto_id가 1001인 경우가 많음
    for edit in candidates:
        try:
            if edit.automation_id() == "1001":
                log("[OK] 파일명 입력칸 선택: auto_id=1001")
                return edit
        except Exception:
            pass

    # fallback: 넓고 아래쪽에 있는 Edit를 파일명 입력칸으로 판단
    candidates.sort(
        key=lambda e: (
            -(e.rectangle().width()),
            -e.rectangle().top,
        )
    )

    log("[OK] 파일명 입력칸 fallback 선택")
    log(f" - rect: {candidates[0].rectangle()}")

    return candidates[0]


def set_save_filename(dialog, save_path: str):
    """
    저장창의 파일 이름 입력칸에 전체 저장 경로를 입력합니다.
    기존 성공 코드처럼 Edit 컨트롤을 직접 잡아서 입력합니다.
    """
    save_path = str(Path(save_path).resolve())

    edit = find_filename_edit_in_save_dialog(dialog)

    if edit is not None:
        try:
            edit.set_focus()
            time.sleep(0.2)

            log("[STEP] 파일명 입력칸에 set_edit_text로 전체 경로 입력")
            edit.set_edit_text(save_path)
            time.sleep(0.3)
            return True

        except Exception as e:
            log(f"[WARN] set_edit_text 실패, 클립보드 입력으로 재시도: {e}")

            try:
                edit.click_input()
                time.sleep(0.2)

                send_keys("^a")
                time.sleep(0.1)

                set_clipboard_text(save_path)
                send_keys("^v")
                time.sleep(0.3)

                return True
            except Exception as e2:
                log(f"[WARN] Edit 클립보드 입력 실패: {e2}")

    # 최후 fallback: 저장창에 포커스를 두고 Ctrl+A / 붙여넣기
    try:
        log("[WARN] 파일명 Edit를 직접 제어하지 못해 dialog 포커스 기준으로 붙여넣기 시도")
        dialog.set_focus()
        time.sleep(0.2)

        send_keys("^a")
        time.sleep(0.1)

        set_clipboard_text(save_path)
        send_keys("^v")
        time.sleep(0.3)

        return True
    except Exception as e:
        log(f"[ERROR] 저장 경로 입력 최종 실패: {e}")
        return False


def find_save_button_in_dialog(dialog):
    """
    저장창 내부의 저장 버튼을 찾습니다.
    """
    save_titles = [
        "저장",
        "저장(&S)",
        "&저장",
        "Save",
        "&Save",
    ]

    candidates = []

    # UIA Button 후보
    try:
        for btn in dialog.descendants(control_type="Button"):
            try:
                if not is_real_visible_control(btn):
                    continue

                candidates.append(btn)
            except Exception:
                pass
    except Exception:
        pass

    # Win32 Button 후보
    if not candidates:
        try:
            for btn in dialog.descendants(class_name="Button"):
                try:
                    if not is_real_visible_control(btn):
                        continue

                    candidates.append(btn)
                except Exception:
                    pass
        except Exception:
            pass

    log("\n===== 저장창 Button 후보 =====")
    for i, btn in enumerate(candidates):
        try:
            log(
                f"[{i}] title='{btn.window_text()}', "
                f"auto_id='{btn.automation_id()}', "
                f"class='{btn.class_name()}', "
                f"rect={btn.rectangle()}"
            )
        except Exception:
            pass
    log("============================\n")

    for title in save_titles:
        for btn in candidates:
            try:
                if btn.window_text().strip() == title:
                    log(f"[OK] 저장 버튼 발견: '{title}'")
                    return btn
            except Exception:
                pass

    return None


def click_save_button_in_dialog(dialog):
    """
    저장창 내부 저장 버튼을 클릭합니다.
    실패 시 Alt+S, Enter 순서로 fallback 합니다.

    주의:
    Alt+S를 전역으로 바로 보내지 않고,
    먼저 실제 저장 dialog를 찾은 다음 dialog에 포커스를 둔 상태에서만 사용합니다.
    """
    btn = find_save_button_in_dialog(dialog)

    if btn is not None:
        try:
            click_control_by_real_mouse(btn, "저장 대화상자 저장 버튼")
            time.sleep(1.0)
            return True
        except Exception as e:
            log(f"[WARN] 저장 버튼 실제 클릭 실패: {e}")

    try:
        log("[STEP] 저장 버튼 클릭 실패 fallback: 저장창 포커스 후 Alt+S")
        dialog.set_focus()
        time.sleep(0.2)
        send_keys("%s")
        time.sleep(1.0)
        return True
    except Exception as e:
        log(f"[WARN] Alt+S 저장 실패: {e}")

    try:
        log("[STEP] Alt+S 실패 fallback: Enter")
        dialog.set_focus()
        time.sleep(0.2)
        send_keys("{ENTER}")
        time.sleep(1.0)
        return True
    except Exception as e:
        log(f"[ERROR] Enter 저장 실패: {e}")

    return False


def handle_overwrite_confirm(timeout=5):
    """
    같은 파일명이 있을 때 뜨는 덮어쓰기 확인창에서 예/Yes/확인을 누릅니다.
    """
    start = time.time()

    yes_titles = [
        "예",
        "예(&Y)",
        "&Yes",
        "Yes",
        "확인",
        "OK",
    ]

    while time.time() - start < timeout:
        for backend in ["uia", "win32"]:
            try:
                for w in Desktop(backend=backend).windows():
                    try:
                        title = w.window_text().strip()

                        if not title:
                            continue

                        if (
                            "확인" in title
                            or "Confirm" in title
                            or "덮어쓰기" in title
                            or "대체" in title
                            or "저장" in title
                            or "Save" in title
                        ):
                            for btn_title in yes_titles:
                                btn = None

                                try:
                                    btn = find_button_by_title(w, btn_title)
                                except Exception:
                                    pass

                                if btn is not None:
                                    click_control_by_real_mouse(btn, f"덮어쓰기 확인 버튼 '{btn_title}'")
                                    time.sleep(0.3)
                                    return True

                            try:
                                w.set_focus()
                                time.sleep(0.2)
                                send_keys("{ENTER}")
                                time.sleep(0.7)
                                return True
                            except Exception:
                                pass

                    except Exception:
                        pass
            except Exception:
                pass

        time.sleep(0.3)

    return False



def _wrap_window_by_handle(hwnd):
    """foreground hwnd를 UIA/Win32 wrapper로 감싸서 반환"""
    for backend in ["uia", "win32"]:
        try:
            return Desktop(backend=backend).window(handle=int(hwnd))
        except Exception:
            pass
    return None


def _window_has_keywords_shallow(win, keywords):
    try:
        title = win.window_text().strip()
        if any(k in title for k in keywords):
            return True
    except Exception:
        pass

    # 전체 데스크톱이 아니라 해당 작은 팝업 내부만 검사하므로 빠름
    try:
        for ctrl in win.descendants():
            try:
                text = ctrl.window_text().strip()
                if text and any(k in text for k in keywords):
                    return True
            except Exception:
                pass
    except Exception:
        pass

    return False


def _looks_like_small_air_popup(win):
    try:
        title = win.window_text().strip()
        rect = win.rectangle()
        is_small = rect.width() < 800 and rect.height() < 500
        title_ok = any(k in title for k in ["검사기록조회", "검사이력", "AIR", "알림", "확인", "Information"])
        return is_small and title_ok
    except Exception:
        return False


def handle_air_message_popup_fast(label: str, keywords, timeout: float = 1.2) -> bool:
    """
    조회 완료/저장 완료 같은 AIR 모달 팝업을 빠르게 닫습니다.
    기존 handle_air_save_complete_popup은 전체 창을 UIA/Win32로 길게 스캔해서 느렸으므로,
    foreground 작은 팝업을 먼저 확인하고 Enter로 닫습니다.
    """
    start = time.time()

    while time.time() - start < timeout:
        try:
            hwnd = ctypes.windll.user32.GetForegroundWindow()
            win = _wrap_window_by_handle(hwnd)

            if win is not None:
                has_keyword = _window_has_keywords_shallow(win, keywords)
                looks_popup = _looks_like_small_air_popup(win)

                if has_keyword or looks_popup:
                    try:
                        title = win.window_text().strip()
                    except Exception:
                        title = ""

                    log(f"[OK] {label} 팝업 빠른 감지: title='{title}'")

                    try:
                        win.set_focus()
                    except Exception:
                        pass

                    send_keys("{ENTER}")
                    time.sleep(0.2)
                    return True
        except Exception:
            pass

        time.sleep(0.1)

    log(f"[INFO] {label} 팝업 빠른 감지 없음")
    return False

def handle_save_dialog_if_present(save_path: str) -> bool:
    """
    저장 대화상자 처리.

    기존 성공 방식 기준:
    1. 저장창을 UIA/Win32 둘 다에서 실제로 찾음
    2. AIR 메인창은 저장창 후보에서 제외
    3. foreground fallback 사용하지 않음
    4. 파일명 Edit를 직접 찾아 전체 경로 입력
    5. 저장창 내부 저장 버튼 클릭
    6. 실패 시에만 Alt+S/Enter fallback
    """
    if not HANDLE_SAVE_DIALOG:
        return False

    save_path = str(Path(save_path).resolve())

    log("[INFO] 저장할 전체 경로:")
    log(f"       {save_path}")

    os.makedirs(os.path.dirname(save_path), exist_ok=True)

    dialog = find_save_dialog(timeout=SAVE_DIALOG_WAIT_SECONDS)

    if dialog is None:
        log("[WARN] 저장 대화상자가 감지되지 않아 저장 처리를 건너뜁니다.")
        log("[WARN] 이 경우 Python 권한, AIR 권한, 저장창 제목, UIA/Win32 인식 여부를 확인해야 합니다.")
        return False

    bring_dialog_to_front(dialog)

    if not set_save_filename(dialog, save_path):
        raise RuntimeError("저장 대화상자 파일명 입력칸에 저장 경로를 입력하지 못했습니다.")

    if not click_save_button_in_dialog(dialog):
        raise RuntimeError("저장 대화상자에서 저장 버튼 클릭 또는 Alt+S 저장에 실패했습니다.")

    handle_overwrite_confirm(timeout=5)

    if wait_until_file_exists(save_path, timeout=25):
        log("[OK] 엑셀 파일 저장 완료:")
        log(f"     {save_path}")
        return True

    log("[INFO] 저장 명령은 실행했지만 지정 경로에서 파일 생성을 아직 확인하지 못했습니다.")
    log(f"       예상 경로: {save_path}")
    return True


def click_excel_export(win, export_save_path: str, form=None) -> str:
    """
    엑셀 내보내기 속도 개선 버전.
    - refresh_air_window()/get_inspection_form() 중복 호출 제거
    - 저장창 대기 시간을 짧게 유지
    - 파일 저장 확인 후 저장 완료 팝업을 foreground 기준으로 즉시 Enter 처리
    """
    before_snapshot = snapshot_excel_files()

    if form is None:
        form = get_inspection_form(win)

    excel_button = get_excel_button(form)
    click_control_by_real_mouse(excel_button, "기간설정 옆 엑셀 내보내기 버튼")

    # 저장창이 뜰 최소 시간만 확보
    time.sleep(POST_EXPORT_DIALOG_WAIT_SECONDS)

    handled = handle_save_dialog_if_present(export_save_path)

    if wait_until_file_exists(export_save_path, timeout=25):
        log(f"[OK] 엑셀 내보내기 파일 확인: {export_save_path}")

        # 빠른 방식으로 먼저 닫고, 실패할 때만 기존 긴 스캔 fallback
        if not handle_air_message_popup_fast(
            label="저장 완료",
            keywords=["저장", "완료", "저장이 완료", "저장되었습니다"],
            timeout=FAST_POPUP_TIMEOUT_SECONDS,
        ):
            handle_air_save_complete_popup(timeout=2)

        return export_save_path

    new_file = wait_for_new_excel_file(before_snapshot, timeout=25)

    if new_file:
        final_path = copy_or_rename_export_file(new_file, export_save_path)
        log(f"[OK] 자동 생성된 엑셀 파일 확인: {new_file}")
        log(f"[OK] 테스트용 파일 경로: {final_path}")

        if not handle_air_message_popup_fast(
            label="저장 완료",
            keywords=["저장", "완료", "저장이 완료", "저장되었습니다"],
            timeout=FAST_POPUP_TIMEOUT_SECONDS,
        ):
            handle_air_save_complete_popup(timeout=2)

        return final_path

    if handled:
        handle_air_message_popup_fast(
            label="저장 완료",
            keywords=["저장", "완료", "저장이 완료", "저장되었습니다"],
            timeout=0.8,
        )
        raise RuntimeError(
            "저장 명령은 실행했지만 파일 생성 확인에 실패했습니다.\n"
            f"예상 경로: {export_save_path}"
        )

    handle_air_message_popup_fast(
        label="저장 완료",
        keywords=["저장", "완료", "저장이 완료", "저장되었습니다"],
        timeout=0.8,
    )

    raise RuntimeError(
        "엑셀 내보내기 파일을 확인하지 못했습니다.\n"
        f"예상 경로: {export_save_path}\n"
        "저장 대화상자 감지 또는 자동 저장 확인이 실패했습니다."
    )


def window_contains_any_text(win, keywords):
    """
    창 내부 텍스트에 특정 문구가 있는지 확인.
    예: '저장이 완료되었습니다'
    """
    try:
        title = win.window_text().strip()
        for keyword in keywords:
            if keyword in title:
                return True
    except Exception:
        pass

    try:
        for ctrl in win.descendants():
            try:
                text = ctrl.window_text().strip()
                if not text:
                    continue

                for keyword in keywords:
                    if keyword in text:
                        return True
            except Exception:
                pass
    except Exception:
        pass

    return False


def find_ok_button_in_dialog(dialog):
    """
    팝업 내부의 OK / 확인 버튼을 찾는다.
    """
    ok_keywords = [
        "OK",
        "Ok",
        "ok",
        "확인",
        "예",
        "Yes",
    ]

    # UIA Button 탐색
    try:
        for btn in dialog.descendants(control_type="Button"):
            try:
                if not is_real_visible_control(btn):
                    continue

                text = btn.window_text().strip()

                if any(k.lower() == text.lower() for k in ok_keywords):
                    return btn

                if any(k.lower() in text.lower() for k in ok_keywords):
                    return btn
            except Exception:
                pass
    except Exception:
        pass

    # Win32 Button 탐색 fallback
    try:
        for btn in dialog.descendants(class_name="Button"):
            try:
                if not is_real_visible_control(btn):
                    continue

                text = btn.window_text().strip()

                if any(k.lower() == text.lower() for k in ok_keywords):
                    return btn

                if any(k.lower() in text.lower() for k in ok_keywords):
                    return btn
            except Exception:
                pass
    except Exception:
        pass

    return None


def handle_air_save_complete_popup(timeout=10):
    """
    AIR 저장 완료 팝업 처리.

    저장 후 AIR에서 '저장이 완료되었습니다' 팝업이 뜨고,
    OK 버튼을 눌러야 다음 부서 조회가 가능하므로 반드시 처리해야 함.
    """
    complete_keywords = [
        "저장이 완료되었습니다",
        "저장 완료",
        "저장되었습니다",
        "완료되었습니다",
        "저장이 완료",
    ]

    popup_title_keywords = [
        "검사기록조회",
        "검사이력",
        "AIR",
        "알림",
        "확인",
        "Information",
    ]

    start = time.time()

    while time.time() - start < timeout:
        for backend in ["uia", "win32"]:
            try:
                for w in Desktop(backend=backend).windows():
                    try:
                        title = w.window_text().strip()
                        rect = w.rectangle()

                        if not title:
                            continue

                        # 1순위: 창 내부에 저장 완료 문구가 있는 경우
                        has_complete_text = window_contains_any_text(w, complete_keywords)

                        # 2순위: 검사기록조회 제목의 작은 팝업인 경우
                        # AIR 메인 화면보다 팝업은 보통 훨씬 작음
                        is_small_popup = rect.width() < 700 and rect.height() < 400
                        title_looks_like_popup = any(k in title for k in popup_title_keywords)

                        if not has_complete_text and not (is_small_popup and title_looks_like_popup):
                            continue

                        log("[OK] AIR 저장 완료 팝업 감지")
                        log(f" - backend: {backend}")
                        log(f" - title: {title}")
                        log(f" - rect: {rect}")

                        try:
                            w.set_focus()
                            time.sleep(0.2)
                        except Exception:
                            pass

                        ok_btn = find_ok_button_in_dialog(w)

                        if ok_btn is not None:
                            log("[STEP] 저장 완료 팝업 OK 버튼 클릭")
                            click_control_by_real_mouse(ok_btn, "저장 완료 팝업 OK 버튼")
                            time.sleep(0.3)
                            return True

                        # 버튼을 못 찾으면 Enter로 닫기
                        log("[STEP] OK 버튼을 못 찾아 Enter로 저장 완료 팝업 닫기")
                        try:
                            w.set_focus()
                            time.sleep(0.2)
                            send_keys("{ENTER}")
                            time.sleep(0.3)
                            return True
                        except Exception as e:
                            log(f"[WARN] Enter로 저장 완료 팝업 닫기 실패: {e}")

                    except Exception:
                        pass
            except Exception:
                pass

        time.sleep(0.3)

    log("[INFO] AIR 저장 완료 팝업이 감지되지 않았습니다.")
    return False


# =========================================================
# 10. 마스터 엑셀 업데이트
# =========================================================

def normalize_excel_text(value):
    """
    엑셀 셀 값을 비교용 문자열로 정규화합니다.
    None, 공백, 날짜/숫자 표현 차이로 인한 불필요한 중복판정 오류를 줄이기 위한 함수입니다.
    """
    if value is None:
        return ""

    # datetime/date는 ISO 형태로 통일
    try:
        import datetime
        if isinstance(value, (datetime.datetime, datetime.date)):
            return value.isoformat(sep=" ").strip()
    except Exception:
        pass

    text = str(value).strip()

    # 엑셀에서 숫자 1.0으로 읽히는 경우 1로 통일
    if re.fullmatch(r"-?\d+\.0", text):
        text = text[:-2]

    return text


def normalize_excel_header(value):
    """
    헤더명 매핑용 정규화.
    예: '툴번호 ▲'와 '툴번호'를 같은 컬럼으로 보기 위해 정렬표시 문자를 제거합니다.
    """
    text = normalize_excel_text(value)
    text = text.replace("▲", "").replace("▼", "")
    text = re.sub(r"\s+", "", text)
    return text.lower()


def get_effective_max_row(ws, max_col=None):
    """
    서식만 남아 있는 빈 행을 제외하고 실제 값이 있는 마지막 행을 찾습니다.
    """
    if max_col is None:
        max_col = ws.max_column

    for row in range(ws.max_row, 0, -1):
        for col in range(1, max_col + 1):
            if ws.cell(row=row, column=col).value not in (None, ""):
                return row

    return 1


def get_effective_max_col(ws, header_row=1):
    """
    헤더가 있는 마지막 열을 찾습니다.
    A열 헤더가 비어 있어도 기존 사용범위는 유지해야 하므로 ws.max_column을 기본으로 사용합니다.
    """
    max_col = ws.max_column

    # 너무 오른쪽에 서식만 있는 경우를 대비해, 헤더 또는 데이터가 있는 마지막 열을 보정
    for col in range(ws.max_column, 0, -1):
        has_value = False
        for row in range(1, min(ws.max_row, 50) + 1):
            if ws.cell(row=row, column=col).value not in (None, ""):
                has_value = True
                break
        if has_value:
            max_col = col
            break

    return max_col


def build_header_map(ws, header_row=1, max_col=None):
    """
    헤더명 -> 열 번호 매핑.
    A열처럼 헤더가 비어 있는 열은 제외합니다.
    """
    if max_col is None:
        max_col = get_effective_max_col(ws, header_row)

    result = {}

    for col in range(1, max_col + 1):
        header = ws.cell(row=header_row, column=col).value
        key = normalize_excel_header(header)

        if key:
            result[key] = col

    return result


def parse_master_recent_datetime(value):
    """
    마스터 엑셀의 '최근검사' 값을 datetime으로 변환합니다.
    - Excel 날짜 셀(datetime/date)
    - Excel serial number
    - 문자열 날짜/시간(2026-05-31 20:30:00, 2026/05/31, 2026.05.31, 오전/오후 포함)
    을 최대한 처리합니다.
    """
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel
            converted = from_excel(value)
            if isinstance(converted, datetime):
                return converted
            if isinstance(converted, date):
                return datetime.combine(converted, datetime.min.time())
        except Exception:
            pass

    text = str(value).strip()
    if not text:
        return None

    # 날짜 추출: 2026-05-31 / 2026.05.31 / 2026/05/31 / 26-05-31 대응
    date_match = re.search(r"(\d{2,4})[\-./년\s]+(\d{1,2})[\-./월\s]+(\d{1,2})", text)
    if not date_match:
        return None

    year = int(date_match.group(1))
    month = int(date_match.group(2))
    day = int(date_match.group(3))

    if year < 100:
        year += 2000

    hour = 0
    minute = 0
    second = 0

    time_match = re.search(r"(\d{1,2})\s*:\s*(\d{1,2})(?:\s*:\s*(\d{1,2}))?", text)
    if time_match:
        hour = int(time_match.group(1))
        minute = int(time_match.group(2))
        second = int(time_match.group(3) or 0)

        if "오후" in text and hour < 12:
            hour += 12
        if "오전" in text and hour == 12:
            hour = 0

    try:
        return datetime(year, month, day, hour, minute, second)
    except ValueError:
        return None


def get_master_global_latest_recent_query_plan():
    """
    AIR 조회 전에 반드시 마스터 엑셀을 먼저 확인합니다.

    요청 로직:
    1. 마스터 엑셀 List 시트를 먼저 확인
    2. 최근검사 열의 값을 수집
    3. 일반사출/스탬핑/도금 전체 데이터 기준으로 최근검사 값을 내림차순 정렬
    4. 정렬 결과의 1번째, 즉 가장 최근 최근검사 날짜를 AIR 조회 시작일 From으로 사용
    5. 조회 종료일 To는 항상 오늘 날짜 사용

    주의:
    - 마스터 엑셀 파일 자체를 실제로 정렬하지 않습니다.
    - 코드 내부에서만 최근검사 값을 리스트로 모아 내림차순 정렬합니다.
    - 날짜 입력 방식은 기존 방식 그대로 사용합니다.
    """
    today = date.today().strftime("%Y-%m-%d")
    fallback_start = (date.today() - timedelta(days=QUERY_START_FALLBACK_DAYS)).strftime("%Y-%m-%d")

    log("\n" + "=" * 80)
    log("[MASTER-CHECK-START] AIR 실행 전 마스터 엑셀 먼저 확인 시작")
    log(f"[MASTER-CHECK] 설정값 USE_MASTER_LATEST_RECENT_DATE_FOR_FROM={USE_MASTER_LATEST_RECENT_DATE_FOR_FROM}")
    log(f"[MASTER-CHECK] 마스터 파일 경로: {MASTER_EXCEL_PATH}")
    log(f"[MASTER-CHECK] 대상 시트명: {MASTER_SHEET_NAME}")
    log(f"[MASTER-CHECK] 최근검사 기준 헤더명: {MASTER_RECENT_INSPECTION_HEADER}")
    log(f"[MASTER-CHECK] 조회 종료일 To 기준: 오늘 날짜 {today}")
    log("=" * 80)

    latest_dt = None
    latest_dept = None
    latest_row = None
    source = ""

    if not USE_MASTER_LATEST_RECENT_DATE_FOR_FROM:
        start_date = fallback_start
        source = "fallback_disabled"
        log(f"[WARN] 마스터 최근검사 기준 조회가 비활성화되어 fallback From={start_date} 사용")
    else:
        from openpyxl import load_workbook

        master_path = Path(MASTER_EXCEL_PATH)
        log(f"[MASTER-CHECK] 마스터 파일 존재 여부 확인: {master_path.exists()}")

        if not master_path.exists():
            raise FileNotFoundError(f"마스터 엑셀 파일을 찾지 못했습니다: {master_path}")

        log("[MASTER-CHECK] 마스터 엑셀 openpyxl 로드 시작")
        wb = load_workbook(master_path, data_only=True)

        try:
            log(f"[MASTER-CHECK] 마스터 엑셀 로드 완료. 시트 목록: {wb.sheetnames}")

            if MASTER_SHEET_NAME in wb.sheetnames:
                ws = wb[MASTER_SHEET_NAME]
                log(f"[MASTER-CHECK] 지정 시트 사용: {ws.title}")
            else:
                ws = wb.active
                log(f"[WARN] '{MASTER_SHEET_NAME}' 시트를 찾지 못해 활성 시트 '{ws.title}'를 사용합니다.")

            max_col = get_effective_max_col(ws, header_row=1)
            last_row = get_effective_max_row(ws, max_col)
            header_map = build_header_map(ws, header_row=1, max_col=max_col)

            dept_col = header_map.get(normalize_excel_header("부서명"), MASTER_DEPARTMENT_COL)
            recent_col = header_map.get(normalize_excel_header(MASTER_RECENT_INSPECTION_HEADER))

            log("\n[MASTER-CHECK] 마스터 구조 확인 결과")
            log(f" - 사용 시트: {ws.title}")
            log(f" - 유효 마지막 행: {last_row}")
            log(f" - 유효 마지막 열: {max_col}")
            log(f" - 부서명 열: {dept_col}열")
            log(f" - 최근검사 열: {recent_col if recent_col is not None else 'NOT FOUND'}")

            if recent_col is None:
                raise RuntimeError(
                    f"마스터 엑셀에서 '{MASTER_RECENT_INSPECTION_HEADER}' 열을 찾지 못했습니다. "
                    "1행 헤더명을 확인하세요."
                )

            target_departments = {normalize_text(job["department"]): job["department"] for job in DEPARTMENT_JOBS}

            dept_latest = {
                job["department"]: {
                    "dt": None,
                    "row": None,
                    "raw": None,
                    "count": 0,
                    "parsed_count": 0,
                }
                for job in DEPARTMENT_JOBS
            }

            recent_records = []

            log("\n[MASTER-CHECK] 최근검사 열 스캔 시작")
            log(" - 기준: 일반사출/스탬핑/도금 전체 데이터")
            log(" - 처리: 최근검사 값을 모두 수집한 뒤 내림차순 정렬")

            for row in range(2, last_row + 1):
                dept_value = ws.cell(row=row, column=dept_col).value
                dept_norm = normalize_text(dept_value)

                if dept_norm not in target_departments:
                    continue

                department = target_departments[dept_norm]
                dept_latest[department]["count"] += 1

                recent_value = ws.cell(row=row, column=recent_col).value
                recent_dt = parse_master_recent_datetime(recent_value)

                if recent_dt is None:
                    continue

                dept_latest[department]["parsed_count"] += 1

                recent_records.append({
                    "dt": recent_dt,
                    "department": department,
                    "row": row,
                    "raw": recent_value,
                })

                if dept_latest[department]["dt"] is None or recent_dt > dept_latest[department]["dt"]:
                    dept_latest[department]["dt"] = recent_dt
                    dept_latest[department]["row"] = row
                    dept_latest[department]["raw"] = recent_value

            log("\n[MASTER-CHECK] 부서별 최근검사 최댓값")
            for job in DEPARTMENT_JOBS:
                department = job["department"]
                info = dept_latest[department]
                if info["dt"] is None:
                    log(
                        f" - {department}: 최근검사 유효값 없음 "
                        f"/ 부서 행 수={info['count']} / 날짜 파싱 성공={info['parsed_count']}"
                    )
                else:
                    log(
                        f" - {department}: 최신 최근검사={info['dt']} "
                        f"/ 행={info['row']} / 원본값={info['raw']} "
                        f"/ 부서 행 수={info['count']} / 날짜 파싱 성공={info['parsed_count']}"
                    )

            log("\n[MASTER-CHECK] 최근검사 내림차순 정렬")
            log(f" - 정렬 대상 유효 최근검사 건수: {len(recent_records)}")

            if not recent_records:
                start_date = fallback_start
                source = "fallback_no_recent"
                log(f"[WARN] 전체 최근검사 값을 찾지 못해 fallback From={start_date} 사용")
            else:
                recent_records.sort(key=lambda item: item["dt"], reverse=True)

                log("[MASTER-CHECK] 최근검사 내림차순 TOP 10")
                for idx, item in enumerate(recent_records[:10], start=1):
                    log(
                        f" {idx}. 최근검사={item['dt']} "
                        f"/ 부서={item['department']} / 행={item['row']} / 원본값={item['raw']}"
                    )

                latest = recent_records[0]
                latest_dt = latest["dt"]
                latest_dept = latest["department"]
                latest_row = latest["row"]

                start_date = latest_dt.strftime("%Y-%m-%d")
                source = "master_recent_sorted_desc_top1"

                log("\n[MASTER-CHECK-RESULT] 최근검사 내림차순 1순위 기준 AIR 조회기간 확정")
                log(f" - 내림차순 1순위 최근검사: {latest_dt}")
                log(f" - 해당 부서: {latest_dept}")
                log(f" - 해당 행: {latest_row}")
                log(f" - AIR 공통 조회 시작일 From: {start_date}")
                log(f" - AIR 공통 조회 종료일 To: {today}")
                log(f" - AIR 조회 시간: From {DAILY_START_TIME} / To {DAILY_END_TIME}")

        finally:
            wb.close()
            log("[MASTER-CHECK] 마스터 엑셀 닫기 완료")

    query_plan = {}

    log("\n[MASTER-CHECK] AIR 부서별 조회 계획")
    for job in DEPARTMENT_JOBS:
        department = job["department"]
        query_plan[department] = {
            "start_date": start_date,
            "end_date": today,
            "latest_recent": latest_dt,
            "source": source,
        }

        log(
            f"[PLAN] {department}: From={start_date} {DAILY_START_TIME}, "
            f"To={today} {DAILY_END_TIME}, source={source}"
        )

    log("[MASTER-CHECK-END] 마스터 기준 조회기간 산정 완료. 이제 AIR 실행/조회 단계로 이동합니다.")
    log("=" * 80 + "\n")
    return query_plan

def build_master_row_key(ws, row, start_col, end_col):
    """
    중복 판정 키 생성.
    A열 순번은 제외하고 B~마지막열 값으로 비교합니다.
    """
    return tuple(
        normalize_excel_text(ws.cell(row=row, column=col).value)
        for col in range(start_col, end_col + 1)
    )


def build_existing_keys(ws, compare_start_col, end_col):
    """
    마스터 전체 기존 데이터의 중복 키 집합을 생성합니다.
    부서명이 B열에 포함되므로 전체 기준으로 비교해도 부서별 중복 판정이 가능합니다.
    """
    last_row = get_effective_max_row(ws, end_col)
    keys = set()

    for row in range(2, last_row + 1):
        key = build_master_row_key(ws, row, compare_start_col, end_col)
        if any(key):
            keys.add(key)

    return keys

def remove_duplicate_rows_in_master(ws, compare_start_col, end_col):
    """
    마스터 시트 전체에서 중복 행을 최종 제거합니다.

    기준:
    - 1행은 헤더이므로 제외
    - A열 순번은 제외
    - B열부터 마지막 열까지 값이 모두 같으면 중복으로 판단
    - 먼저 나온 행은 유지
    - 나중에 나온 중복 행은 삭제
    """
    last_row = get_effective_max_row(ws, end_col)

    seen_keys = set()
    duplicate_rows = []

    for row in range(2, last_row + 1):
        key = build_master_row_key(
            ws=ws,
            row=row,
            start_col=compare_start_col,
            end_col=end_col,
        )

        # 완전히 빈 행은 중복 판단 대상에서 제외
        if not any(key):
            continue

        if key in seen_keys:
            duplicate_rows.append(row)
        else:
            seen_keys.add(key)

    if not duplicate_rows:
        log("[DUP-CLEAN] 마스터 전체 중복 행 없음")
        return 0

    # 아래에서 위로 삭제해야 행 번호 밀림 문제가 없음
    for row in reversed(duplicate_rows):
        ws.delete_rows(row, 1)

    log(f"[DUP-CLEAN] 마스터 전체 중복 행 삭제 완료: {len(duplicate_rows)}건")
    return len(duplicate_rows)

def find_department_last_row(ws, department, dept_col=MASTER_DEPARTMENT_COL, max_col=None):
    """
    마스터에서 특정 부서명의 마지막 행을 찾습니다.
    """
    if max_col is None:
        max_col = ws.max_column

    target = normalize_text(department)
    last_row = get_effective_max_row(ws, max_col)

    for row in range(last_row, 1, -1):
        value = ws.cell(row=row, column=dept_col).value
        if normalize_text(value) == target:
            return row

    return None


def copy_row_style(ws, src_row, dst_row, max_col):
    """
    기존 부서 마지막 행의 서식을 새 삽입 행에 복사합니다.
    값은 복사하지 않고 스타일만 복사합니다.
    """
    from copy import copy

    if src_row is None or src_row < 1:
        return

    try:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    except Exception:
        pass

    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)

        if src.has_style:
            dst._style = copy(src._style)

        if src.number_format:
            dst.number_format = src.number_format

        if src.alignment:
            dst.alignment = copy(src.alignment)

        if src.font:
            dst.font = copy(src.font)

        if src.fill:
            dst.fill = copy(src.fill)

        if src.border:
            dst.border = copy(src.border)

        if src.protection:
            dst.protection = copy(src.protection)


def reindex_department_numbers(ws, department, index_col=MASTER_INDEX_COL, dept_col=MASTER_DEPARTMENT_COL, max_col=None):
    """
    특정 부서 데이터의 A열 순번을 1부터 다시 부여합니다.
    Daily 데이터 삽입 후에도 부서별 순번이 끊기지 않도록 하기 위한 로직입니다.
    """
    if max_col is None:
        max_col = ws.max_column

    target = normalize_text(department)
    last_row = get_effective_max_row(ws, max_col)
    seq = 1

    for row in range(2, last_row + 1):
        dept_value = ws.cell(row=row, column=dept_col).value
        if normalize_text(dept_value) == target:
            ws.cell(row=row, column=index_col).value = seq
            seq += 1

    return seq - 1


def get_daily_department_from_filename(file_path):
    """
    saved_exports가 문자열 경로만 들어온 경우 파일명에서 부서명을 보조 추출합니다.
    """
    name = Path(file_path).name

    for job in DEPARTMENT_JOBS:
        dept = job["department"]
        if dept in name:
            return dept

    return ""


def normalize_saved_exports(saved_exports):
    """
    saved_exports를 [{'department': ..., 'path': ...}] 형태로 통일합니다.
    """
    normalized = []

    for item in saved_exports:
        if isinstance(item, dict):
            path = item.get("path") or item.get("file") or item.get("saved_path")
            department = item.get("department") or get_daily_department_from_filename(path or "")
        else:
            path = str(item)
            department = get_daily_department_from_filename(path)

        if not path:
            continue

        if not department:
            log(f"[WARN] Daily 파일에서 부서명을 판단하지 못해 건너뜁니다: {path}")
            continue

        normalized.append({"department": department, "path": str(path)})

    return normalized


def read_daily_rows_for_master(daily_path, department, master_headers, master_max_col):
    """
    Daily 엑셀 파일에서 마스터 컬럼 구조에 맞춰 행 데이터를 읽습니다.
    - Daily 헤더와 마스터 헤더를 이름 기준으로 매핑
    - A열 순번은 나중에 마스터에서 재부여하므로 비워둠
    - B열 부서명은 비어 있으면 파일 기준 department로 채움
    """
    from openpyxl import load_workbook

    daily_path = str(daily_path)

    if not os.path.exists(daily_path):
        raise FileNotFoundError(f"Daily 파일을 찾지 못했습니다: {daily_path}")

    wb = load_workbook(daily_path, data_only=True)

    if MASTER_SHEET_NAME in wb.sheetnames:
        ws = wb[MASTER_SHEET_NAME]
    else:
        ws = wb.active

    daily_max_col = get_effective_max_col(ws, header_row=1)
    daily_last_row = get_effective_max_row(ws, daily_max_col)
    daily_header_map = build_header_map(ws, header_row=1, max_col=daily_max_col)

    rows = []
    dept_key = normalize_excel_header("부서명")
    target_dept_norm = normalize_text(department)

    for src_row in range(2, daily_last_row + 1):
        # 빈 행 제외
        has_value = False
        for col in range(1, daily_max_col + 1):
            if ws.cell(row=src_row, column=col).value not in (None, ""):
                has_value = True
                break
        if not has_value:
            continue

        new_values = [None] * master_max_col

        for master_col in range(1, master_max_col + 1):
            # A열 순번은 마스터에서 재인덱싱
            if master_col == MASTER_INDEX_COL:
                new_values[master_col - 1] = None
                continue

            master_header = master_headers.get(master_col, "")
            header_key = normalize_excel_header(master_header)

            if not header_key:
                continue

            daily_col = daily_header_map.get(header_key)
            if daily_col:
                new_values[master_col - 1] = ws.cell(row=src_row, column=daily_col).value

        # 부서명 컬럼 보정
        daily_dept_value = None
        daily_dept_col = daily_header_map.get(dept_key)
        if daily_dept_col:
            daily_dept_value = ws.cell(row=src_row, column=daily_dept_col).value

        if normalize_excel_text(daily_dept_value):
            # Daily 파일 안에 부서명이 있는데 현재 처리 부서와 다르면 스킵
            if normalize_text(daily_dept_value) != target_dept_norm:
                continue
        else:
            daily_dept_value = department

        new_values[MASTER_DEPARTMENT_COL - 1] = daily_dept_value
        rows.append(new_values)

    wb.close()
    return rows


def insert_department_rows_to_master(ws, department, rows_to_add, master_max_col):
    """
    중복 제거 후 Daily 데이터를 마스터에 삽입합니다.
    APPEND_NEW_ROWS_TO_BOTTOM=True이면 마스터 전체 마지막 행 아래에 추가하고,
    False이면 해당 부서 마지막 행 바로 아래에 삽입합니다.
    삽입 후 해당 부서 A열을 1부터 재인덱싱합니다.
    """
    if not rows_to_add:
        return {"department": department, "read": 0, "inserted": 0, "duplicates": 0, "last_count": 0}

    existing_keys = build_existing_keys(
        ws,
        compare_start_col=DUPLICATE_COMPARE_START_COL,
        end_col=master_max_col,
    )

    unique_rows = []
    duplicate_count = 0

    for values in rows_to_add:
        key = tuple(normalize_excel_text(values[col - 1]) for col in range(DUPLICATE_COMPARE_START_COL, master_max_col + 1))

        if not any(key):
            continue

        if key in existing_keys:
            duplicate_count += 1
            continue

        existing_keys.add(key)
        unique_rows.append(values)

    if not unique_rows:
        last_count = reindex_department_numbers(ws, department, max_col=master_max_col)
        return {
            "department": department,
            "read": len(rows_to_add),
            "inserted": 0,
            "duplicates": duplicate_count,
            "last_count": last_count,
        }

    dept_last_row = find_department_last_row(ws, department, max_col=master_max_col)

    if APPEND_NEW_ROWS_TO_BOTTOM:
        # 요청사항 반영: 공정 순서와 무관하게 마스터 전체 마지막 행 아래에 추가
        insert_at = get_effective_max_row(ws, master_max_col) + 1
        template_row = dept_last_row if dept_last_row is not None else insert_at - 1
        log(f"[STEP] {department}: 마스터 전체 마지막 행 아래에 추가합니다.")
    else:
        # 기존 방식: 해당 부서의 마지막 행 바로 아래에 삽입하여 부서 블록을 유지
        if dept_last_row is None:
            insert_at = get_effective_max_row(ws, master_max_col) + 1
            template_row = insert_at - 1
            log(f"[WARN] 마스터에서 '{department}' 기존 행을 찾지 못했습니다. 마지막 행 아래에 추가합니다.")
        else:
            insert_at = dept_last_row + 1
            template_row = dept_last_row

    log(f"[STEP] {department}: 삽입 위치={insert_at}행, 삽입 대상={len(unique_rows)}건, 중복 제외={duplicate_count}건")

    ws.insert_rows(insert_at, amount=len(unique_rows))

    for offset, values in enumerate(unique_rows):
        dst_row = insert_at + offset
        copy_row_style(ws, template_row, dst_row, master_max_col)

        for col in range(1, master_max_col + 1):
            ws.cell(row=dst_row, column=col).value = values[col - 1]

    last_count = reindex_department_numbers(ws, department, max_col=master_max_col)

    return {
        "department": department,
        "read": len(rows_to_add),
        "inserted": len(unique_rows),
        "duplicates": duplicate_count,
        "last_count": last_count,
    }


def update_master_excel_with_daily_exports(saved_exports):
    """
    AIR에서 부서별로 내보낸 Daily 엑셀 파일을 마스터 파일에 반영합니다.

    처리 내용:
    1. Daily 파일을 부서명 기준으로 마스터 List 시트에 매핑
    2. A열 제외 B~마지막열 전체 값 기준 중복 제거
    3. 신규 데이터는 설정에 따라 마스터 전체 마지막 행 또는 해당 부서 마지막 행 아래에 삽입
    4. 삽입 후 각 부서 A열 순번 재인덱싱
    5. 백업 생성 후 마스터 파일에 저장
    """
    if not UPDATE_MASTER_AFTER_EXPORT:
        log("[SKIP] UPDATE_MASTER_AFTER_EXPORT=False 이므로 마스터 업데이트를 생략합니다.")
        return []

    from openpyxl import load_workbook

    master_path = Path(MASTER_EXCEL_PATH)

    if not master_path.exists():
        raise FileNotFoundError(f"마스터 엑셀 파일을 찾지 못했습니다: {master_path}")

    exports = normalize_saved_exports(saved_exports)

    if not exports:
        log("[WARN] 마스터에 반영할 Daily 파일이 없습니다.")
        return []

    log("\n" + "=" * 70)
    log("[START] 마스터 엑셀 Daily 데이터 삽입 시작")
    log(f" - 마스터 파일: {master_path}")
    log(f" - 대상 파일 수: {len(exports)}")
    log("=" * 70)

    if CREATE_MASTER_BACKUP:
        backup_path = master_path.with_name(
            f"{master_path.stem}_backup_{time.strftime('%Y%m%d_%H%M%S')}{master_path.suffix}"
        )
        shutil.copy2(master_path, backup_path)
        log(f"[OK] 마스터 백업 생성: {backup_path}")

    wb = load_workbook(master_path)

    if MASTER_SHEET_NAME in wb.sheetnames:
        ws = wb[MASTER_SHEET_NAME]
    else:
        ws = wb.active
        log(f"[WARN] '{MASTER_SHEET_NAME}' 시트를 찾지 못해 활성 시트 '{ws.title}'를 사용합니다.")

    master_max_col = get_effective_max_col(ws, header_row=1)
    master_headers = {
        col: ws.cell(row=1, column=col).value
        for col in range(1, master_max_col + 1)
    }

    summary = []

    for export in exports:
        department = export["department"]
        daily_path = export["path"]

        log("\n" + "-" * 60)
        log(f"[DEPT] {department}")
        log(f"[FILE] {daily_path}")

        daily_rows = read_daily_rows_for_master(
            daily_path=daily_path,
            department=department,
            master_headers=master_headers,
            master_max_col=master_max_col,
        )

        result = insert_department_rows_to_master(
            ws=ws,
            department=department,
            rows_to_add=daily_rows,
            master_max_col=master_max_col,
        )

        summary.append(result)

        log(
            f"[RESULT] {department}: Daily읽음={result['read']}건, "
            f"신규삽입={result['inserted']}건, 중복제외={result['duplicates']}건, "
            f"재인덱싱후총건수={result['last_count']}건"
        )

    # ------------------------------------------------------------
    # 최종 안전장치: 모든 Daily 업데이트 완료 후 마스터 전체 중복 제거
    # ------------------------------------------------------------
    final_duplicate_deleted = remove_duplicate_rows_in_master(
        ws=ws,
        compare_start_col=DUPLICATE_COMPARE_START_COL,
        end_col=master_max_col,
    )

    # 중복 삭제 후 A열 부서별 순번 다시 정리
    for job in DEPARTMENT_JOBS:
        reindex_department_numbers(
            ws=ws,
            department=job["department"],
            max_col=master_max_col,
        )

    log(f"[DUP-CLEAN] 최종 중복 삭제 수: {final_duplicate_deleted}건")

    try:
        wb.save(master_path)
    except PermissionError:
        raise PermissionError(
            "마스터 엑셀 파일 저장에 실패했습니다. "
            "파일이 Excel에서 열려 있으면 닫은 뒤 다시 실행하세요. "
            f"대상 파일: {master_path}"
        )
    finally:
        wb.close()

    log("\n" + "=" * 70)
    log("[OK] 마스터 엑셀 업데이트 완료")
    log(f" - 저장 파일: {master_path}")
    log("=" * 70)

    return summary


# =========================================================
# 11. 부서별 AIR 내보내기 실행
# =========================================================

def run_air_export_for_department(
    department: str,
    department_index: int,
    start_date: str,
    end_date: str,
    start_time: str,
    end_time: str,
    set_period: bool = True,
) -> str:
    export_path = make_export_path(department)

    log("\n" + "=" * 70)
    log(f"[START] AIR 조회/내보내기 시작: {department}")
    log(f" - 부서 인덱스: {department_index}")
    log(f" - 기간: {start_date} {start_time} ~ {end_date} {end_time}")
    log(f" - 기간/시간 설정 여부: {set_period}")
    log(f" - 저장 경로: {export_path}")
    log("=" * 70)

    win = refresh_air_window()
    win = enter_inspection_history_strict(win)

    form = input_conditions(
        win=win,
        department=department,
        department_index=department_index,
        start_date=start_date,
        end_date=end_date,
        start_time=start_time,
        end_time=end_time,
        set_period=set_period,
    )

    click_search(win, form=form)

    saved_path = click_excel_export(win, export_path, form=form)

    log(f"[OK] AIR 내보내기 완료: {saved_path}")

    return saved_path


# =========================================================
# 12. 메인 실행
# =========================================================

PIPELINE_STEPS = [
    "마스터 최근검사 확인",
    "일반사출 조회/내보내기",
    "스탬핑 조회/내보내기",
    "도금 조회/내보내기",
    "자동으로 열린 Excel 닫기",
    "마스터 업데이트",
    "중복 제거 및 저장",
]


def execute_automation_pipeline(progress_callback=None):
    global RUN_EXPORT_DIR
    global STOP_REQUESTED

    STOP_REQUESTED = False
    CONTROL_CACHE.clear()

    def update_progress(step_name: str, status: str, detail: str = ""):
        if progress_callback:
            progress_callback(step_name, status, detail)

    log("===== AIR 검사이력 부서별 엑셀 내보내기 + 마스터 업데이트 =====")
    log(f"CODE_VERSION: {CODE_VERSION}")

    Path(EXPORT_DIR).mkdir(parents=True, exist_ok=True)

    # 이번 실행에서 생성되는 Daily 파일 3개만 따로 관리하기 위한 실행 폴더를 먼저 생성합니다.
    # 예: AIR_Daily_Export\20260601_1
    RUN_EXPORT_DIR = create_run_export_dir()

    log("\n[INFO] 이번 실행 Daily 엑셀 저장 폴더")
    log(f"오늘 내보낸 파일 저장 폴더- {RUN_EXPORT_DIR}")

    # AIR/Excel에서 생성한 xlsx의 pageSetup collated 속성 때문에
    # openpyxl load_workbook()이 실패하는 문제를 방지합니다.
    # 마스터 엑셀 확인도 load_workbook()을 사용하므로 AIR 실행 전 먼저 적용합니다.
    patch_openpyxl_print_page_setup_collated()

    start_time = DAILY_START_TIME
    end_time = DAILY_END_TIME

    ensure_not_stopped()
    update_progress("마스터 최근검사 확인", "in_progress", "마스터 최근검사 확인 진행 중")

    # 핵심 변경:
    # AIR를 조회하기 전에 항상 마스터 엑셀을 먼저 확인하여
    # 전체 최근검사 최댓값의 날짜를 모든 부서 공통 From으로 사용합니다.
    query_plan = get_master_global_latest_recent_query_plan()

    latest_start_date = ""
    if DEPARTMENT_JOBS:
        first_department = DEPARTMENT_JOBS[0]["department"]
        latest_start_date = str(query_plan.get(first_department, {}).get("start_date", ""))

    if latest_start_date:
        update_progress("__QUERY_BASIS__", "info", latest_start_date)

    update_progress("마스터 최근검사 확인", "completed", "마스터 최근검사 확인 완료")

    log("\n[INFO] AIR 공통 조회 조건")
    for job in DEPARTMENT_JOBS:
        department = job["department"]
        plan = query_plan[department]
        log(f" - {department}: From={plan['start_date']} {start_time} / To={plan['end_date']} {end_time}")

    log(f" - 기본 저장 폴더: {EXPORT_DIR}")
    log(f" - 이번 실행 저장 폴더: {RUN_EXPORT_DIR}")
    log(f" - 마스터 파일: {MASTER_EXCEL_PATH}")

    ensure_not_stopped()
    launch_air_if_needed()
    refresh_air_window()

    saved_exports = []

    for idx, job in enumerate(DEPARTMENT_JOBS):
        department = job["department"]
        department_index = job["combo_index"]
        plan = query_plan[department]

        # 3개 부서는 같은 조회기간을 사용하므로 첫 번째 부서에서만 기간/시간을 설정합니다.
        # 이후 부서는 AIR 화면에 남아 있는 기간/시간을 그대로 사용하고 부서명만 변경합니다.
        set_period = idx == 0

        step_name = f"{department} 조회/내보내기"
        update_progress(step_name, "in_progress", f"{department} 조회/내보내기 진행 중")

        try:
            ensure_not_stopped()
            saved_path = run_air_export_for_department(
                department=department,
                department_index=department_index,
                start_date=plan["start_date"],
                end_date=plan["end_date"],
                start_time=start_time,
                end_time=end_time,
                set_period=set_period,
            )

            saved_exports.append({
                "department": department,
                "path": saved_path,
                "from": plan["start_date"],
                "to": plan["end_date"],
            })
            update_progress(step_name, "completed", f"{department} 조회/내보내기 완료")

        except Exception as e:
            log(f"[ERROR] {department} 내보내기 실패")
            log(str(e))
            update_progress(step_name, "error", f"{department} 조회/내보내기 실패")
            continue

    if saved_exports:
        log("\n[INFO] AIR 부서별 엑셀 내보내기 완료")
        for export in saved_exports:
            log(f" - {export['department']}: {export['path']}")

        ensure_not_stopped()
        update_progress("자동으로 열린 Excel 닫기", "in_progress", "자동으로 열린 Excel 닫기 진행 중")

        # 저장 완료 팝업 OK 후 자동으로 열린 Daily 엑셀 창 때문에
        # openpyxl이 파일을 읽거나 마스터를 업데이트할 때 문제가 생길 수 있습니다.
        # 단, 사용자가 작업 중인 다른 엑셀 파일은 절대 닫지 않고, 이번에 내보낸 파일만 닫습니다.
        closed_ok = close_only_exported_excel_workbooks(saved_exports, timeout=20)

        if not closed_ok:
            update_progress("자동으로 열린 Excel 닫기", "error", "자동으로 열린 Excel 닫기 실패")
            raise RuntimeError(
                "내보낸 엑셀 파일이 아직 Excel에서 열려 있거나 잠금 상태입니다. "
                "자동으로 열린 Daily 엑셀 파일을 닫은 뒤 다시 실행하세요."
            )

        update_progress("자동으로 열린 Excel 닫기", "completed", "자동으로 열린 Excel 닫기 완료")

        patch_openpyxl_print_page_setup_collated()

        try:
            ensure_not_stopped()
            update_progress("마스터 업데이트", "in_progress", "마스터 업데이트 진행 중")
            update_progress("중복 제거 및 저장", "in_progress", "중복 제거 및 저장 진행 중")
            update_master_excel_with_daily_exports(saved_exports)
            update_progress("마스터 업데이트", "completed", "마스터 업데이트 완료")
            update_progress("중복 제거 및 저장", "completed", "중복 제거 및 저장 완료")
        except Exception as e:
            log("[ERROR] 마스터 엑셀 업데이트 실패")
            log(str(e))
            update_progress("마스터 업데이트", "error", "마스터 업데이트 실패")
            update_progress("중복 제거 및 저장", "error", "중복 제거 및 저장 실패")
            raise
    else:
        log("[WARN] 저장된 Daily 파일이 없어 마스터 엑셀 업데이트를 생략합니다.")

    return saved_exports


def main():
    execute_automation_pipeline()


class AirAutomationApp(ctk.CTk if ctk else object):
    CLEAN_KEYWORDS = [
        "[OK]",
        "[WARN]",
        "[ERROR]",
        "[START]",
        "[STEP]",
        "[RESULT]",
        "[DUP-CLEAN]",
        "[MASTER-CHECK-RESULT]",
    ]

    def __init__(self):
        super().__init__()

        self.title("AIR 검사이력 자동 내보내기 + 마스터 업데이트")
        self.geometry("1180x960")
        self.minsize(1024, 840)

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.font_family = "맑은 고딕"
        self.font_title = ctk.CTkFont(family=self.font_family, size=24, weight="bold")
        self.font_section = ctk.CTkFont(family=self.font_family, size=16, weight="bold")
        self.font_body = ctk.CTkFont(family=self.font_family, size=12)
        self.font_body_bold = ctk.CTkFont(family=self.font_family, size=13, weight="bold")

        self.raw_log_visible = False
        self.worker_thread = None

        self.step_states = {step: "대기" for step in PIPELINE_STEPS}
        self.step_labels = {}
        self.department_vars = {}

        self._build_layout()
        self._reset_progress_ui()
        self._load_paths_on_startup()

    def _build_layout(self):
        self.main_container = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.main_container.pack(fill="both", expand=True, padx=16, pady=16)
        self.main_container.grid_columnconfigure(0, weight=1)

        self.header_frame = ctk.CTkFrame(self.main_container, corner_radius=10)
        self.header_frame.grid(row=0, column=0, sticky="ew", padx=0, pady=(0, 10))
        self.header_frame.grid_columnconfigure(0, weight=1)

        self.title_label = ctk.CTkLabel(
            self.header_frame,
            text="AIR 검사이력 자동 내보내기 + 마스터 업데이트",
            font=self.font_title,
            anchor="w",
        )
        self.title_label.grid(row=0, column=0, sticky="w", padx=14, pady=(12, 2))

        self.version_label = ctk.CTkLabel(
            self.header_frame,
            text=f"Version: {GUI_VERSION}",
            font=self.font_body,
            anchor="w",
        )
        self.version_label.grid(row=1, column=0, sticky="w", padx=14, pady=(0, 12))

        self.action_frame = ctk.CTkFrame(self.main_container, corner_radius=10)
        self.action_frame.grid(row=1, column=0, sticky="ew", padx=0, pady=(0, 10))

        self.action_button_row = ctk.CTkFrame(self.action_frame, fg_color="transparent")
        self.action_button_row.grid(row=0, column=0, sticky="w", padx=14, pady=(14, 10))

        self.run_button = ctk.CTkButton(
            self.action_button_row,
            text="실행",
            width=120,
            fg_color="#3b82f6",
            hover_color="#2563eb",
            text_color="#ffffff",
            font=self.font_body_bold,
            command=self._on_run_clicked,
        )
        self.run_button.grid(row=0, column=0, padx=(0, 8), pady=0)

        self.stop_button = ctk.CTkButton(
            self.action_button_row,
            text="중지",
            width=90,
            fg_color="#ff727e",
            hover_color="#fd6d75",
            text_color="#ffffff",
            font=self.font_body_bold,
            command=self._on_stop_clicked,
            state="disabled",
        )
        self.stop_button.grid(row=0, column=1, padx=(0, 8), pady=0)

        self.open_export_folder_button = ctk.CTkButton(
            self.action_button_row,
            text="오늘 내보낸 파일 저장 폴더 열기",
            width=130,
            fg_color="#60a5fa",
            hover_color="#3b82f6",
            text_color="#ffffff",
            font=self.font_body_bold,
            command=self._open_export_folder,
        )
        self.open_export_folder_button.grid(row=0, column=2, padx=(0, 8), pady=0)

        self.open_master_file_button = ctk.CTkButton(
            self.action_button_row,
            text="마스터 파일 열기",
            width=160,
            fg_color="#60a5fa",
            hover_color="#3b82f6",
            text_color="#ffffff",
            font=self.font_body_bold,
            command=self._open_master_file,
        )
        self.open_master_file_button.grid(row=0, column=3, padx=(0, 0), pady=0)

        self.current_status_label = ctk.CTkLabel(
            self.action_frame,
            text="현재 상태: 대기 중",
            font=self.font_body_bold,
            anchor="w",
            corner_radius=8,
            padx=10,
            pady=4,
            fg_color="#7CADFC",
            text_color="#ffffff",
        )
        self.current_status_label.grid(row=1, column=0, sticky="w", padx=14, pady=(0, 6))

        self.run_folder_label = ctk.CTkLabel(
            self.action_frame,
            text=f"오늘 내보낸 파일 저장 위치: {EXPORT_DIR}",
            font=self.font_body,
            anchor="w",
            justify="left",
        )
        self.run_folder_label.grid(row=2, column=0, sticky="ew", padx=14, pady=(0, 4))

        self.master_file_label = ctk.CTkLabel(
            self.action_frame,
            text=f"마스터 파일: {MASTER_EXCEL_PATH}",
            font=self.font_body,
            anchor="w",
            justify="left",
        )
        self.master_file_label.grid(row=3, column=0, sticky="ew", padx=14, pady=(0, 14))

        self.condition_summary_frame = ctk.CTkFrame(self.main_container, corner_radius=10)
        self.condition_summary_frame.grid(row=2, column=0, sticky="ew", padx=0, pady=(0, 10))
        self.condition_summary_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            self.condition_summary_frame,
            text="조회 조건 요약",
            font=self.font_section,
            anchor="w",
        ).grid(row=0, column=0, sticky="w", padx=14, pady=(12, 6))

        self.start_time_label = ctk.CTkLabel(
            self.condition_summary_frame,
            text=f"조회 시작 시간: {DAILY_START_TIME}",
            font=self.font_body,
            anchor="w",
        )
        self.start_time_label.grid(row=1, column=0, sticky="w", padx=14, pady=2)

        self.end_time_label = ctk.CTkLabel(
            self.condition_summary_frame,
            text=f"조회 종료 시간: {DAILY_END_TIME}",
            font=self.font_body,
            anchor="w",
        )
        self.end_time_label.grid(row=2, column=0, sticky="w", padx=14, pady=2)

        self.query_basis_label = ctk.CTkLabel(
            self.condition_summary_frame,
            text="조회 시작일 기준: 실행 후 자동 계산",
            font=self.font_body,
            anchor="w",
        )
        self.query_basis_label.grid(row=3, column=0, sticky="w", padx=14, pady=(2, 8))

        self.department_check_area = ctk.CTkFrame(self.condition_summary_frame, fg_color="transparent")
        self.department_check_area.grid(row=4, column=0, sticky="w", padx=14, pady=(0, 12))

        for idx, job in enumerate(DEPARTMENT_JOBS):
            department = job["department"]
            var = ctk.BooleanVar(value=False)
            chk = ctk.CTkCheckBox(
                self.department_check_area,
                text=department,
                variable=var,
                font=self.font_body,
                text_color="#ffffff",
                border_color="#93c5fd",
                fg_color="#3b82f6",
                hover_color="#2563eb",
                checkmark_color="#ffffff",
                state="disabled",
            )
            chk.grid(row=0, column=idx, padx=(0, 20), pady=2, sticky="w")
            self.department_vars[department] = var

        self.progress_frame = ctk.CTkFrame(self.main_container, corner_radius=10)
        self.progress_frame.grid(row=3, column=0, sticky="ew", padx=0, pady=(0, 10))
        self.progress_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            self.progress_frame,
            text="진행 상태",
            font=self.font_section,
            anchor="w",
        ).grid(row=0, column=0, sticky="w", padx=14, pady=(12, 6))

        self.total_progress_label = ctk.CTkLabel(
            self.progress_frame,
            text="전체 진행률: 0%",
            font=self.font_body,
            anchor="w",
        )
        self.total_progress_label.grid(row=1, column=0, sticky="w", padx=14, pady=(0, 4))

        self.total_progress_bar = ctk.CTkProgressBar(self.progress_frame)
        self.total_progress_bar.grid(row=2, column=0, sticky="ew", padx=14, pady=(0, 10))
        self.total_progress_bar.set(0.0)

        self.step_status_list = ctk.CTkFrame(self.progress_frame)
        self.step_status_list.grid(row=3, column=0, sticky="ew", padx=14, pady=(0, 12))
        self.step_status_list.grid_columnconfigure(0, weight=1)

        for idx, step in enumerate(PIPELINE_STEPS, start=1):
            label = ctk.CTkLabel(
                self.step_status_list,
                text=f"{idx}. {step:<20} 대기",
                font=self.font_body,
                anchor="w",
                justify="left",
            )
            label.grid(row=idx - 1, column=0, sticky="ew", pady=1)
            self.step_labels[step] = label

        self.clean_log_frame = ctk.CTkFrame(self.main_container, corner_radius=10)
        self.clean_log_frame.grid(row=4, column=0, sticky="ew", padx=0, pady=(0, 10))
        self.clean_log_frame.grid_columnconfigure(0, weight=1)
        self.clean_log_frame.grid_rowconfigure(1, weight=1)

        self.clean_log_title = ctk.CTkLabel(
            self.clean_log_frame,
            text="로그",
            font=self.font_section,
            anchor="w",
        )
        self.clean_log_title.grid(row=0, column=0, sticky="w", padx=14, pady=(12, 6))

        self.clean_log_textbox = ctk.CTkTextbox(self.clean_log_frame, height=240, wrap="word", font=self.font_body)
        self.clean_log_textbox.grid(row=1, column=0, sticky="nsew", padx=14, pady=(0, 12))

        self.raw_log_frame = ctk.CTkFrame(self.main_container, corner_radius=10)
        self.raw_log_frame.grid(row=5, column=0, sticky="ew", padx=0, pady=(0, 0))
        self.raw_log_frame.grid_columnconfigure(0, weight=1)
        self.raw_log_frame.grid_rowconfigure(1, weight=1)

        self.raw_log_toggle_button = ctk.CTkButton(
            self.raw_log_frame,
            text="Raw Log 보기 ⏷",
            font=self.font_body_bold,
            command=self._toggle_raw_log,
        )
        self.raw_log_toggle_button.grid(row=0, column=0, sticky="ew", padx=14, pady=(12, 10))

        self.raw_log_textbox = ctk.CTkTextbox(self.raw_log_frame, height=180, wrap="none", font=self.font_body)
        self.raw_log_textbox.grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 12))
        self.raw_log_textbox.grid_remove()

    def _set_status_text(self, text: str):
        color_map = {
            "대기 중": ("#374151", "#ffffff"),
            "실행 중": ("#1d4ed8", "#ffffff"),
            "중지 요청됨": ("#b45309", "#ffffff"),
            "중지됨": ("#6b7280", "#ffffff"),
            "완료": ("#047857", "#ffffff"),
            "오류": ("#b91c1c", "#ffffff"),
        }
        fg, tc = color_map.get(text, ("#64799B", "#e5e7eb"))
        self.current_status_label.configure(text=f"현재 상태: {text}", fg_color=fg, text_color=tc)

    def _reset_progress_ui(self):
        self.step_states = {step: "대기" for step in PIPELINE_STEPS}

        for idx, step in enumerate(PIPELINE_STEPS, start=1):
            self.step_labels[step].configure(text=f"{idx}. {step:<20} 대기")

        for department, var in self.department_vars.items():
            var.set(False)

        self.query_basis_label.configure(text="조회 시작일 기준: 실행 후 자동 계산")

        self.total_progress_bar.set(0.0)
        self.total_progress_label.configure(text="전체 진행률: 0%")

    def _on_progress(self, step_name: str, status: str, detail: str = ""):
        self.after(0, self._apply_progress, step_name, status, detail)

    def _apply_progress(self, step_name: str, status: str, detail: str = ""):
        if step_name == "__QUERY_BASIS__":
            self.query_basis_label.configure(
                text=f"조회 시작일 기준: {detail} (마스터 List 시트의 최근검사 최신 날짜)"
            )
            return

        status_text_map = {
            "pending": "대기",
            "in_progress": "진행 중",
            "completed": "완료",
            "error": "오류",
        }

        text_status = status_text_map.get(status, status)
        self.step_states[step_name] = text_status

        if step_name in self.step_labels:
            idx = PIPELINE_STEPS.index(step_name) + 1
            self.step_labels[step_name].configure(text=f"{idx}. {step_name:<20} {text_status}")

        for department, var in self.department_vars.items():
            if step_name == f"{department} 조회/내보내기":
                var.set(status == "completed")

        completed_count = sum(1 for s in self.step_states.values() if s == "완료")
        ratio = completed_count / max(1, len(PIPELINE_STEPS))
        self.total_progress_bar.set(ratio)
        self.total_progress_label.configure(text=f"전체 진행률: {int(ratio * 100)}%")

        if detail:
            self._append_clean_log(detail)

    def _append_raw_log(self, message: str):
        self.raw_log_textbox.insert("end", f"{message}\n")
        self.raw_log_textbox.see("end")

    def _append_clean_log(self, message: str):
        stamp = datetime.now().strftime("%H:%M:%S")
        self.clean_log_textbox.insert("end", f"{stamp}  {message}\n")
        self.clean_log_textbox.see("end")

    def _is_clean_log_target(self, message: str) -> bool:
        if any(keyword in message for keyword in self.CLEAN_KEYWORDS):
            return True

        summary_words = ["실행", "조회", "내보내기", "저장", "업데이트", "중복", "완료", "실패", "오류", "중지"]
        return any(word in message for word in summary_words)

    def _on_log_line(self, message: str):
        self.after(0, self._apply_log_line, message)

    def _apply_log_line(self, message: str):
        self._append_raw_log(message)

        if self._is_clean_log_target(message):
            self._append_clean_log(message)

    def _toggle_raw_log(self):
        self.raw_log_visible = not self.raw_log_visible

        if self.raw_log_visible:
            self.raw_log_textbox.grid()
            self.raw_log_toggle_button.configure(text="Raw Log 숨기기 ⏶")
        else:
            self.raw_log_textbox.grid_remove()
            self.raw_log_toggle_button.configure(text="Raw Log 보기 ⏷")

    def _refresh_path_labels(self):
        self.run_folder_label.configure(text=f"오늘 내보낸 파일 저장 위치: {EXPORT_DIR}")
        self.master_file_label.configure(text=f"마스터 파일: {MASTER_EXCEL_PATH}")

    def _persist_paths(self):
        try:
            save_user_path_settings(EXPORT_DIR, MASTER_EXCEL_PATH)
        except Exception as e:
            messagebox.showerror("설정 저장 실패", f"경로 설정 저장에 실패했습니다.\n{e}")
            return False
        return True

    def _open_export_folder(self):
        target = RUN_EXPORT_DIR if RUN_EXPORT_DIR else EXPORT_DIR

        if not target:
            messagebox.showwarning("경로 없음", "열 수 있는 실행 폴더가 없습니다.")
            return

        try:
            Path(target).mkdir(parents=True, exist_ok=True)
        except Exception as e:
            messagebox.showerror("폴더 열기 실패", f"저장 폴더를 준비할 수 없습니다.\n{target}\n\n{e}")
            return

        os.startfile(target)

    def _open_master_file(self):
        if not os.path.exists(MASTER_EXCEL_PATH):
            messagebox.showwarning("파일 없음", f"마스터 파일을 찾지 못했습니다.\n{MASTER_EXCEL_PATH}")
            return

        os.startfile(MASTER_EXCEL_PATH)

    def _ensure_paths_ready(self) -> bool:
        try:
            Path(EXPORT_DIR).mkdir(parents=True, exist_ok=True)
        except Exception as e:
            messagebox.showerror("저장 폴더 오류", f"저장 폴더를 준비할 수 없습니다.\n{EXPORT_DIR}\n\n{e}")
            return False

        if not MASTER_EXCEL_PATH or not os.path.isfile(MASTER_EXCEL_PATH):
            messagebox.showerror("마스터 파일 오류", f"마스터 파일을 찾지 못했습니다.\n{MASTER_EXCEL_PATH}")
            return False

        if not self._persist_paths():
            return False

        self._refresh_path_labels()
        return True

    def _load_paths_on_startup(self):
        apply_saved_path_settings()
        self._refresh_path_labels()

    def _on_run_clicked(self):
        if self.worker_thread and self.worker_thread.is_alive():
            return

        if not self._ensure_paths_ready():
            self._set_status_text("대기 중")
            return

        self.clean_log_textbox.delete("1.0", "end")
        self.raw_log_textbox.delete("1.0", "end")
        self._reset_progress_ui()

        self._set_status_text("실행 중")
        self.run_folder_label.configure(text="실행 폴더: 준비 중")

        self.run_button.configure(state="disabled")
        self.stop_button.configure(state="normal")

        register_log_listener(self._on_log_line)

        self.worker_thread = threading.Thread(target=self._run_worker, daemon=True)
        self.worker_thread.start()

    def _on_stop_clicked(self):
        request_stop()
        self._set_status_text("중지 요청됨")
        self._append_clean_log("중지 요청됨")

    def _run_worker(self):
        try:
            execute_automation_pipeline(progress_callback=self._on_progress)

            self.after(0, self._set_status_text, "완료")
            self.after(0, self._append_clean_log, "전체 작업 완료")

        except UserStopRequested:
            self.after(0, self._set_status_text, "중지됨")
            self.after(0, self._append_clean_log, "사용자 요청으로 작업 중지")

        except Exception as e:
            self.after(0, self._set_status_text, "오류")
            self.after(0, self._append_clean_log, f"오류 요약: {e}")
            self.after(0, self._append_raw_log, traceback.format_exc())

        finally:
            unregister_log_listener(self._on_log_line)

            folder_text = RUN_EXPORT_DIR if RUN_EXPORT_DIR else "-"
            self.after(0, lambda: self.run_folder_label.configure(text=f"실행 폴더: {folder_text}"))
            self.after(0, lambda: self.run_button.configure(state="normal"))
            self.after(0, lambda: self.stop_button.configure(state="disabled"))


def launch_gui():
    app = AirAutomationApp()
    app.mainloop()


if __name__ == "__main__":
    if ctk is None:
        print("[WARN] customtkinter를 찾지 못해 CLI 모드로 실행합니다. (pip install customtkinter)")
        try:
            main()
        except Exception as e:
            print("\n[ERROR] 자동화 중 오류 발생")
            print(e)
            print("\n오류 직전 로그를 확인하세요.")
        finally:
            input("\n종료하려면 Enter를 누르세요...")
    else:
        launch_gui()
