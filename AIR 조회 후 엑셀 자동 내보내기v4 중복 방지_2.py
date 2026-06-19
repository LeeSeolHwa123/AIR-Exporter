import os
import sys
import json
import re
import time
import ctypes
import shutil
import tempfile
import base64
import threading
import traceback
from ctypes import wintypes
from datetime import date, datetime, timedelta
from pathlib import Path
from tkinter import filedialog, messagebox
from zipfile import BadZipFile

from pywinauto import Desktop
from pywinauto.keyboard import send_keys

try:
    import psutil as _psutil  # type: ignore[import]
    _HAS_PSUTIL = True
except ImportError:
    _HAS_PSUTIL = False

try:
    import win32process as _win32process
    _HAS_WIN32PROCESS = True
except ImportError:
    _HAS_WIN32PROCESS = False

try:
    import customtkinter as ctk
except Exception:
    ctk = None


# =========================================================
# 0. 기본 설정
# =========================================================

CODE_VERSION = "AIR_DAILY_EXPORT_MASTER_GLOBAL_LATEST_V11_FAST_DEPARTMENT_CHANGE"
GUI_VERSION = "QEX AIR Exporter v3 (2026-06-19)"

STOP_REQUESTED = False
LOG_LISTENERS = []
RECENT_EXPORTED_EXCEL_PATHS = []


class UserStopRequested(Exception):
    pass

try:
    ctypes.windll.user32.SetProcessDPIAware()
except Exception:
    pass


# =========================================================
# 1. 사용자 설정
# =========================================================

AIR_START_SEARCH_TEXT = "newair"

CURRENT_WINDOWS_USER = Path.home().name

DEFAULT_EXPORT_DIR = rf"C:\Users\{CURRENT_WINDOWS_USER}\OneDrive - kochind.com\QIC_QEX - 98. AIR Inspection Monitoring\AIR Raw Data"
DEFAULT_MASTER_EXCEL_PATH = rf"C:\Users\{CURRENT_WINDOWS_USER}\OneDrive - kochind.com\QIC_QEX - 98. AIR Inspection Monitoring\History_202605291240.xlsx"
EXPORT_DIR = DEFAULT_EXPORT_DIR

# 이번 실행에서 내보낸 Daily 파일만 명확히 구분하기 위한 실행 단위 폴더
# 예: C:/Users/seolhl/OneDrive - kochind.com/.../AIR Raw Data/20260601_1
RUN_DATE = time.strftime("%Y%m%d")
RUN_EXPORT_DIR = None

DEPARTMENT_JOBS = [
    {
        "department": "일반사출",
        "combo_index": 3,
    },
    {
        "department": "스탬핑",
        "combo_index": 1,
    },
    {
        "department": "도금",
        "combo_index": 2,
    },
    {
        "department": "조립",
        "combo_index": 5,
    },
]

DAILY_START_TIME = "08:30" #20:30으로 설정하면 휴일 전날 아침에 20:30으로 조회했을 경우 휴일 이후에 조회 시, 휴일 전날 아침~20:30까지의 데이터 누락 위험이 있음 
                           #안전하게 무조건 08:30으로 설정하되, 이러면 중복이 발생하므로 모든 업데이트 이후 중복 삭제
DAILY_END_TIME = "20:30"

SEARCH_WAIT_SECONDS = 0.2
POST_SEARCH_STABILIZE_SECONDS = 0.6
SAVE_DIALOG_WAIT_SECONDS = 0.7

# 속도 최적화: 조회/저장 완료 팝업은 길게 전체 창 스캔하지 않고
# foreground 작은 팝업을 먼저 즉시 Enter 처리합니다.
FAST_POPUP_TIMEOUT_SECONDS = 1.0 
SEARCH_COMPLETE_TIMEOUT_SECONDS = 13
POST_EXPORT_DIALOG_WAIT_SECONDS = 0.5
FILE_CHECK_INTERVAL_SECONDS = 0.2

SEARCH_COMPLETE_KEYWORDS = [
    "조회가 완료",
    "조회 완료",
    "검색이 완료",
    "검색 완료",
    "완료되었습니다",
    "조회되었습니다",
    "조회된 데이터가 없습니다",
    "데이터가 없습니다",
]

HANDLE_SAVE_DIALOG = True

# 반복 조회 중 같은 AIR 화면 컨트롤을 매번 전체 탐색하지 않기 위한 캐시
CONTROL_CACHE = {}
LAST_AIR_PIDS: set = set()

# =========================================================
# 1-1. 마스터 엑셀 업데이트 설정
# =========================================================

# AIR에서 부서별 Daily 파일을 내보낸 뒤, 아래 마스터 파일에 자동 삽입합니다.
UPDATE_MASTER_AFTER_EXPORT = True

MASTER_EXCEL_PATH = DEFAULT_MASTER_EXCEL_PATH
MASTER_SHEET_NAME = "List"

# 마스터 파일이 열려 있으면 저장에 실패할 수 있으므로, 업데이트 전에 Excel에서 닫아두는 것을 권장합니다.
CREATE_MASTER_BACKUP = True
MASTER_BACKUP_DIR_NAME = "BackUp Files"

# 중복 기준: A열 순번 제외, B~마지막열 값이 모두 같으면 중복으로 판단합니다.
DUPLICATE_COMPARE_START_COL = 2

# 마스터 구조 기준: A열=부서별 순번, B열=부서명
MASTER_INDEX_COL = 1
MASTER_DEPARTMENT_COL = 2

# AIR 조회기간 자동 산정 기준
# 마스터 파일을 먼저 열어서 전체 부서 중 '최근검사' 열의 가장 최근 날짜를 From으로 사용합니다.
MASTER_RECENT_INSPECTION_HEADER = "최근검사"
USE_MASTER_LATEST_RECENT_DATE_FOR_FROM = True
QUERY_START_FALLBACK_DAYS = 1

# True: 신규 Daily 데이터는 마스터 전체 마지막 행 아래에 추가합니다.
# False: 기존처럼 해당 부서의 마지막 행 바로 아래에 삽입합니다.
APPEND_NEW_ROWS_TO_BOTTOM = True


def get_settings_file_path() -> Path:
    app_data = os.getenv("APPDATA")
    if app_data:
        base_dir = Path(app_data)
    else:
        base_dir = Path.home()
    return base_dir / "QEXAirExporter" / "settings.json"


def load_user_path_settings() -> dict:
    settings_path = get_settings_file_path()

    if not settings_path.exists():
        return {}

    try:
        with settings_path.open("r", encoding="utf-8") as fp:
            data = json.load(fp)

        if not isinstance(data, dict):
            return {}

        return data
    except Exception as e:
        print(f"[WARN] 사용자 설정 파일 로드 실패: {settings_path} / {e}", flush=True)
        return {}


def remap_path_to_current_windows_user(path: str) -> str:
    """
    저장된 경로의 사용자명(C:\\Users\\<name>)만 현재 사용자로 치환합니다.
    나머지 경로는 그대로 유지합니다.
    """
    value = str(path or "").strip()
    if not value:
        return value

    match = re.match(r"^([A-Za-z]:\\Users\\)([^\\]+)(\\.*)$", value, re.IGNORECASE)
    if not match:
        return value

    return f"{match.group(1)}{CURRENT_WINDOWS_USER}{match.group(3)}"


def normalize_user_path(path: str) -> str:
    """
    사용자가 입력한 경로를 비교/검사하기 좋게 정리합니다.
    - 앞뒤 공백 제거
    - 바깥따옴표 제거
    - 환경변수/사용자 홈 경로 확장
    - Windows 경로 정규화
    """
    value = str(path or "").strip().strip('"').strip("'")

    if not value:
        return ""

    value = os.path.expandvars(os.path.expanduser(value))
    return os.path.normpath(value)


def normalize_master_excel_path(path: str) -> str:
    """
    저장된 마스터 파일 경로를 현재 사용자 기준으로 정규화합니다.
    값이 비어 있으면 기본 경로를 사용합니다.
    """
    value = normalize_user_path(remap_path_to_current_windows_user(path))
    if value:
        return value
    return DEFAULT_MASTER_EXCEL_PATH


def get_master_backup_dir(master_path: Path) -> Path:
    """마스터 파일과 같은 OneDrive 폴더 아래에 백업 폴더를 생성합니다."""
    return master_path.parent / MASTER_BACKUP_DIR_NAME


def save_user_path_settings(export_dir: str, master_excel_path: str):
    settings_path = get_settings_file_path()
    settings_path.parent.mkdir(parents=True, exist_ok=True)

    # 기존 설정을 읽어서 자격증명 등 다른 키를 보존합니다.
    existing = {}
    if settings_path.exists():
        try:
            with settings_path.open("r", encoding="utf-8") as fp:
                existing = json.load(fp)
        except Exception:
            pass

    existing["export_dir"] = str(export_dir or "").strip()
    existing["master_excel_path"] = str(master_excel_path or "").strip()

    with settings_path.open("w", encoding="utf-8") as fp:
        json.dump(existing, fp, ensure_ascii=False, indent=2)


# =========================================================
# 1-0. AIR 자격증명 저장 / 로드
# =========================================================

_CRED_SERVICE = "QEXAirExporter"


def save_air_credentials(user_id: str, password: str):
    """AIR 로그인 자격증명을 Windows Credential Manager(keyring)에 저장합니다."""
    try:
        import keyring  # type: ignore[import]
        keyring.set_password(_CRED_SERVICE, "user_id", user_id)
        keyring.set_password(_CRED_SERVICE, "password", password)
        log("[OK] 자격증명을 Windows Credential Manager에 저장했습니다.")
        return
    except Exception:
        pass
    # fallback: settings.json에 base64로 저장
    settings_path = get_settings_file_path()
    existing = {}
    if settings_path.exists():
        try:
            with settings_path.open("r", encoding="utf-8") as fp:
                existing = json.load(fp)
        except Exception:
            pass
    existing["air_user_id"] = base64.b64encode(user_id.encode("utf-8")).decode()
    existing["air_password"] = base64.b64encode(password.encode("utf-8")).decode()
    settings_path.parent.mkdir(parents=True, exist_ok=True)
    with settings_path.open("w", encoding="utf-8") as fp:
        json.dump(existing, fp, ensure_ascii=False, indent=2)
    log("[OK] 자격증명을 설정 파일에 저장했습니다.")


def load_air_credentials() -> tuple:
    """저장된 AIR 자격증명을 로드합니다. (user_id, password) 반환"""
    try:
        import keyring  # type: ignore[import]
        uid = keyring.get_password(_CRED_SERVICE, "user_id") or ""
        pw = keyring.get_password(_CRED_SERVICE, "password") or ""
        if uid:
            return uid, pw
    except Exception:
        pass
    # fallback: settings.json
    settings_path = get_settings_file_path()
    if settings_path.exists():
        try:
            with settings_path.open("r", encoding="utf-8") as fp:
                data = json.load(fp)
            raw_uid = data.get("air_user_id", "")
            raw_pw = data.get("air_password", "")
            uid = base64.b64decode(raw_uid.encode()).decode("utf-8") if raw_uid else ""
            pw = base64.b64decode(raw_pw.encode()).decode("utf-8") if raw_pw else ""
            return uid, pw
        except Exception:
            pass
    return "", ""


def apply_saved_path_settings():
    global EXPORT_DIR
    global MASTER_EXCEL_PATH

    settings = load_user_path_settings()

    export_dir = remap_path_to_current_windows_user(settings.get("export_dir", ""))
    master_excel_path = normalize_master_excel_path(settings.get("master_excel_path", ""))

    if export_dir:
        EXPORT_DIR = export_dir

    if master_excel_path:
        MASTER_EXCEL_PATH = master_excel_path

    # 경로 정규화 결과를 설정 파일에 반영합니다.
    try:
        save_user_path_settings(EXPORT_DIR, MASTER_EXCEL_PATH)
    except Exception:
        pass


apply_saved_path_settings()


# =========================================================
# 2. AIR 컨트롤 설정
# =========================================================

AIR_TITLE_PATTERNS = [
    r".*CCS AIR System.*",
    r".*OPC AIR System.*",
    r".*검사기록조회.*",
]

INSPECTION_BUTTON_AUTO_ID = "button6"
INSPECTION_BUTTON_TITLE = "검사이력"

INSPECTION_WINDOW_KEYWORD = "검사기록조회"

INSPECTION_FORM_AUTO_ID = "frmSetupList_1"
INSPECTION_FORM_TITLE = "검사기록조회"

PLANT_COMBO_AUTO_ID = "cboPlant"
DEPARTMENT_COMBO_AUTO_ID = "cboCostCenter"

START_DATE_PICKER_AUTO_ID = "dFrom"
END_DATE_PICKER_AUTO_ID = "dTo"

START_TIME_COMBO_AUTO_ID = "cboStart"
END_TIME_COMBO_AUTO_ID = "cboEnd"

CONDITION_PANEL_AUTO_ID = "panel1"
PERIOD_GROUP_AUTO_ID = "gbPeroid"

INSPECTION_SEARCH_BUTTON_AUTO_ID = "button1"

EXCEL_BUTTON_AUTO_ID = "cmdExcel"

FIELD_AUTO_IDS = {
    "제품번호": "txtPNo",
    "기계번호": "txtmcno",
    "Lot No": "txtLotNo",
    "Barcode": "txtBarcode",
    "툴번호": "txtToolNo",
    "PO No": "txtPOno",
    "판정": "txtDisposition",
}


# =========================================================
# 3. Win32 ComboBox 메시지 상수
# =========================================================
CB_GETCOUNT = 0x0146
CB_GETCURSEL = 0x0147
CB_GETLBTEXT = 0x0148
CB_GETLBTEXTLEN = 0x0149
CB_SETCURSEL = 0x014E

WM_COMMAND = 0x0111
CBN_SELCHANGE = 1
CBN_SELENDOK = 9

DTM_FIRST = 0x1000
DTM_SETSYSTEMTIME = DTM_FIRST + 2
GDT_VALID = 0


class SYSTEMTIME(ctypes.Structure):
    _fields_ = [
        ("wYear", wintypes.WORD),
        ("wMonth", wintypes.WORD),
        ("wDayOfWeek", wintypes.WORD),
        ("wDay", wintypes.WORD),
        ("wHour", wintypes.WORD),
        ("wMinute", wintypes.WORD),
        ("wSecond", wintypes.WORD),
        ("wMilliseconds", wintypes.WORD),
    ]

WPARAM_T = ctypes.c_size_t
LPARAM_T = ctypes.c_ssize_t
LRESULT_T = ctypes.c_ssize_t

_SendMessageW = ctypes.windll.user32.SendMessageW
_SendMessageW.argtypes = [wintypes.HWND, wintypes.UINT, WPARAM_T, LPARAM_T]
_SendMessageW.restype = LRESULT_T


# =========================================================
# 4. 공통 유틸
# =========================================================

def log(msg: str):
    message = str(msg)

    try:
        if getattr(sys, "stdout", None):
            print(message, flush=True)
    except Exception:
        pass

    for listener in list(LOG_LISTENERS):
        try:
            listener(message)
        except Exception:
            pass


def register_log_listener(listener):
    if listener not in LOG_LISTENERS:
        LOG_LISTENERS.append(listener)


def unregister_log_listener(listener):
    if listener in LOG_LISTENERS:
        LOG_LISTENERS.remove(listener)


def request_stop():
    global STOP_REQUESTED
    STOP_REQUESTED = True


def ensure_not_stopped():
    if STOP_REQUESTED:
        raise UserStopRequested("사용자 중지 요청")


def normalize_text(text: str) -> str:
    return str(text).strip().replace(" ", "")


def set_clipboard_text(text: str):
    """
    Set clipboard text without creating a second Tk root.
    Creating/destroying tkinter.Tk() from the worker thread can close or destabilize
    the customtkinter GUI in a PyInstaller windowed executable.
    """
    value = str(text)
    user32 = ctypes.windll.user32
    kernel32 = ctypes.windll.kernel32

    GMEM_MOVEABLE = 0x0002
    CF_UNICODETEXT = 13

    kernel32.GlobalAlloc.restype = wintypes.HGLOBAL
    kernel32.GlobalLock.restype = wintypes.LPVOID
    user32.SetClipboardData.restype = wintypes.HANDLE

    data = value.encode("utf-16le") + b"\x00\x00"
    handle = kernel32.GlobalAlloc(GMEM_MOVEABLE, len(data))
    if not handle:
        raise RuntimeError("클립보드 메모리 할당 실패")

    locked = kernel32.GlobalLock(handle)
    if not locked:
        kernel32.GlobalFree(handle)
        raise RuntimeError("클립보드 메모리 잠금 실패")

    try:
        ctypes.memmove(locked, data, len(data))
    finally:
        kernel32.GlobalUnlock(handle)

    if not user32.OpenClipboard(None):
        kernel32.GlobalFree(handle)
        raise RuntimeError("클립보드를 열 수 없습니다.")

    try:
        user32.EmptyClipboard()
        if not user32.SetClipboardData(CF_UNICODETEXT, handle):
            kernel32.GlobalFree(handle)
            raise RuntimeError("클립보드 데이터 설정 실패")
        handle = None
    finally:
        user32.CloseClipboard()


def parse_date_ymd(value: str):
    value = str(value).strip()

    if "-" in value:
        parts = value.split("-")
    elif "/" in value:
        parts = value.split("/")
    elif "." in value:
        parts = value.split(".")
    else:
        value = value.replace(" ", "")
        if len(value) != 8:
            raise ValueError(f"날짜 형식 오류: {value}")
        return value[0:4], value[4:6], value[6:8]

    if len(parts) != 3:
        raise ValueError(f"날짜 형식 오류: {value}")

    year = parts[0].strip()
    if len(year) == 2:
        year = f"20{year}"
    else:
        year = year.zfill(4)

    return year, parts[1].zfill(2), parts[2].zfill(2)


def create_run_export_dir() -> str:
    """
    오늘 날짜 기준 실행 폴더를 생성합니다.

    예:
    C:/Users/seolhl/OneDrive - kochind.com/.../AIR Raw Data/20260601_1
    C:/Users/seolhl/OneDrive - kochind.com/.../AIR Raw Data/20260601_2

    같은 날짜 폴더가 이미 있으면 뒤의 실행횟수를 1씩 증가시킵니다.
    """
    base_dir = Path(EXPORT_DIR)
    base_dir.mkdir(parents=True, exist_ok=True)

    run_no = 1

    while True:
        candidate = base_dir / f"{RUN_DATE}_{run_no}"

        if not candidate.exists():
            candidate.mkdir(parents=True, exist_ok=False)

            log("[RUN-FOLDER] 이번 실행 저장 폴더 생성 완료")
            log(f" - 기준 폴더: {base_dir}")
            log(f" - 실행 날짜: {RUN_DATE}")
            log(f" - 실행 횟수: {run_no}")
            log(f" - 저장 폴더: {candidate}")

            return str(candidate)

        run_no += 1


def make_export_path(department: str) -> str:
    """
    이번 실행 폴더 안에 부서별 Daily 엑셀 저장 경로를 생성합니다.

    예:
    C:/Users/seolhl/OneDrive - kochind.com/.../AIR Raw Data/20260601_1/AIR_검사이력_일반사출_20260601_171000.xlsx
    """
    global RUN_EXPORT_DIR

    if RUN_EXPORT_DIR is None:
        RUN_EXPORT_DIR = create_run_export_dir()

    Path(RUN_EXPORT_DIR).mkdir(parents=True, exist_ok=True)

    safe_dept = re.sub(r'[\\/:*?"<>|]', "_", department)
    filename = f"AIR_검사이력_{safe_dept}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    export_path = os.path.join(RUN_EXPORT_DIR, filename)

    log("[EXPORT-PATH] 부서별 저장 경로 생성")
    log(f" - 부서명: {department}")
    log(f" - 저장 경로: {export_path}")

    return export_path


def wait_until_file_exists(path: str, timeout: int = 30, interval: float = None) -> bool:
    """
    파일 생성 확인.
    기존 1초 단위 polling은 저장 후 대기 시간이 길어져서 0.2초 단위로 줄입니다.
    """
    if interval is None:
        interval = FILE_CHECK_INTERVAL_SECONDS

    start = time.time()

    while time.time() - start < timeout:
        if os.path.exists(path) and os.path.getsize(path) > 0:
            return True
        time.sleep(interval)

    return False


def _check_excel_file_ready(path: str, require_writable: bool = True) -> tuple[bool, str]:
    """
    xlsx 파일이 완전히 저장되어 openpyxl로 읽을 수 있는지 확인합니다.
    파일 존재/크기만으로는 zip 구조가 완성됐다고 볼 수 없으므로 실제 workbook 로드를 확인합니다.
    """
    file_path = Path(path)

    if not file_path.exists():
        return False, "파일 없음"

    try:
        first_size = file_path.stat().st_size
    except OSError as e:
        return False, f"파일 크기 확인 실패: {e}"

    if first_size <= 0:
        return False, "파일 크기 0"

    time.sleep(max(0.2, FILE_CHECK_INTERVAL_SECONDS))

    try:
        second_size = file_path.stat().st_size
    except OSError as e:
        return False, f"파일 크기 재확인 실패: {e}"

    if second_size <= 0:
        return False, "파일 크기 0"

    if first_size != second_size:
        return False, f"파일 저장 중(size {first_size} -> {second_size})"

    if require_writable:
        try:
            with open(file_path, "r+b"):
                pass
        except (PermissionError, OSError) as e:
            return False, f"파일 잠금/권한 대기: {e}"

    try:
        patch_openpyxl_print_page_setup_collated()
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        wb.close()
        return True, "OK"
    except BadZipFile as e:
        return False, f"xlsx zip 구조 미완성: {e}"
    except Exception as e:
        return False, f"openpyxl 로드 대기: {e}"


def wait_until_excel_file_ready(
    path: str,
    timeout: int = 45,
    interval: float = None,
    require_writable: bool = True,
    label: str = "엑셀 파일",
) -> bool:
    """
    xlsx 저장 완료를 파일 생성이 아니라 실제 openpyxl 로드 가능 상태로 판정합니다.
    AIR/Excel/OneDrive가 파일을 쓰는 중이면 BadZipFile이 날 수 있어 재시도합니다.
    """
    if interval is None:
        interval = FILE_CHECK_INTERVAL_SECONDS

    start = time.time()
    last_log = 0
    last_reason = ""

    while time.time() - start < timeout:
        ready, reason = _check_excel_file_ready(path, require_writable=require_writable)

        if ready:
            log(f"[OK] {label} 준비 완료: {path}")
            return True

        last_reason = reason

        if time.time() - last_log >= 2:
            log(f"[INFO] {label} 저장 완료 대기 중: {reason}")
            log(f"       {path}")
            last_log = time.time()

        time.sleep(interval)

    log(f"[WARN] {label} 준비 확인 timeout: {path} / 마지막 상태: {last_reason}")
    return False


def load_workbook_after_ready(
    path: str,
    data_only: bool = True,
    read_only: bool = False,
    timeout: int = 45,
    label: str = "엑셀 파일",
    require_writable: bool = False,
):
    if not wait_until_excel_file_ready(
        path,
        timeout=timeout,
        require_writable=require_writable,
        label=label,
    ):
        raise RuntimeError(f"{label}이 아직 완전히 저장되지 않았거나 읽을 수 없습니다: {path}")

    from openpyxl import load_workbook

    last_error = None

    for attempt in range(1, 6):
        try:
            return load_workbook(path, data_only=data_only, read_only=read_only)
        except (BadZipFile, PermissionError, OSError) as e:
            last_error = e
            log(f"[WARN] {label} 로드 재시도({attempt}/5): {e}")
            time.sleep(0.5)

    raise RuntimeError(f"{label} 로드 실패: {path} / 원인: {last_error}")


def normalize_excel_file_path(path: str) -> str:
    """
    Excel COM의 Workbook.FullName과 비교하기 위한 경로 정규화.
    대소문자/슬래시 차이 때문에 같은 파일을 못 찾는 문제를 줄입니다.
    """
    return os.path.normcase(os.path.abspath(str(path)))


def extract_export_paths(export_items):
    """
    saved_exports 구조에서 실제 파일 경로만 추출합니다.

    허용 형태:
    1. ["C:\\...\\AIR_검사이력_일반사출.xlsx", ...]
    2. [{"department": "일반사출", "path": "C:\\...xlsx"}, ...]
    """
    paths = []

    for item in export_items:
        if isinstance(item, dict):
            path = (
                item.get("path")
                or item.get("file_path")
                or item.get("saved_path")
                or item.get("export_path")
            )
        else:
            path = item

        if not path:
            continue

        path = str(path)

        if os.path.exists(path):
            paths.append(path)
        else:
            log(f"[WARN] 내보낸 파일 경로가 존재하지 않아 닫기 대상에서 제외: {path}")

    return paths


def is_file_unlocked(path: str) -> bool:
    """
    파일이 Excel/OneDrive 등에 의해 잠겨 있지 않고 openpyxl로 실제 읽을 수 있는지 확인합니다.
    """
    ready, reason = _check_excel_file_ready(path, require_writable=True)
    if not ready:
        log(f"[INFO] 내보낸 엑셀 파일이 아직 읽기 가능한 상태가 아닙니다: {path} / {reason}")
        return False
    return True


def wait_until_export_files_unlocked(export_items, timeout: int = 10) -> bool:
    """
    내보낸 엑셀 파일들이 닫혀서 잠금 해제될 때까지 대기합니다.
    """
    paths = extract_export_paths(export_items)

    if not paths:
        return True

    start = time.time()

    while time.time() - start < timeout:
        locked = []

        for path in paths:
            if not is_file_unlocked(path):
                locked.append(path)

        if not locked:
            log("[OK] 내보낸 엑셀 파일 잠금 해제 확인 완료")
            return True

        log("[INFO] 아직 열려 있는 내보낸 엑셀 파일 대기 중:")
        for path in locked:
            log(f" - {path}")

        time.sleep(0.5)

    log("[WARN] 일부 내보낸 엑셀 파일 잠금이 해제되지 않았습니다.")
    return False


def close_only_exported_excel_workbooks(
    export_items,
    timeout: int = 20,
    wait_for_excel_appearance: bool = False,
) -> bool:
    """
    Excel에서 열려 있는 워크북 중 이번 자동화로 내보낸 파일만 닫습니다.

    개선점:
    1. Excel COM Workbook.FullName 전체 경로 비교
    2. 경로 비교 실패 시 파일명 기준 fallback 비교
    3. 자동으로 열린 Excel이 늦게 COM에 등록될 수 있으므로 timeout 동안 반복 시도
    4. 닫기 실패 시 False 반환
    """
    paths = extract_export_paths(export_items)

    if not paths:
        log("[INFO] 닫을 내보낸 엑셀 파일이 없습니다.")
        return True

    target_paths = {
        normalize_excel_file_path(path): path
        for path in paths
    }

    target_file_names = {
        os.path.basename(path).lower(): path
        for path in paths
    }

    log("\n[EXCEL-CLOSE] 마스터 업데이트 전, 자동으로 열린 내보낸 엑셀 파일만 닫기 시작")
    for path in paths:
        log(f" - 닫기 대상: {path}")

    try:
        import win32com.client
    except ImportError:
        log("[WARN] pywin32가 설치되어 있지 않아 Excel 창 자동 닫기를 건너뜁니다.")
        log("       설치 명령: pip install pywin32")
        return False

    start = time.time()
    wait_until = start + (min(timeout, 6) if wait_for_excel_appearance else 0)
    total_closed_count = 0
    excel_app = None

    while time.time() - start < timeout:
        try:
            excel_app = win32com.client.GetActiveObject("Excel.Application")
        except Exception:
            log("[INFO] 실행 중인 Excel COM 인스턴스를 찾지 못했습니다.")
            if wait_for_excel_appearance and time.time() < wait_until:
                log("[INFO] 저장 완료 후 Excel 자동 실행 대기 중...")
                time.sleep(0.5)
                continue
            return wait_until_export_files_unlocked(paths, timeout=2)

        try:
            workbooks = excel_app.Workbooks
            workbook_count = workbooks.Count

            log(f"[EXCEL-CLOSE] 현재 Excel COM Workbook 수: {workbook_count}")

            closed_this_round = 0

            for i in range(workbook_count, 0, -1):
                wb = workbooks.Item(i)

                try:
                    wb_fullname = str(wb.FullName)
                except Exception:
                    continue

                normalized_wb_path = normalize_excel_file_path(wb_fullname)
                wb_file_name = os.path.basename(wb_fullname).lower()

                log(f"[EXCEL-CHECK] 열려 있는 Workbook: {wb_fullname}")

                matched = False

                # 1차: 전체 경로 정확 비교
                if normalized_wb_path in target_paths:
                    matched = True

                # 2차: 파일명 fallback
                # AIR export 파일명은 초 단위 timestamp가 들어가므로 파일명만으로도 충돌 가능성이 낮음
                elif wb_file_name in target_file_names:
                    matched = True

                if matched:
                    log(f"[EXCEL-CLOSE] 내보낸 파일 닫기: {wb_fullname}")
                    wb.Close(SaveChanges=False)
                    closed_this_round += 1
                    total_closed_count += 1
                    time.sleep(0.3)

            if closed_this_round > 0:
                log(f"[EXCEL-CLOSE] 이번 반복에서 닫은 파일 수: {closed_this_round}")

        except Exception as e:
            log(f"[WARN] Excel Workbook 닫기 중 오류: {e}")

        # 닫힌 뒤 실제 파일 잠금이 풀렸는지 확인
        if wait_until_export_files_unlocked(paths, timeout=2):
            if wait_for_excel_appearance and total_closed_count <= 0 and time.time() < wait_until:
                log("[INFO] 내보낸 파일은 준비됐지만 Excel 자동 실행 여부를 조금 더 확인합니다.")
                time.sleep(0.5)
                continue

            log(f"[EXCEL-CLOSE] 닫은 내보낸 엑셀 파일 총 수: {total_closed_count}")
            quit_excel_if_empty(excel_app, total_closed_count)
            return True

        log("[INFO] Excel 자동 실행/잠금 해제를 기다린 뒤 닫기 재시도")
        time.sleep(0.5)

    log(f"[EXCEL-CLOSE] 닫은 내보낸 엑셀 파일 총 수: {total_closed_count}")
    log("[WARN] timeout 동안 내보낸 엑셀 파일 잠금이 해제되지 않았습니다.")
    return False


def quit_excel_if_empty(excel_app, closed_count: int = 0) -> bool:
    """
    이번 자동화가 내보낸 워크북을 닫은 뒤 빈 Excel 화면만 남으면 종료합니다.
    다른 Workbook이 열려 있으면 사용자가 작업 중인 Excel로 보고 유지합니다.
    """
    if excel_app is None or closed_count <= 0:
        return False

    try:
        remaining_count = excel_app.Workbooks.Count
    except Exception:
        return False

    if remaining_count > 0:
        log(f"[EXCEL-CLOSE] 다른 Excel Workbook {remaining_count}개가 열려 있어 Excel 앱은 유지합니다.")
        return False

    try:
        excel_app.DisplayAlerts = False
    except Exception:
        pass

    try:
        excel_app.Quit()
        log("[EXCEL-CLOSE] 빈 Excel 잔류 화면 종료 완료")
        return True
    except Exception as e:
        log(f"[WARN] 빈 Excel 앱 종료 실패: {e}")
        return False


def patch_openpyxl_print_page_setup_collated():
    """
    AIR가 만든 xlsx에 pageSetup collated 속성이 들어 있으면 openpyxl에서 아래 오류가 날 수 있습니다.

    PrintPageSetup.__init__() got an unexpected keyword argument 'collated'

    load_workbook() 전에 이 함수를 실행해서 openpyxl이 모르는 collated 속성만 무시하게 합니다.
    """
    try:
        from openpyxl.worksheet.page import PrintPageSetup

        original_init = PrintPageSetup.__init__

        if getattr(original_init, "_air_collated_patch", False):
            return

        def patched_init(self, *args, **kwargs):
            kwargs.pop("collated", None)
            return original_init(self, *args, **kwargs)

        patched_init._air_collated_patch = True
        PrintPageSetup.__init__ = patched_init

        log("[OK] openpyxl PrintPageSetup collated 무시 패치 적용 완료")

    except Exception as e:
        log(f"[WARN] openpyxl collated 패치 적용 실패: {e}")


def load_workbook_resilient(path: str, data_only: bool = True):
    """
    OneDrive/Excel 잠금 이슈를 줄이기 위해 workbook 로드를 재시도하고,
    직접 열기에 실패하면 임시 복사본으로 읽습니다.

    반환값: (workbook, temp_copy_path)
    - temp_copy_path가 None이 아니면 사용 후 삭제가 필요합니다.
    """
    from openpyxl import load_workbook

    workbook_path = Path(path)
    last_error = None

    for attempt in range(1, 4):
        try:
            wb = load_workbook(workbook_path, data_only=data_only)
            return wb, None
        except PermissionError as e:
            last_error = e
            log(f"[WARN] 마스터 파일 직접 열기 권한 오류(시도 {attempt}/3): {e}")
        except OSError as e:
            last_error = e
            log(f"[WARN] 마스터 파일 직접 열기 OS 오류(시도 {attempt}/3): {e}")

        time.sleep(0.4)

    temp_copy_path = None

    try:
        tmp_dir = Path(tempfile.gettempdir()) / "QEXAirExporter"
        tmp_dir.mkdir(parents=True, exist_ok=True)

        temp_copy_path = tmp_dir / f"{workbook_path.stem}_tmp_{int(time.time() * 1000)}{workbook_path.suffix}"
        shutil.copy2(workbook_path, temp_copy_path)

        log(f"[INFO] 임시 복사본으로 마스터 파일 로드 시도: {temp_copy_path}")
        wb = load_workbook(temp_copy_path, data_only=data_only)
        return wb, str(temp_copy_path)

    except Exception as e:
        last_error = e

    raise RuntimeError(
        "마스터 엑셀 파일 로드에 실패했습니다. "
        "파일 잠금/권한/손상 여부를 확인하세요. "
        f"대상 파일: {workbook_path} / 원인: {last_error}"
    )


def close_workbook_and_cleanup_temp(wb, temp_copy_path: str = None):
    try:
        if wb is not None:
            wb.close()
    finally:
        if temp_copy_path:
            try:
                if os.path.exists(temp_copy_path):
                    os.remove(temp_copy_path)
            except Exception:
                pass


def snapshot_excel_files():
    """
    이번 실행 폴더 안의 엑셀 파일 상태만 기록합니다.
    과거 실행 폴더나 다른 폴더는 보지 않습니다.
    """
    global RUN_EXPORT_DIR

    result = {}

    if RUN_EXPORT_DIR is None:
        return result

    folder = Path(RUN_EXPORT_DIR)

    if not folder.exists():
        return result

    for ext in ("*.xlsx", "*.xls"):
        for file in folder.glob(ext):
            try:
                result[str(file.resolve())] = file.stat().st_mtime
            except Exception:
                pass

    return result


def wait_for_new_excel_file(before_snapshot: dict, timeout: int = 30):
    """
    이번 실행 폴더 안에서 새로 생성된 엑셀 파일만 찾습니다.
    과거 Daily 파일이나 다른 폴더의 엑셀 파일은 무시합니다.
    """
    global RUN_EXPORT_DIR

    if RUN_EXPORT_DIR is None:
        return None

    folder = Path(RUN_EXPORT_DIR)

    if not folder.exists():
        return None

    start = time.time()

    while time.time() - start < timeout:
        candidates = []

        for ext in ("*.xlsx", "*.xls"):
            for file in folder.glob(ext):
                try:
                    full = str(file.resolve())
                    mtime = file.stat().st_mtime
                    size = file.stat().st_size

                    if size <= 0:
                        continue

                    if full not in before_snapshot:
                        candidates.append(file)
                    elif mtime > before_snapshot[full] + 0.5:
                        candidates.append(file)
                except Exception:
                    pass

        if candidates:
            candidates.sort(key=lambda f: f.stat().st_mtime, reverse=True)
            return str(candidates[0].resolve())

        time.sleep(FILE_CHECK_INTERVAL_SECONDS)

    return None


def copy_or_rename_export_file(src: str, dst: str):
    src_path = Path(src)
    dst_path = Path(dst)

    if src_path.resolve() == dst_path.resolve():
        return str(dst_path)

    dst_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        shutil.copy2(src_path, dst_path)
        return str(dst_path)
    except Exception:
        return str(src_path)


def is_real_visible_control(ctrl) -> bool:
    try:
        rect = ctrl.rectangle()

        if rect.width() <= 0 or rect.height() <= 0:
            return False

        if rect.left < -30000 or rect.top < -30000:
            return False

        if not ctrl.is_visible():
            return False

        if not ctrl.is_enabled():
            return False

        return True

    except Exception:
        return False


def is_inside_screen(x: int, y: int) -> bool:
    try:
        SM_XVIRTUALSCREEN = 76
        SM_YVIRTUALSCREEN = 77
        SM_CXVIRTUALSCREEN = 78
        SM_CYVIRTUALSCREEN = 79

        left = ctypes.windll.user32.GetSystemMetrics(SM_XVIRTUALSCREEN)
        top = ctypes.windll.user32.GetSystemMetrics(SM_YVIRTUALSCREEN)
        width = ctypes.windll.user32.GetSystemMetrics(SM_CXVIRTUALSCREEN)
        height = ctypes.windll.user32.GetSystemMetrics(SM_CYVIRTUALSCREEN)

        return left <= x <= left + width and top <= y <= top + height
    except Exception:
        return True


def get_control_rect_tuple_safe(ctrl, label: str = "control"):
    try:
        rect = ctrl.rectangle()
        return (int(rect.left), int(rect.top), int(rect.right), int(rect.bottom), rect)
    except Exception as e:
        log(f"[WARN] {label} UIA 좌표 읽기 실패, Win32 좌표로 재시도: {e}")

    hwnd = get_control_hwnd(ctrl)
    if not hwnd:
        return None

    try:
        rect = wintypes.RECT()
        if ctypes.windll.user32.GetWindowRect(wintypes.HWND(int(hwnd)), ctypes.byref(rect)):
            return (int(rect.left), int(rect.top), int(rect.right), int(rect.bottom), rect)
    except Exception as e:
        log(f"[WARN] {label} Win32 좌표 읽기 실패: {e}")

    return None


def real_mouse_click(x: int, y: int):
    ctypes.windll.user32.SetCursorPos(x, y)
    time.sleep(0.05)

    MOUSEEVENTF_LEFTDOWN = 0x0002
    MOUSEEVENTF_LEFTUP = 0x0004

    ctypes.windll.user32.mouse_event(MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    time.sleep(0.05)
    ctypes.windll.user32.mouse_event(MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    time.sleep(0.15)


def click_control_by_real_mouse(ctrl, label: str):
    rect = ctrl.rectangle()
    x = int((rect.left + rect.right) / 2)
    y = int((rect.top + rect.bottom) / 2)

    log(f"[INFO] {label} 실제 클릭: x={x}, y={y}, rect={rect}")

    if not is_inside_screen(x, y):
        raise RuntimeError(f"{label} 클릭 좌표가 화면 범위 밖입니다: ({x}, {y})")

    try:
        top = ctrl.top_level_parent()
        top.set_focus()
        time.sleep(0.1)
    except Exception:
        pass

    real_mouse_click(x, y)


def click_control_fast(ctrl, label: str):
    """
    이미 화면에 보이는 AIR 내부 버튼을 빠르게 클릭합니다.
    top-level focus 재설정 대기 없이 좌표만 검증한 뒤 바로 클릭합니다.
    """
    rect = ctrl.rectangle()
    x = int((rect.left + rect.right) / 2)
    y = int((rect.top + rect.bottom) / 2)

    if not is_inside_screen(x, y):
        raise RuntimeError(f"{label} 클릭 좌표가 화면 범위 밖입니다: ({x}, {y})")

    real_mouse_click(x, y)


def get_foreground_window_info():
    hwnd = ctypes.windll.user32.GetForegroundWindow()

    length = ctypes.windll.user32.GetWindowTextLengthW(hwnd)
    buff = ctypes.create_unicode_buffer(length + 1)
    ctypes.windll.user32.GetWindowTextW(hwnd, buff, length + 1)

    return hwnd, buff.value


# =========================================================
# 5. AIR 창 찾기 / 진입
# =========================================================

# =========================================================
# 5-A. AIR PID 기반 팝업 필터링 헬퍼
# =========================================================

def _snapshot_pids() -> set:
    """현재 실행 중인 모든 프로세스 PID 스냅샷을 반환합니다."""
    if not _HAS_PSUTIL:
        return set()
    pids = set()
    try:
        for proc in _psutil.process_iter(['pid']):
            try:
                pids.add(proc.info['pid'])
            except Exception:
                pass
    except Exception:
        pass
    return pids


def _collect_air_pids(before_pids: set) -> set:
    """
    AIR 실행 후 생성된 프로세스 PID를 수집합니다.
    - 이름에 AIR 관련 키워드 포함
    - AIR 창 제목과 매칭되는 윈도우의 PID
    - 위 조건을 만족하는 프로세스의 자식 프로세스
    """
    if not _HAS_PSUTIL:
        return set()

    air_keywords = ["air", "newair", "ccs", "opc"]
    air_pids: set = set()

    try:
        for proc in _psutil.process_iter(['pid', 'name']):
            try:
                pid = proc.info['pid']
                if pid == os.getpid():
                    continue
                name = (proc.info['name'] or "").lower()
                if any(kw in name for kw in air_keywords):
                    air_pids.add(pid)
                    try:
                        for child in proc.children(recursive=True):
                            if child.pid != os.getpid():
                                air_pids.add(child.pid)
                    except Exception:
                        pass
            except (_psutil.NoSuchProcess, _psutil.AccessDenied):
                pass
    except Exception:
        pass

    if _HAS_WIN32PROCESS:
        try:
            for w in Desktop(backend="uia").windows():
                try:
                    title = w.window_text().strip()
                    if not title:
                        continue
                    hwnd = int(w.handle)
                    if is_own_app_window(hwnd, title):
                        continue
                    if any(re.search(p, title) for p in AIR_TITLE_PATTERNS):
                        pid = _get_window_pid(hwnd)
                        if pid is not None and pid != os.getpid():
                            air_pids.add(pid)
                except Exception:
                    pass
        except Exception:
            pass

    global LAST_AIR_PIDS
    if air_pids:
        LAST_AIR_PIDS = set(air_pids)

    log(f"[INFO] AIR 관련 PID {len(air_pids)}개 수집")
    return air_pids


def _get_window_pid(hwnd: int) -> int | None:
    """창 핸들로 해당 창을 소유한 프로세스의 PID를 반환합니다."""
    if not _HAS_WIN32PROCESS:
        return None
    try:
        _, pid = _win32process.GetWindowThreadProcessId(hwnd)
        return pid
    except Exception:
        return None


def is_own_app_window(hwnd: int, title: str = "") -> bool:
    """Return True for this exporter's own GUI/window so popup handlers never close it."""
    try:
        pid = _get_window_pid(hwnd)
        if pid == os.getpid():
            return True
    except Exception:
        pass

    title_lower = (title or "").lower()
    own_title_keywords = [
        "air 검사이력 자동 내보내기",
        "qex air exporter",
        "qex_air_exporter",
    ]
    return any(keyword in title_lower for keyword in own_title_keywords)


def is_air_owned_window(hwnd: int, title: str = "", air_pids: set | None = None) -> bool:
    """Only allow popup actions on windows that plausibly belong to AIR, never this GUI."""
    if is_own_app_window(hwnd, title):
        return False

    if air_pids and _HAS_WIN32PROCESS:
        pid = _get_window_pid(hwnd)
        return pid is not None and pid in air_pids

    title_lower = (title or "").lower()
    return any(
        keyword in title_lower
        for keyword in [
            "ccs air",
            "opc air",
            "air system",
            "automatic inspection report",
            "welcome to automatic inspection report",
            "검사기록조회",
            "검사이력",
        ]
    )


def is_startup_notification_popup(win, title: str = "") -> bool:
    """
    AIR 시작 직후 닫아도 되는 알림/공지성 팝업만 True로 판단합니다.
    사용자별로 뜨는 환경설정/업데이트/인증류 팝업은 건드리지 않고 로그인 흐름으로 넘깁니다.
    """
    notification_keywords = [
        "알림",
        "공지",
        "안내",
        "information",
        "notification",
        "notice",
    ]

    title_lower = (title or "").lower()
    if any(keyword in title_lower for keyword in notification_keywords):
        return True

    try:
        for ctrl in win.descendants():
            try:
                text = ctrl.window_text().strip()
                text_lower = text.lower()
                if text and any(keyword in text_lower for keyword in notification_keywords):
                    return True
            except Exception:
                pass
    except Exception:
        pass

    return False


def _startup_popup_button_texts(win) -> list[str]:
    texts = []
    try:
        for btn in win.descendants(control_type="Button"):
            try:
                text = btn.window_text().strip()
                if text:
                    texts.append(text)
            except Exception:
                pass
    except Exception:
        pass
    return texts


def _startup_popup_all_text(win, title: str = "") -> str:
    texts = [str(title or "")]
    try:
        for ctrl in win.descendants():
            try:
                text = ctrl.window_text().strip()
                if text:
                    texts.append(text)
            except Exception:
                pass
    except Exception:
        pass
    return "\n".join(texts).lower()


def _startup_has_continue_button(win) -> bool:
    return any(
        "계속" in text.lower() or "continue" in text.lower()
        for text in _startup_popup_button_texts(win)
    )


def _startup_has_login_input_fields(win) -> bool:
    """
    Welcome...Login 제목은 시작용 계속 팝업과 실제 로그인 창이 모두 사용할 수 있습니다.
    Edit 입력칸이 있어야 실제 로그인 창으로 봅니다.
    """
    try:
        return len(win.descendants(control_type="Edit")) >= 1
    except Exception:
        return False


def is_startup_actionable_popup(win, title: str = "") -> bool:
    """
    AIR 시작 시 로그인 전 막는 팝업인지 판단합니다.
    종료 전용 Alt+Q 로직과 분리되어 있으며, 시작 단계에서는 계속/확인/닫기만 처리합니다.
    """
    title_lower = (title or "").lower()
    startup_keywords = [
        "환경설정",
        "설정",
        "업데이트",
        "안내",
        "공지",
        "알림",
        "notice",
        "information",
        "notification",
        "update",
        "setup",
        "setting",
    ]
    if any(keyword in title_lower for keyword in startup_keywords):
        return True

    if is_startup_notification_popup(win, title):
        return True

    button_texts = _startup_popup_button_texts(win)
    action_keywords = ["계속", "continue", "확인", "ok", "닫기", "close", "예", "yes"]
    return any(
        any(keyword.lower() in text.lower() for keyword in action_keywords)
        for text in button_texts
    )


def looks_like_small_popup_window(win) -> bool:
    try:
        rect = win.rectangle()
        width = rect.width()
        height = rect.height()
        return 80 <= width <= 900 and 40 <= height <= 650
    except Exception:
        return False


def _popup_has_close_or_ok_button(win) -> bool:
    """
    종료 중 닫아도 되는 일반 팝업인지 판단할 때 쓰는 버튼 검사입니다.
    최종 종료창의 끝내기/종료/계속 버튼은 여기서 제외합니다.
    """
    captions = _get_button_caption_norms(win)
    allowed = ["닫기", "close", "확인", "ok", "예", "yes"]
    blocked = ["계속", "continue", "끝내기", "종료", "exit", "quit"]

    return any(
        any(key in caption for key in allowed)
        and not any(key in caption for key in blocked)
        for caption in captions
    )


def _popup_has_error_timeout_text(win, title: str = "") -> bool:
    all_text = _startup_popup_all_text(win, title) if win is not None else str(title or "").lower()
    keywords = [
        "crash at",
        "crash",
        "timeout",
        "time out",
        "timed out",
        "오류",
        "에러",
        "error",
        "exception",
        "예외",
        "failed",
        "failure",
        "실패",
        "경고",
        "warning",
    ]
    return any(keyword in all_text for keyword in keywords)


def _popup_has_crash_at_text(win, title: str = "") -> bool:
    all_text = _startup_popup_all_text(win, title) if win is not None else str(title or "").lower()
    title_lower = str(title or "").lower()
    return title_lower.startswith("crash at") or "crash at" in all_text


def _looks_like_air_shutdown_popup_to_close(win, title: str = "") -> bool:
    """
    AIR 종료 중 PID/제목 매칭이 애매하게 잡히는 작은 오류/timeout 팝업 보강.
    다른 프로그램 창을 건드리지 않도록 작은 창 + AIR/오류/알림 성격일 때만 True로 둡니다.
    """
    if win is None or not looks_like_small_popup_window(win):
        return False

    if _looks_like_final_air_exit_window(win):
        return False

    title_lower = (title or "").lower()
    all_text = _startup_popup_all_text(win, title)

    if _popup_has_error_timeout_text(win, title):
        return True

    shutdown_popup_keywords = [
        "알림",
        "공지",
        "안내",
        "환경설정",
        "설정",
        "notice",
        "information",
        "notification",
        "setting",
        "setup",
        "newair",
        "air",
        "automatic inspection report",
        "검사",
    ]

    if any(keyword in title_lower for keyword in shutdown_popup_keywords):
        return True

    return _popup_has_close_or_ok_button(win) and any(
        keyword in all_text for keyword in shutdown_popup_keywords
    )


def _click_window_titlebar_close(win, title: str = "") -> bool:
    """pywinauto rectangle()/close() 대기를 피하기 위해 hwnd의 GetWindowRect로 X를 클릭합니다."""
    try:
        hwnd = int(win.handle)
    except Exception:
        hwnd = 0

    try:
        if hwnd:
            rect = wintypes.RECT()
            if ctypes.windll.user32.GetWindowRect(wintypes.HWND(hwnd), ctypes.byref(rect)):
                x = int(rect.right) - 18
                y = int(rect.top) + 16
            else:
                return False
        else:
            rect = win.rectangle()
            x = int(rect.right) - 18
            y = int(rect.top) + 16

        if not is_inside_screen(x, y):
            return False
        real_mouse_click(x, y)
        time.sleep(0.12)
        log(f"[OK] 팝업 상단 X 클릭 처리(GetWindowRect): {title}")
        return True
    except Exception:
        return False


def _click_foreground_nonfinal_popup_x(timeout: float = 1.5, label: str = "팝업") -> bool:
    """
    종료 중 전면에 뜬 비최종 팝업을 제목/본문 탐지에 의존하지 않고 상단 X로 닫습니다.
    Crash at 오류창처럼 텍스트를 못 읽어도 창 자체가 전면에 있으면 닫을 수 있게 하기 위한 루틴입니다.
    """
    start = time.time()

    while time.time() - start < timeout:
        try:
            hwnd, title = get_foreground_window_info()
            title = (title or "").strip()
            if not hwnd or is_own_app_window(int(hwnd), title):
                time.sleep(0.05)
                continue

            if title and any(re.search(pattern, title) for pattern in AIR_TITLE_PATTERNS):
                time.sleep(0.05)
                continue

            win = _desktop_window_from_handle(int(hwnd))
            if win is not None:
                if _is_ignorable_air_internal_window(win, title):
                    time.sleep(0.05)
                    continue
                if _looks_like_final_air_exit_window(win):
                    log(f"[AIR-CLOSE] 최종 종료 팝업 감지 - X 닫기 보류: title='{title}'")
                    return False

            rect = wintypes.RECT()
            if not ctypes.windll.user32.GetWindowRect(wintypes.HWND(int(hwnd)), ctypes.byref(rect)):
                time.sleep(0.05)
                continue

            width = int(rect.right - rect.left)
            height = int(rect.bottom - rect.top)
            if not (80 <= width <= 1100 and 40 <= height <= 750):
                time.sleep(0.05)
                continue

            x = int(rect.right) - 18
            y = int(rect.top) + 16
            if not is_inside_screen(x, y):
                return False

            real_mouse_click(x, y)
            time.sleep(0.18)
            log(f"[AIR-CLOSE] {label} 전면 팝업 X 닫기(GetWindowRect): title='{title}'")
            return True
        except Exception as e:
            log(f"[WARN] {label} 전면 팝업 X 닫기 실패: {e}")
            break

    return False


def _press_enter_on_foreground_after_air_close(delay: float = 0.15, timeout: float = 1.2) -> bool:
    """
    AIR 메인 X 클릭 직후에는 알림창이 전면에 뜨는 것이 정해진 흐름입니다.
    창 탐색/필터가 흔들려 Enter를 놓치지 않도록 짧게 기다린 뒤 전면창에 바로 Enter를 보냅니다.
    """
    time.sleep(delay)
    start = time.time()

    while time.time() - start < timeout:
        try:
            hwnd, title = get_foreground_window_info()
            title = (title or "").strip()
            if not hwnd or is_own_app_window(int(hwnd), title):
                time.sleep(0.05)
                continue

            if title and any(re.search(pattern, title) for pattern in AIR_TITLE_PATTERNS):
                time.sleep(0.05)
                continue

            send_keys("{ENTER}")
            time.sleep(0.15)
            log(f"[AIR-CLOSE] X 클릭 후 전면 알림 Enter 입력: title='{title}'")
            return True
        except Exception as e:
            log(f"[WARN] X 클릭 후 전면 알림 Enter 실패: {e}")
            break

    return False


def _window_handle_exists(hwnd: int) -> bool:
    try:
        user32 = ctypes.windll.user32
        return bool(user32.IsWindow(hwnd)) and bool(user32.IsWindowVisible(hwnd))
    except Exception:
        return True


def _desktop_window_from_handle(hwnd: int):
    for backend in ("uia", "win32"):
        try:
            for win in Desktop(backend=backend).windows():
                try:
                    if int(win.handle) == int(hwnd):
                        return win
                except Exception:
                    pass
        except Exception:
            pass
    return None


def _classify_startup_popup(win, title: str = "", hwnd: int | None = None, air_pids: set | None = None) -> str | None:
    title_lower = (title or "").lower()
    all_text = _startup_popup_all_text(win, title) if win is not None else title_lower

    jit_keywords = [
        "jit",
        "just-in-time",
        "just in time",
        "objectdisposedexception",
        "frmscreensaver",
        "이 대화 상자 대신",
        "예외 텍스트",
        "jit(Just-in-time)",
    ]
    if any(keyword.lower() in all_text for keyword in jit_keywords):
        return "jit_continue"

    pid_is_air = False
    if hwnd is not None and _HAS_WIN32PROCESS:
        try:
            pid = _get_window_pid(int(hwnd))
            pid_is_air = pid is not None and (
                (air_pids and pid in air_pids) or (LAST_AIR_PIDS and pid in LAST_AIR_PIDS)
            )
        except Exception:
            pid_is_air = False

    if any(keyword in title_lower for keyword in ["newair", "microsoft .net framework"]) and (
        pid_is_air or win is None or _startup_has_continue_button(win) or "continue" in all_text or "계속" in all_text
    ):
        return "jit_continue"

    if win is not None and _startup_has_login_input_fields(win):
        return None

    # Welcome to Automatic Inspection Report - Login은 정상 시작/로그인 화면입니다.
    # 팝업으로 닫지 않고, 로그인 입력창 탐색/로그인 시도를 먼저 진행합니다.
    if "welcome to automatic inspection report" in title_lower:
        return None

    setting_keywords = ["환경설정", "설정", "setting", "setup"]
    if any(keyword in title_lower for keyword in setting_keywords):
        return "setting"

    try:
        if win is not None and is_startup_notification_popup(win, title):
            return "notice"
    except Exception:
        pass

    notice_keywords = ["알림", "공지", "안내", "notice", "information", "notification"]
    if any(keyword in title_lower for keyword in notice_keywords):
        return "notice"

    if win is not None and _popup_has_error_timeout_text(win, title):
        return "generic"

    if win is not None and is_startup_actionable_popup(win, title):
        return "generic"

    return None


def _close_startup_popup_by_type(win, hwnd: int, title: str, popup_type: str) -> bool:
    """
    시작 팝업 종류별 처리:
    - 알림: Enter 우선, 실패 시 X
    - 환경설정: X만
    - JIT/예외 계속 대화상자: X 금지, Alt+C 우선, 실패 시 계속 버튼
    """
    try:
        if win is not None:
            bring_window_to_front(win)
            try:
                win.set_focus()
            except Exception:
                pass
        else:
            ctypes.windll.user32.SetForegroundWindow(int(hwnd))
    except Exception:
        pass

    if popup_type == "jit_continue":
        try:
            send_keys("%c")
            time.sleep(0.15)
            log(f"[OK] 시작 계속 대화상자 Alt+C 처리({popup_type}): {title}")
            if not _window_handle_exists(hwnd):
                return True
        except Exception as e:
            log(f"[WARN] 시작 계속 대화상자 Alt+C 실패({popup_type}): {title} / {e}")

        if win is not None:
            clicked = _click_air_popup_button(win, target_texts=["계속", "continue"], blocked_texts=["끝내기", "종료", "exit", "quit"])
            if clicked:
                log(f"[OK] 시작 계속 대화상자 버튼 처리({popup_type}): '{clicked}' / {title}")
                time.sleep(0.12)
                return True
        return False

    if popup_type == "notice":
        try:
            if win is not None:
                try:
                    bring_window_to_front(win)
                    win.set_focus()
                except Exception:
                    pass
            send_keys("{ENTER}")
            time.sleep(0.12)
            log(f"[OK] 알림 팝업 Enter 처리: {title}")
            if not _window_handle_exists(hwnd):
                return True
        except Exception as e:
            log(f"[WARN] 알림 팝업 Enter 실패: {title} / {e}")

        if win is not None:
            clicked = _close_air_nonfinal_popup_by_button(win, title)
            if clicked:
                log(f"[OK] 알림 팝업 버튼 fallback 처리: '{clicked}' / {title}")
                return True

    if popup_type == "setting":
        if win is not None and _click_window_titlebar_close(win, title):
            if not _window_handle_exists(hwnd):
                return True

        if win is not None:
            clicked = _close_air_nonfinal_popup_by_button(win, title)
            if clicked:
                log(f"[OK] 환경설정 팝업 버튼 fallback 처리: '{clicked}' / {title}")
                return True

        try:
            send_keys("%{F4}")
            time.sleep(0.12)
            log(f"[OK] 환경설정 팝업 Alt+F4 fallback 처리: {title}")
            return True
        except Exception as e:
            log(f"[WARN] 환경설정 팝업 Alt+F4 실패: {title} / {e}")

    if popup_type in ("notice", "setting"):
        try:
            if win is not None:
                win.close()
            else:
                send_keys("%{F4}")
            time.sleep(0.12)
            log(f"[OK] {'환경설정' if popup_type == 'setting' else '알림'} 팝업 X 닫기 처리: {title}")
            return True
        except Exception as e:
            log(f"[WARN] 시작 팝업 X 닫기 실패: {title} / {e}")
            return False

    if popup_type == "generic" and win is not None:
        clicked = _click_air_popup_button(
            win,
            target_texts=["계속", "continue", "확인", "ok", "닫기", "close"],
            blocked_texts=["끝내기", "종료", "exit", "quit"],
        )
        if clicked:
            log(f"[OK] 시작 버튼형 팝업 처리: '{clicked}' / {title}")
            time.sleep(0.12)
            return True

    return False


def handle_foreground_air_startup_popup(air_pids: set | None = None) -> bool:
    """
    AIR 시작 직후 전면에 떠 있는 알림/환경설정/Welcome 팝업을 먼저 Enter 처리합니다.
    전체 창 탐색보다 빠르고, 버튼 텍스트를 못 읽는 팝업에도 동작합니다.
    """
    try:
        hwnd, title = get_foreground_window_info()
    except Exception:
        return False

    title = (title or "").strip()
    if not hwnd or is_own_app_window(int(hwnd), title):
        return False

    if "welcome to automatic inspection report" in title.lower():
        return False

    win = _desktop_window_from_handle(int(hwnd))
    active_air_pids = set(air_pids or LAST_AIR_PIDS)
    popup_type = _classify_startup_popup(win, title, hwnd=int(hwnd), air_pids=active_air_pids)

    pid_is_air = False
    if active_air_pids and _HAS_WIN32PROCESS:
        pid = _get_window_pid(int(hwnd))
        pid_is_air = pid is not None and pid in active_air_pids

    if popup_type is None and pid_is_air and win is not None and looks_like_small_popup_window(win):
        popup_type = "generic"

    if popup_type is None:
        return False

    log(f"[INFO] 전면 시작 팝업 처리 대상({popup_type}): title='{title}'")
    return _close_startup_popup_by_type(win, int(hwnd), title, popup_type)


def wait_for_jit_continue_dialog_after_settings(timeout: float = 5.0) -> bool:
    """
    환경설정 팝업을 닫은 직후 뜨는 .NET/JIT 계속 대화상자를 기다렸다가 Alt+C로 넘깁니다.
    """
    start = time.time()
    while time.time() - start < timeout:
        try:
            hwnd, title = get_foreground_window_info()
            if hwnd and not is_own_app_window(int(hwnd), title):
                win = _desktop_window_from_handle(int(hwnd))
                popup_type = _classify_startup_popup(win, title, hwnd=int(hwnd), air_pids=LAST_AIR_PIDS)
                if popup_type == "jit_continue":
                    log(f"[INFO] 환경설정 후 JIT 계속 대화상자 감지: {title}")
                    return _close_startup_popup_by_type(win, int(hwnd), title, popup_type)
        except Exception:
            pass

        try:
            for backend in ("uia", "win32"):
                for win in Desktop(backend=backend).windows():
                    try:
                        title = win.window_text().strip()
                        if not title:
                            continue
                        hwnd = int(win.handle)
                        if is_own_app_window(hwnd, title):
                            continue
                        popup_type = _classify_startup_popup(win, title, hwnd=hwnd, air_pids=LAST_AIR_PIDS)
                        if popup_type == "jit_continue":
                            log(f"[INFO] 환경설정 후 JIT 계속 대화상자 감지: {title}")
                            return _close_startup_popup_by_type(win, hwnd, title, popup_type)
                    except Exception:
                        pass
        except Exception:
            pass

        time.sleep(0.2)

    log("[INFO] 환경설정 후 JIT 계속 대화상자 감지 없음")
    return False


def press_windows_key():
    """Open the Windows Start menu/search without relying on a shortcut path."""
    user32 = ctypes.windll.user32
    VK_LWIN = 0x5B
    KEYEVENTF_KEYUP = 0x0002

    user32.keybd_event(VK_LWIN, 0, 0, 0)
    time.sleep(0.05)
    user32.keybd_event(VK_LWIN, 0, KEYEVENTF_KEYUP, 0)


def launch_air_from_windows_search(search_text: str = AIR_START_SEARCH_TEXT):
    """
    Launch AIR through Windows Start search.
    This avoids user-specific .appref-ms/.lnk shortcut paths.
    """
    before_pids = _snapshot_pids()

    log(f"[INFO] Windows 검색으로 AIR 실행 시도: {search_text}")
    press_windows_key()
    time.sleep(0.8)
    send_keys("^a")
    send_keys(search_text, with_spaces=True)
    time.sleep(1.5)
    send_keys("{ENTER}")
    log("[OK] Windows 검색 실행 요청 완료")

    return before_pids


def wait_for_air_startup_signal(timeout: float = 12.0) -> bool:
    """
    AIR 실행 요청 후 로그인/팝업/메인 창이 실제로 뜰 때까지만 기다립니다.
    Welcome...Login은 정상 시작/로그인 화면 후보로 두고, JIT/예외 계속 대화상자만 Alt+C 처리합니다.
    고정 sleep보다 빠르고, AIR가 늦게 뜨는 경우에도 시작 팝업 감시를 놓치지 않습니다.
    """
    start = time.time()
    while time.time() - start < timeout:
        try:
            for w in Desktop(backend="uia").windows():
                try:
                    title = w.window_text().strip()
                    if not title:
                        continue
                    hwnd = int(w.handle)
                    if is_own_app_window(hwnd, title):
                        continue

                    title_lower = title.lower()
                    if "welcome to automatic inspection report" in title_lower:
                        log(f"[OK] AIR 시작/로그인 화면 감지: {title}")
                        return True

                    popup_type = _classify_startup_popup(w, title, hwnd=hwnd, air_pids=LAST_AIR_PIDS)
                    if popup_type == "jit_continue":
                        if _startup_has_login_input_fields(w):
                            log(f"[OK] AIR 실제 로그인 입력창 감지: {title}")
                            return True

                        log(f"[INFO] 시작 계속 대화상자 감지 - Alt+C 처리({popup_type}): {title}")
                        _close_startup_popup_by_type(w, hwnd, title, popup_type)
                        time.sleep(0.2)
                        continue

                    if is_air_owned_window(hwnd, title) or (
                        looks_like_small_popup_window(w) and is_startup_actionable_popup(w, title)
                    ):
                        log(f"[OK] AIR 시작 신호 감지: {title}")
                        return True
                except Exception:
                    pass
        except Exception:
            pass
        time.sleep(0.2)

    log("[WARN] AIR 시작 창을 제한 시간 내 감지하지 못했습니다. 팝업 감시는 계속 진행합니다.")
    return False


def launch_air_if_needed():

    # AIR가 이미 실행 중이면 재실행하지 않습니다.
    try:
        if find_air_window(timeout=2):
            log("[INFO] AIR가 이미 실행 중입니다. 실행 생략.")
            return
    except Exception:
        pass

    before_pids = launch_air_from_windows_search()
    _collect_air_pids(before_pids)
    wait_for_air_startup_signal(timeout=12)


def dismiss_air_startup_popups(watch_seconds: int = 20, air_pids: set | None = None):
    """
    AIR 시작 직후 나타나는 팝업창들을 자동으로 닫습니다.

    air_pids가 제공되면 해당 PID 소속 창만 처리합니다(안전 모드).
    제공되지 않으면 창 크기 기반 폴백 모드로 동작합니다.
    """
    use_pid = bool(air_pids) and _HAS_WIN32PROCESS
    log(f"[INFO] AIR 팝업 닫기 시작 (모드: {'PID 필터' if use_pid else '크기 필터'}, 최대 {watch_seconds}초)")

    skip_keywords = ["로그인", "login", "logon", "sign in",
                     "ccs air", "opc air", "검사기록조회", "air system"]
    close_btn_keywords = ["계속", "continue", "닫기", "close", "확인", "ok", "예", "yes"]

    start = time.time()
    total_closed = 0
    size_mode_dismissed: set = set()
    login_first_seen: float | None = None   # 로그인 창 첫 감지 시각

    while time.time() - start < watch_seconds:
        if handle_foreground_air_startup_popup(air_pids=air_pids):
            total_closed += 1
            time.sleep(0.1)
            continue

        try:
            desktop_wins = Desktop(backend="uia").windows()
        except Exception:
            time.sleep(0.1)
            continue

        found_login_or_main = False
        closed_this_round = 0

        for w in desktop_wins:
            try:
                title = w.window_text().strip()
                if not title:
                    continue

                try:
                    hwnd = int(w.handle)
                except Exception:
                    hwnd = id(w)

                if is_own_app_window(hwnd, title):
                    continue

                title_lower = title.lower()
                popup_type = _classify_startup_popup(w, title, hwnd=hwnd, air_pids=air_pids or LAST_AIR_PIDS)
                actionable_popup = popup_type is not None

                # ── PID 필터 모드: AIR 프로세스 소속 창만 통과 ──────────
                if use_pid:
                    pid = _get_window_pid(hwnd)
                    # AIR 시작 알림/환경설정 팝업은 런처/별도 PID로 뜨는 경우가 있어
                    # PID가 달라도 작은 시작 팝업이면 처리 대상으로 허용합니다.
                    if (pid is None or pid not in air_pids) and not actionable_popup:
                        continue
                else:
                    if not is_air_owned_window(hwnd, title) and not actionable_popup:
                        continue

                    # ── 크기 필터 모드 (폴백) ───────────────────────────
                    if hwnd in size_mode_dismissed:
                        continue
                    try:
                        r = w.rectangle()
                        w_w = r.right - r.left
                        w_h = r.bottom - r.top
                        if w_w < 80 or w_h < 40:
                            continue
                        if w_w > 1400 or w_h > 900:
                            continue
                    except Exception:
                        continue

                # ── 로그인/메인 창 판별: 큰 메인/로그인 창은 닫지 않음 ──
                # AIR 제목을 가진 작은 팝업은 메인창으로 오인하지 않고 아래에서 버튼 처리합니다.
                is_main_login = (
                    any(kw in title_lower for kw in skip_keywords)
                    or any(re.search(p, title) for p in AIR_TITLE_PATTERNS)
                ) and not actionable_popup
                if is_main_login:
                    log(f"[INFO] 로그인/메인 창으로 판단해 시작 팝업 처리 제외: {title}")
                    found_login_or_main = True
                    continue  # break 대신 continue → 이후 팝업도 계속 처리

                if not actionable_popup:
                    continue

                log(f"[INFO] 시작 팝업 처리 대상({popup_type}): {title}")
                closed = _close_startup_popup_by_type(w, hwnd, title, popup_type)
                if closed:
                    total_closed += 1
                    closed_this_round += 1
                else:
                    log(f"[INFO] 시작 팝업 처리 실패({popup_type}): {title}")

                if not use_pid:
                    size_mode_dismissed.add(hwnd)

            except Exception:
                pass

        # ── 종료 조건 ────────────────────────────────────────────────────
        if found_login_or_main:
            if login_first_seen is None:
                login_first_seen = time.time()
                log("[INFO] 로그인/AIR 메인 창 감지 → 추가 팝업 짧게 감시")
            elif time.time() - login_first_seen >= 0.6 and closed_this_round == 0:
                # 로그인 창이 안정적으로 보이고 이번 라운드에 닫은 팝업이 없으면 종료
                log("[INFO] 로그인 창 안정화 확인 → 팝업 감시 종료")
                break
        else:
            login_first_seen = None  # 로그인 창이 사라지면 타이머 리셋

        time.sleep(0.15)

    log(f"[INFO] AIR 팝업 자동 닫기 완료 (총 {total_closed}개 처리)")
    return total_closed > 0


def find_air_login_window(timeout=40):
    """AIR 로그인 창을 탐지합니다. 로그인 다이얼로그 또는 메인 창의 로그인 폼을 찾습니다."""
    log("[INFO] AIR 로그인 창 탐색 중...")
    start = time.time()
    login_keywords = ["로그인", "login", "logon", "sign in", "air system", "ccs air", "opc air"]

    while time.time() - start < timeout:
        for w in Desktop(backend="uia").windows():
            try:
                title = w.window_text().strip()
                if not title:
                    continue
                title_lower = title.lower()
                title_matched = any(kw in title_lower for kw in login_keywords) or any(
                    re.search(pattern, title) for pattern in AIR_TITLE_PATTERNS
                )
                if not title_matched:
                    continue

                edits = w.descendants(control_type="Edit")
                buttons = w.descendants(control_type="Button")
                if len(edits) >= 1 and len(buttons) >= 1:
                    log(f"[OK] AIR 로그인 창 발견: {title}")
                    return w
            except Exception:
                pass
        time.sleep(0.2)

    return None


def air_auto_login(user_id: str, password: str, login_timeout: int = 60):
    """AIR 로그인 창을 찾아 자격증명을 입력하고 로그인합니다."""
    login_win = None
    start = time.time()

    while time.time() - start < login_timeout:
        login_win = find_air_login_window(timeout=2)
        if login_win is not None:
            break

        log("[INFO] 로그인 입력창 미감지 → 시작 팝업 처리 후 재시도")
        handled = handle_foreground_air_startup_popup(air_pids=LAST_AIR_PIDS)
        if not handled:
            handled = dismiss_air_startup_popups(watch_seconds=2, air_pids=LAST_AIR_PIDS)

        if not handled:
            time.sleep(0.3)

    if login_win is None:
        # 이미 로그인된 AIR 메인 창이 있는지 확인
        try:
            find_air_window(timeout=5)
            log("[OK] AIR가 이미 로그인된 상태입니다.")
            return
        except Exception:
            pass
        raise RuntimeError("AIR 로그인 창을 찾지 못했습니다. AIR가 정상적으로 실행 중인지 확인하세요.")

    log(f"[OK] AIR 로그인 창 발견: {login_win.window_text()}")
    bring_window_to_front(login_win)
    time.sleep(0.1)

    # ── 로그인 창 내부 알림/패널 닫기 ──────────────────────────────────
    # 'Welcome to AIR - Login' 창 안에 닫기 패널이 내장돼 있을 수 있습니다.
    # cmdExit / cmdLogin 이외의 '닫기' 버튼이 있으면 먼저 클릭합니다.
    INNER_CLOSE_TEXTS = ["계속", "continue", "확인", "ok", "닫기", "close"]
    try:
        for btn in login_win.descendants(control_type="Button"):
            try:
                txt  = btn.window_text().strip()
                try:
                    aid = btn.automation_id()
                except Exception:
                    aid = ""
                if aid in ("cmdExit", "cmdLogin"):
                    continue
                if any(kw.lower() in txt.lower() for kw in INNER_CLOSE_TEXTS):
                    btn.click_input()
                    time.sleep(0.1)
                    log(f"[OK] 로그인 창 내부 패널 닫기: '{txt}'")
                    break
            except Exception:
                pass
    except Exception:
        pass

    # ── 임베드된 자식 창(Window:1) 및 2차 패널 제거 ─────────────────────
    # 환경설정 닫기 후 추가 팝업이 로그인 창 내부에 나타날 수 있습니다.
    # 자식 Window 컨트롤을 짧게 탐색해 닫습니다.
    saw_setting_popup = False
    _t = time.time()
    while time.time() - _t < 3:
        found_child = False
        try:
            for child_win in login_win.descendants(control_type="Window"):
                try:
                    child_title = child_win.window_text().strip()
                    child_btns = child_win.descendants(control_type="Button")
                    if child_btns:
                        # 닫기 키워드 버튼 우선, 없으면 첫 번째 버튼
                        clicked = False
                        for cb in child_btns:
                            try:
                                ct = cb.window_text().strip()
                                if any(kw.lower() in ct.lower() for kw in INNER_CLOSE_TEXTS + ["확인", "ok", "계속"]):
                                    cb.click_input()
                                    time.sleep(0.1)
                                    log(f"[OK] 자식 창 닫기('{ct}'): '{child_title}'")
                                    if "환경설정" in child_title or "setting" in child_title.lower():
                                        saw_setting_popup = True
                                    clicked = True
                                    found_child = True
                                    break
                            except Exception:
                                pass
                        if not clicked:
                            child_btns[0].click_input()
                            time.sleep(0.1)
                            log(f"[OK] 자식 창 첫버튼 클릭: '{child_title}'")
                            if "환경설정" in child_title or "setting" in child_title.lower():
                                saw_setting_popup = True
                            found_child = True
                    else:
                        # 버튼 없는 자식 창 → Enter
                        try:
                            child_win.set_focus()
                            send_keys("{ENTER}")
                            time.sleep(0.1)
                            log(f"[OK] 자식 창 Enter: '{child_title}'")
                            if "환경설정" in child_title or "setting" in child_title.lower():
                                saw_setting_popup = True
                            found_child = True
                        except Exception:
                            pass
                except Exception:
                    pass
        except Exception:
            pass

        if not found_child:
            break
        time.sleep(0.1)

    log("[INFO] 로그인 창 내부 팝업 처리 완료")
    if saw_setting_popup:
        wait_for_jit_continue_dialog_after_settings(timeout=5.0)
    time.sleep(0.1)

    edits = login_win.descendants(control_type="Edit")
    if not edits:
        raise RuntimeError("AIR 로그인 창에서 입력 필드를 찾지 못했습니다.")

    # AIR exposes login fields in reverse order: password first, user id second.
    user_id_edit = edits[1] if len(edits) >= 2 else edits[0]
    password_edit = edits[0] if len(edits) >= 2 else None

    try:
        user_id_edit.click_input()
        time.sleep(0.2)
        user_id_edit.set_text(user_id)
        log("[OK] 아이디 입력 완료")
    except Exception as e:
        log(f"[WARN] 아이디 직접 입력 실패, 키보드 방식 시도: {e}")
        try:
            user_id_edit.click_input()
            send_keys("^a")
            send_keys(user_id, with_spaces=True)
        except Exception:
            pass

    # First Edit = password.
    if password_edit is not None:
        try:
            password_edit.click_input()
            time.sleep(0.2)
            password_edit.set_text(password)
            log("[OK] 비밀번호 입력 완료")
        except Exception as e:
            log(f"[WARN] 비밀번호 직접 입력 실패, 키보드 방식 시도: {e}")
            try:
                password_edit.click_input()
                send_keys("^a")
                send_keys(password, with_spaces=True)
            except Exception:
                pass
    else:
        # Tab으로 다음 필드로 이동
        try:
            send_keys("{TAB}")
            time.sleep(0.2)
            send_keys(password, with_spaces=True)
            log("[OK] 비밀번호 입력 완료 (Tab)")
        except Exception as e:
            log(f"[WARN] 비밀번호 입력 실패: {e}")

    time.sleep(0.08)

    # 로그인 버튼 찾기: auto_id='cmdLogin' 우선, 없으면 텍스트 키워드
    buttons = login_win.descendants(control_type="Button")
    login_btn = None
    try:
        for btn in buttons:
            try:
                if btn.automation_id() == "cmdLogin":
                    login_btn = btn
                    break
            except Exception:
                pass
    except Exception:
        pass
    if login_btn is None:
        for btn in buttons:
            try:
                txt = btn.window_text().strip().lower()
                if any(kw in txt for kw in ["로그인", "login", "확인", "ok", "sign"]):
                    login_btn = btn
                    break
            except Exception:
                pass
    if login_btn is None and buttons:
        login_btn = buttons[0]

    if login_btn:
        try:
            login_btn.click_input()
            log("[OK] 로그인 버튼 클릭 완료")
        except Exception as e:
            send_keys("{ENTER}")
            log(f"[OK] 로그인 Enter 키 입력 (버튼 클릭 실패: {e})")
    else:
        send_keys("{ENTER}")
        log("[OK] 로그인 Enter 키 입력")

    # 메인 창 로드 대기
    log("[INFO] AIR 메인 화면 로드 대기 중...")
    try:
        find_air_window(timeout=30)
        log("[OK] AIR 로그인 및 메인 화면 진입 완료")
    except Exception:
        log("[WARN] AIR 메인 화면 확인 실패 - 로그인이 완료됐을 수 있습니다.")


def find_air_window(timeout=30):
    start = time.time()

    while time.time() - start < timeout:
        for w in Desktop(backend="uia").windows():
            try:
                title = w.window_text()

                if not title.strip():
                    continue

                for pattern in AIR_TITLE_PATTERNS:
                    if re.search(pattern, title):
                        log(f"[OK] AIR 창 발견: {title}")
                        return w

            except Exception:
                pass

        time.sleep(0.2)

    raise RuntimeError("AIR 창을 찾지 못했습니다. AIR가 실행 중인지 확인하세요.")


def bring_window_to_front(win):
    try:
        win.restore()
    except Exception:
        pass

    try:
        win.set_focus()
    except Exception:
        pass

    try:
        handle = int(win.handle)
        ctypes.windll.user32.ShowWindow(handle, 9)
        ctypes.windll.user32.SetForegroundWindow(handle)
    except Exception:
        pass

    time.sleep(0.05)


def refresh_air_window():
    cached = CONTROL_CACHE.get("air_window")
    if cached is not None:
        try:
            title = cached.window_text()
            if any(re.search(pattern, title) for pattern in AIR_TITLE_PATTERNS):
                bring_window_to_front(cached)
                return cached
        except Exception:
            CONTROL_CACHE.pop("air_window", None)

    win = find_air_window(timeout=15)
    bring_window_to_front(win)
    CONTROL_CACHE["air_window"] = win
    return win


def _normalize_button_caption(text: str) -> str:
    return re.sub(r"\s+", "", str(text or "").strip()).lower()


def _get_window_text_by_hwnd(hwnd: int) -> str:
    try:
        length = ctypes.windll.user32.GetWindowTextLengthW(wintypes.HWND(int(hwnd)))
        buff = ctypes.create_unicode_buffer(length + 1)
        ctypes.windll.user32.GetWindowTextW(wintypes.HWND(int(hwnd)), buff, length + 1)
        return buff.value.strip()
    except Exception:
        return ""


def _get_class_name_by_hwnd(hwnd: int) -> str:
    try:
        buff = ctypes.create_unicode_buffer(256)
        ctypes.windll.user32.GetClassNameW(wintypes.HWND(int(hwnd)), buff, 256)
        return buff.value.strip()
    except Exception:
        return ""


def _click_native_child_button_by_caption(parent_hwnd: int, target_texts, blocked_texts=None) -> str:
    target_norms = [_normalize_button_caption(text) for text in target_texts]
    blocked_norms = [_normalize_button_caption(text) for text in (blocked_texts or [])]
    matches = []

    try:
        enum_proc_type = ctypes.WINFUNCTYPE(ctypes.c_bool, wintypes.HWND, wintypes.LPARAM)

        def enum_proc(child_hwnd, lparam):
            try:
                child = int(child_hwnd)
                caption = _get_window_text_by_hwnd(child)
                caption_norm = _normalize_button_caption(caption)
                if not caption_norm:
                    return True

                class_name = _get_class_name_by_hwnd(child).lower()
                if "button" not in class_name:
                    return True

                if any(blocked and blocked in caption_norm for blocked in blocked_norms):
                    return True

                if caption_norm in target_norms or any(target in caption_norm for target in target_norms):
                    rect = wintypes.RECT()
                    if ctypes.windll.user32.GetWindowRect(wintypes.HWND(child), ctypes.byref(rect)):
                        matches.append((int(rect.left), int(rect.top), int(rect.right), int(rect.bottom), caption))
            except Exception:
                pass
            return True

        ctypes.windll.user32.EnumChildWindows(
            wintypes.HWND(int(parent_hwnd)),
            enum_proc_type(enum_proc),
            0,
        )
    except Exception:
        return ""

    if not matches:
        return ""

    left, top, right, bottom, caption = matches[0]
    x = int((left + right) / 2)
    y = int((top + bottom) / 2)
    if not is_inside_screen(x, y):
        return ""

    real_mouse_click(x, y)
    time.sleep(0.12)
    return caption or "native button"


def _native_child_button_captions(parent_hwnd: int) -> list[str]:
    captions = []

    try:
        enum_proc_type = ctypes.WINFUNCTYPE(ctypes.c_bool, wintypes.HWND, wintypes.LPARAM)

        def enum_proc(child_hwnd, lparam):
            try:
                child = int(child_hwnd)
                caption = _get_window_text_by_hwnd(child)
                if not caption:
                    return True

                class_name = _get_class_name_by_hwnd(child).lower()
                if "button" in class_name:
                    captions.append(caption)
            except Exception:
                pass
            return True

        ctypes.windll.user32.EnumChildWindows(
            wintypes.HWND(int(parent_hwnd)),
            enum_proc_type(enum_proc),
            0,
        )
    except Exception:
        pass

    return captions


def _native_window_has_exit_button(hwnd: int) -> bool:
    for caption in _native_child_button_captions(hwnd):
        caption_norm = _normalize_button_caption(caption)
        if any(keyword in caption_norm for keyword in ["끝내기", "종료", "exit", "quit"]):
            return True
    return False


def _is_button_like_control(ctrl) -> bool:
    try:
        control_type = str(ctrl.element_info.control_type or "").lower()
        if control_type == "button":
            return True
    except Exception:
        pass

    try:
        friendly = str(ctrl.friendly_class_name() or "").lower()
        if "button" in friendly:
            return True
    except Exception:
        pass

    try:
        class_name = str(ctrl.class_name() or "").lower()
        if "button" in class_name:
            return True
    except Exception:
        pass

    return False


def _popup_button_controls(win) -> list:
    buttons = []
    seen = set()

    for kwargs in ({"control_type": "Button"}, {}):
        try:
            controls = win.descendants(**kwargs)
        except Exception:
            controls = []

        for ctrl in controls:
            try:
                handle = int(getattr(ctrl, "handle", 0) or 0)
            except Exception:
                handle = 0
            key = handle or id(ctrl)
            if key in seen:
                continue
            if kwargs or _is_button_like_control(ctrl):
                buttons.append(ctrl)
                seen.add(key)

    return buttons


def _click_air_popup_button(win, target_texts, blocked_texts=None) -> str:
    target_norms = [_normalize_button_caption(text) for text in target_texts]
    blocked_norms = {_normalize_button_caption(text) for text in (blocked_texts or [])}

    try:
        native_clicked = _click_native_child_button_by_caption(int(win.handle), target_texts, blocked_texts)
        if native_clicked:
            return native_clicked
    except Exception:
        pass

    for btn in _popup_button_controls(win):
        try:
            caption = btn.window_text().strip()
            if not caption:
                try:
                    texts = [text for text in btn.texts() if str(text).strip()]
                    caption = texts[0].strip() if texts else ""
                except Exception:
                    caption = ""
            caption_norm = _normalize_button_caption(caption)
            if not caption_norm:
                continue
            if any(blocked and blocked in caption_norm for blocked in blocked_norms):
                continue

            if caption_norm in target_norms or any(target in caption_norm for target in target_norms):
                try:
                    btn.click_input()
                except Exception:
                    rect = btn.rectangle()
                    real_mouse_click(int((rect.left + rect.right) / 2), int((rect.top + rect.bottom) / 2))
                return caption
        except Exception:
            pass

    return ""


def _click_control_by_hwnd_rect(ctrl) -> bool:
    try:
        hwnd = int(getattr(ctrl, "handle", 0) or 0)
    except Exception:
        hwnd = 0

    try:
        if hwnd:
            rect = wintypes.RECT()
            if ctypes.windll.user32.GetWindowRect(wintypes.HWND(hwnd), ctypes.byref(rect)):
                x = int((rect.left + rect.right) / 2)
                y = int((rect.top + rect.bottom) / 2)
                if is_inside_screen(x, y):
                    real_mouse_click(x, y)
                    return True

        rect = ctrl.rectangle()
        x = int((rect.left + rect.right) / 2)
        y = int((rect.top + rect.bottom) / 2)
        if is_inside_screen(x, y):
            real_mouse_click(x, y)
            return True
    except Exception:
        pass

    return False


def _click_any_air_exit_popup_button_or_x(win, title: str = "") -> str:
    """
    AIR 종료 중 뜨는 팝업은 종류가 흔들리므로, 버튼 문구를 넓게 보고 클릭합니다.
    마지막 종료 팝업(끝내기/종료 버튼)은 이 함수에서 처리하지 않고 Alt+Q 전용 함수로 넘깁니다.
    버튼 클릭은 가능한 한 버튼 hwnd의 GetWindowRect 좌표로 처리하고, 실패하면 창 X를 클릭합니다.
    """
    if _looks_like_final_air_exit_window(win):
        return ""

    target_norms = [
        _normalize_button_caption(text)
        for text in [
            "닫기",
            "닫기(&c)",
            "close",
            "확인",
            "확인(&o)",
            "ok",
            "OK",
            "예",
            "yes",
        ]
    ]
    blocked_norms = [
        _normalize_button_caption(text)
        for text in ["취소", "cancel", "아니오", "no", "계속", "continue", "끝내기", "종료", "exit", "quit"]
    ]

    try:
        native_clicked = _click_native_child_button_by_caption(
            int(win.handle),
            [
                "닫기",
                "닫기(&c)",
                "close",
                "확인",
                "확인(&o)",
                "ok",
                "OK",
                "예",
                "yes",
            ],
            ["취소", "cancel", "아니오", "no", "계속", "continue", "끝내기", "종료", "exit", "quit"],
        )
        if native_clicked:
            return native_clicked
    except Exception:
        pass

    for btn in _popup_button_controls(win):
        try:
            caption = btn.window_text().strip()
            if not caption:
                try:
                    texts = [text for text in btn.texts() if str(text).strip()]
                    caption = texts[0].strip() if texts else ""
                except Exception:
                    caption = ""

            caption_norm = _normalize_button_caption(caption)
            if not caption_norm:
                continue
            if any(blocked and blocked in caption_norm for blocked in blocked_norms):
                continue
            if not (caption_norm in target_norms or any(target in caption_norm for target in target_norms)):
                continue

            try:
                btn.click_input()
            except Exception:
                if not _click_control_by_hwnd_rect(btn):
                    continue

            time.sleep(0.15)
            return caption or "button"
        except Exception:
            pass

    if _click_window_titlebar_close(win, title):
        return "X"

    return ""


def _get_button_caption_norms(win) -> list[str]:
    captions = []
    for btn in _popup_button_controls(win):
        try:
            caption = btn.window_text().strip()
            if not caption:
                try:
                    texts = [text for text in btn.texts() if str(text).strip()]
                    caption = texts[0].strip() if texts else ""
                except Exception:
                    caption = ""
            caption = _normalize_button_caption(caption)
            if caption:
                captions.append(caption)
        except Exception:
            pass
    return captions


def _looks_like_final_air_exit_window(win) -> bool:
    title_has_welcome = False

    try:
        title = win.window_text().strip().lower()
        if "welcome to automatic inspection report" in title or "대화상자" in title:
            title_has_welcome = True
    except Exception:
        pass

    try:
        if _startup_has_login_input_fields(win):
            return False
    except Exception:
        pass

    captions = _get_button_caption_norms(win)
    has_continue = any("계속" in caption or "continue" in caption for caption in captions)
    has_exit = any(
        "끝내기" in caption
        or "종료" in caption
        or "exit" in caption
        or "quit" in caption
        for caption in captions
    )

    if has_exit and looks_like_small_popup_window(win):
        return True

    if has_exit and (has_continue or title_has_welcome):
        return True

    return False


def _has_air_exit_button(win) -> bool:
    """
    종료/끝내기 버튼이 있는 AIR 팝업인지 확인합니다.
    이런 창은 상단 X로 닫지 않고 버튼 클릭으로만 처리합니다.
    """
    captions = _get_button_caption_norms(win)
    return any(
        "끝내기" in caption
        or "종료" in caption
        or "exit" in caption
        or "quit" in caption
        for caption in captions
    )


def _looks_like_air_exit_popup_window(win) -> bool:
    """
    AIR 메인창이 아니라 종료 확인용 팝업 크기인지 확인합니다.
    """
    try:
        rect = win.rectangle()
        width = rect.width()
        height = rect.height()
        return 120 <= width <= 900 and 80 <= height <= 650
    except Exception:
        return False


def _is_ignorable_air_internal_window(win, title: str = "") -> bool:
    """
    AIR 프로세스 소유이지만 사용자가 조작하는 팝업이 아닌 내부/보조 창은 제외합니다.
    """
    title_lower = (title or "").strip().lower()

    if not title_lower:
        return True

    internal_title_keywords = [
        "default ime",
        "gdi+ window",
        ".net-broadcasteventwindow",
        "broadcasteventwindow",
    ]
    if any(keyword in title_lower for keyword in internal_title_keywords):
        return True

    if title_lower in {"m"}:
        return True

    try:
        hwnd = int(win.handle)
        if not ctypes.windll.user32.IsWindowVisible(hwnd):
            return True
    except Exception:
        pass

    try:
        class_name = win.class_name().strip().lower()
        internal_class_keywords = ["ime", "broadcasteventwindow"]
        if any(keyword in class_name for keyword in internal_class_keywords):
            return True
    except Exception:
        pass

    return False


def _click_rightmost_air_popup_button(win, blocked_texts=None) -> str:
    blocked_norms = {_normalize_button_caption(text) for text in (blocked_texts or [])}
    candidates = []

    try:
        buttons = win.descendants(control_type="Button")
    except Exception:
        buttons = []

    for btn in buttons:
        try:
            caption = btn.window_text().strip()
            caption_norm = _normalize_button_caption(caption)
            if any(blocked and blocked in caption_norm for blocked in blocked_norms):
                continue

            rect = btn.rectangle()
            if rect.width() <= 0 or rect.height() <= 0:
                continue

            candidates.append((rect.right, rect.left, caption, btn))
        except Exception:
            pass

    if not candidates:
        return ""

    candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
    _, _, caption, btn = candidates[0]

    try:
        try:
            btn.click_input()
        except Exception:
            rect = btn.rectangle()
            real_mouse_click(int((rect.left + rect.right) / 2), int((rect.top + rect.bottom) / 2))
        return caption or "오른쪽 끝 버튼"
    except Exception:
        return ""


def _close_final_air_exit_window(win, title: str = "") -> bool:
    """
    최종 AIR 종료 대화상자는 상단 X/버튼 클릭을 쓰지 않고 Alt+Q만 사용합니다.
    """
    try:
        bring_window_to_front(win)
    except Exception:
        pass

    try:
        win.set_focus()
    except Exception:
        pass

    try:
        send_keys("%q")
        log(f"[AIR-CLOSE] 최종 종료 대화상자 Alt+Q 입력: title='{title}'")
        time.sleep(0.25)
        try:
            hwnd = int(win.handle)
            if not _window_handle_exists(hwnd):
                return True
        except Exception:
            return True
    except Exception as e:
        log(f"[WARN] 최종 종료 대화상자 Alt+Q 입력 실패: {e}")

    return False


def close_final_air_exit_window_if_present(timeout: float = 2.0) -> bool:
    """
    최종 Welcome 종료 대화상자를 PID 필터 없이 제목/버튼 기준으로 찾아 닫습니다.
    알림/환경설정 처리 직후 새로 뜨는 타이밍을 놓치지 않기 위한 보강 루틴입니다.
    """
    start = time.time()

    while time.time() - start < timeout:
        if _has_air_settings_popup(timeout=0.1):
            time.sleep(0.1)
            continue

        candidates = []

        for backend in ["uia", "win32"]:
            try:
                for win in Desktop(backend=backend).windows():
                    try:
                        title = win.window_text().strip()
                        if not title:
                            continue

                        hwnd = int(win.handle)
                        if is_own_app_window(hwnd, title):
                            continue

                        if _is_ignorable_air_internal_window(win, title):
                            continue

                        if _looks_like_final_air_exit_window(win):
                            candidates.append((title, win))
                    except Exception:
                        pass
            except Exception:
                pass

        if candidates:
            title, win = candidates[0]
            return _close_final_air_exit_window(win, title)

        time.sleep(0.1)

    return False


def _close_air_nonfinal_popup_by_button(win, title: str = "") -> str:
    """
    종료 과정에서 최종 Welcome 대화상자가 아닌 AIR 팝업을 닫습니다.
    닫기/확인 계열 버튼만 누르고, 계속/끝내기/종료 버튼은 건드리지 않습니다.
    """
    clicked = _click_air_popup_button(
        win,
        target_texts=[
            "닫기",
            "닫기(&c)",
            "close",
            "확인",
            "확인(&o)",
            "ok",
            "OK",
            "예",
            "yes",
        ],
        blocked_texts=[
            "계속",
            "continue",
            "끝내기",
            "종료",
            "exit",
            "quit",
        ],
    )
    if clicked:
        return clicked

    try:
        win.set_focus()
    except Exception:
        pass

    try:
        send_keys("{ENTER}")
        time.sleep(0.12)
        return "Enter"
    except Exception:
        pass

    try:
        send_keys("%{F4}")
        time.sleep(0.12)
        return "Alt+F4"
    except Exception:
        return ""


def handle_foreground_air_exit_popup(air_pids: set | None = None) -> bool:
    """
    종료 중 전면에 떠 있는 팝업을 시작 팝업 처리처럼 최우선으로 처리합니다.
    알림/환경설정/오류 팝업이 같은 Welcome 제목을 써도 최종 종료창으로 오판하지 않게 합니다.
    """
    try:
        hwnd, title = get_foreground_window_info()
    except Exception:
        return False

    title = (title or "").strip()
    if not hwnd or not title or is_own_app_window(int(hwnd), title):
        return False

    win = _desktop_window_from_handle(int(hwnd))
    if win is None or _is_ignorable_air_internal_window(win, title):
        return False

    if _looks_like_final_air_exit_window(win):
        return _close_final_air_exit_window(win, title)

    active_air_pids = set(air_pids or LAST_AIR_PIDS)
    is_air_window = is_air_owned_window(int(hwnd), title, air_pids=active_air_pids)
    is_shutdown_popup = _looks_like_air_shutdown_popup_to_close(win, title)

    if not is_air_window and not is_shutdown_popup:
        return False

    popup_type = _classify_startup_popup(win, title, hwnd=int(hwnd), air_pids=active_air_pids)

    if popup_type is None and "welcome to automatic inspection report" in title.lower():
        if is_startup_notification_popup(win, title):
            popup_type = "notice"
        elif _popup_has_close_or_ok_button(win):
            popup_type = "generic"

    if popup_type is None and is_shutdown_popup:
        popup_type = "generic"

    if popup_type in ("notice", "setting"):
        if _close_startup_popup_by_type(win, int(hwnd), title, popup_type):
            log(f"[AIR-CLOSE] 전면 {'알림' if popup_type == 'notice' else '환경설정'} 팝업 처리 완료: title='{title}'")
            return True

    if popup_type == "generic":
        clicked = _close_air_nonfinal_popup_by_button(win, title)
        if clicked:
            log(f"[AIR-CLOSE] 전면 기타 팝업 처리: '{clicked}' / title='{title}'")
            return True

    return False


def handle_air_exit_popups(timeout: float = 6.0, air_pids: set | None = None) -> bool:
    """
    AIR 종료 중 뜨는 알림/환경설정/최종 종료 확인창을 처리합니다.
    - 알림: 시작 화면과 동일하게 Enter 우선, 필요 시 X
    - 환경설정: 시작 화면과 동일하게 X
    - 최종 Welcome 대화상자: Alt+Q만 사용
    """
    start = time.time()
    handled_any = False

    while time.time() - start < timeout:
        handled_this_round = False
        windows = []
        final_exit_popups = []

        if handle_foreground_air_exit_popup(air_pids=air_pids):
            handled_any = True
            handled_this_round = True
            time.sleep(0.15)
            continue

        for backend in ["uia", "win32"]:
            try:
                windows.extend(Desktop(backend=backend).windows())
            except Exception:
                pass

        if not windows:
            time.sleep(0.1)
            continue

        for popup in windows:
            try:
                title = popup.window_text().strip()
                if not title:
                    continue

                hwnd = int(popup.handle)
                if is_own_app_window(hwnd, title):
                    continue

                if _is_ignorable_air_internal_window(popup, title):
                    continue

                is_final_exit_window = _looks_like_final_air_exit_window(popup)

                is_air_window = is_air_owned_window(hwnd, title, air_pids=air_pids)
                is_shutdown_popup = _looks_like_air_shutdown_popup_to_close(popup, title)

                if not is_final_exit_window and not is_air_window and not is_shutdown_popup:
                    continue

                popup_type = None if is_final_exit_window else _classify_startup_popup(
                    popup,
                    title,
                    hwnd=hwnd,
                    air_pids=air_pids,
                    )

                if is_final_exit_window:
                    final_exit_popups.append((popup, title))
                    continue

                if popup_type is None and "welcome to automatic inspection report" in title.lower():
                    if is_startup_notification_popup(popup, title):
                        popup_type = "notice"
                    elif _popup_has_close_or_ok_button(popup):
                        popup_type = "generic"

                if is_shutdown_popup and popup_type is None:
                    popup_type = "generic"

                if popup_type in ("notice", "setting"):
                    if _close_startup_popup_by_type(popup, hwnd, title, popup_type):
                        log(f"[AIR-CLOSE] 종료 중 {'알림' if popup_type == 'notice' else '환경설정'} 팝업 처리 완료: title='{title}'")
                        handled_any = True
                        handled_this_round = True
                        time.sleep(0.15)
                    else:
                        clicked = _close_air_nonfinal_popup_by_button(popup, title)
                        if clicked:
                            log(f"[AIR-CLOSE] 종료 중 {'알림' if popup_type == 'notice' else '환경설정'} 팝업 fallback 처리: '{clicked}' / title='{title}'")
                            handled_any = True
                            handled_this_round = True
                            time.sleep(0.15)
                    continue

                if (popup_type == "generic" or popup_type is None) and looks_like_small_popup_window(popup):
                    clicked = _close_air_nonfinal_popup_by_button(popup, title)
                    if clicked:
                        log(f"[AIR-CLOSE] 종료 중 기타 팝업 처리: '{clicked}' / title='{title}'")
                        handled_any = True
                        handled_this_round = True
                        time.sleep(0.12)
                    continue

                # 알림/환경설정/최종 종료 대화상자가 아닌 AIR 내부 창은 건드리지 않습니다.
                continue
            except Exception as e:
                log(f"[WARN] AIR 종료 팝업 처리 중 오류: {e}")

        if not handled_this_round and final_exit_popups:
            for popup, title in final_exit_popups:
                if _close_final_air_exit_window(popup, title):
                    handled_any = True
                    handled_this_round = True
                    break

        if not handled_this_round and close_final_air_exit_window_if_present(timeout=0.5):
            handled_any = True
            handled_this_round = True

        if not handled_this_round:
            time.sleep(0.1)

    return handled_any


def _iter_candidate_air_exit_windows(air_pids: set | None = None):
    seen = set()

    for backend in ("uia", "win32"):
        try:
            windows = Desktop(backend=backend).windows()
        except Exception:
            windows = []

        for win in windows:
            try:
                hwnd = int(win.handle)
                if hwnd in seen:
                    continue
                seen.add(hwnd)

                title = win.window_text().strip()
                if not title:
                    continue
                if is_own_app_window(hwnd, title):
                    continue
                if _is_ignorable_air_internal_window(win, title):
                    continue

                is_final = _looks_like_final_air_exit_window(win)
                is_air = is_air_owned_window(hwnd, title, air_pids=air_pids)
                is_shutdown_popup = _looks_like_air_shutdown_popup_to_close(win, title)

                if is_final or is_air or is_shutdown_popup:
                    yield win, hwnd, title
            except Exception:
                pass


def drain_any_air_exit_popups_by_button_or_x(timeout: float = 1.5, air_pids: set | None = None, max_count: int = 6) -> int:
    """
    종료 중 분류가 안 되는 팝업도 버튼 문구 또는 X로 처리합니다.
    메인 AIR 화면과 최종 종료 팝업은 제외합니다. 최종 종료 팝업은 항상 Alt+Q 전용입니다.
    """
    closed = 0
    start = time.time()

    while closed < max_count and time.time() - start < timeout:
        handled = False

        try:
            hwnd, title = get_foreground_window_info()
            title = (title or "").strip()
            if hwnd and title and not is_own_app_window(int(hwnd), title):
                win = _desktop_window_from_handle(int(hwnd))
                if win is not None and not _is_ignorable_air_internal_window(win, title):
                    if _looks_like_final_air_exit_window(win):
                        return closed

                    is_popup = looks_like_small_popup_window(win)
                    is_main_air = any(re.search(pattern, title) for pattern in AIR_TITLE_PATTERNS)
                    if is_popup and not is_main_air:
                        clicked = _click_any_air_exit_popup_button_or_x(win, title)
                        if clicked:
                            log(f"[AIR-CLOSE] 전면 종료 팝업 공용 처리: '{clicked}' / title='{title}'")
                            closed += 1
                            handled = True
                            time.sleep(0.15)
        except Exception:
            pass

        if handled:
            continue

        for win, hwnd, title in _iter_candidate_air_exit_windows(air_pids=air_pids):
            try:
                if _looks_like_final_air_exit_window(win):
                    continue

                is_popup = looks_like_small_popup_window(win)
                is_main_air = any(re.search(pattern, title) for pattern in AIR_TITLE_PATTERNS)
                if not is_popup or is_main_air:
                    continue

                clicked = _click_any_air_exit_popup_button_or_x(win, title)
                if clicked:
                    log(f"[AIR-CLOSE] 종료 팝업 공용 처리: '{clicked}' / title='{title}'")
                    closed += 1
                    handled = True
                    time.sleep(0.15)
                    break
            except Exception:
                pass

        if not handled:
            time.sleep(0.1)
            break

    return closed


def _wait_for_air_exit_popup(predicate, timeout: float, air_pids: set | None = None):
    start = time.time()

    while time.time() - start < timeout:
        try:
            hwnd, title = get_foreground_window_info()
            title = (title or "").strip()
            if hwnd and title and not is_own_app_window(int(hwnd), title):
                win = _desktop_window_from_handle(int(hwnd))
                if win is not None and not _is_ignorable_air_internal_window(win, title):
                    is_final = _looks_like_final_air_exit_window(win)
                    is_air = is_air_owned_window(int(hwnd), title, air_pids=air_pids)
                    is_shutdown_popup = _looks_like_air_shutdown_popup_to_close(win, title)
                    if (is_final or is_air or is_shutdown_popup) and predicate(win, int(hwnd), title):
                        return win, int(hwnd), title
        except Exception:
            pass

        for win, hwnd, title in _iter_candidate_air_exit_windows(air_pids=air_pids):
            try:
                if predicate(win, hwnd, title):
                    return win, hwnd, title
            except Exception:
                pass

        time.sleep(0.1)

    return None


def _close_pending_air_timeout_warning(timeout: float = 1.0, air_pids: set | None = None) -> bool:
    """
    종료 순서 중간에 끼는 timeout/오류/경고 팝업은 최종 종료창이 아니므로 닫기 버튼으로 제거합니다.
    """
    item = _wait_for_air_exit_popup(
        lambda win, hwnd, title: (
            not _looks_like_final_air_exit_window(win)
            and _popup_has_error_timeout_text(win, title)
            and _popup_has_close_or_ok_button(win)
        ),
        timeout=timeout,
        air_pids=air_pids,
    )
    if not item:
        return False

    win, hwnd, title = item
    clicked = _close_air_nonfinal_popup_by_button(win, title)
    if clicked:
        log(f"[AIR-CLOSE] timeout/경고 팝업 닫기 처리: '{clicked}' / title='{title}'")
        return True

    return False


def _is_blocking_nonfinal_air_exit_popup(win, title: str = "", hwnd: int | None = None, air_pids: set | None = None) -> bool:
    """
    종료 순서 중간에 끼어 진행을 막는 일반 팝업인지 판단합니다.
    오류 문구를 못 읽어도 닫기/확인 버튼이 있는 작은 팝업이면 막힘 팝업으로 봅니다.
    """
    if win is None:
        return False

    if _looks_like_final_air_exit_window(win):
        return False

    if not looks_like_small_popup_window(win):
        return False

    title_lower = (title or "").lower()
    popup_type = _classify_startup_popup(win, title, hwnd=hwnd, air_pids=air_pids)

    if popup_type in ("notice", "setting", "generic", "jit_continue"):
        return True

    if _popup_has_error_timeout_text(win, title):
        return True

    if _popup_has_close_or_ok_button(win):
        candidate_keywords = [
            "welcome to automatic inspection report",
            "automatic inspection report",
            "newair",
            "air",
            "알림",
            "안내",
            "공지",
            "환경설정",
            "설정",
            "오류",
            "에러",
            "경고",
            "timeout",
            "error",
            "warning",
            "exception",
            "대화상자",
        ]
        all_text = _startup_popup_all_text(win, title)
        if any(keyword in title_lower or keyword in all_text for keyword in candidate_keywords):
            return True

        if hwnd is not None and is_air_owned_window(int(hwnd), title, air_pids=air_pids):
            return True

    return False


def _is_air_settings_popup(win, title: str = "", hwnd: int | None = None, air_pids: set | None = None) -> bool:
    if win is None:
        return False

    if _looks_like_final_air_exit_window(win):
        return False

    title_lower = (title or "").lower()
    if "환경설정" in title or "setting" in title_lower or "setup" in title_lower:
        return True

    try:
        return _classify_startup_popup(win, title, hwnd=hwnd, air_pids=air_pids) == "setting"
    except Exception:
        return False


def _find_air_settings_popup(timeout: float = 0.8, air_pids: set | None = None):
    return _wait_for_air_exit_popup(
        lambda win, hwnd, title: _is_air_settings_popup(
            win,
            title,
            hwnd=hwnd,
            air_pids=air_pids,
        ),
        timeout=timeout,
        air_pids=air_pids,
    )


def _has_air_settings_popup(timeout: float = 0.2, air_pids: set | None = None) -> bool:
    return _find_air_settings_popup(timeout=timeout, air_pids=air_pids) is not None


def _close_air_settings_popup(win, title: str = "") -> bool:
    """
    환경설정 창은 Alt+Q 대상이 아닙니다.
    종료 단계에서는 상단 X를 먼저 누르고, 실패할 때만 내부 닫기 버튼을 누릅니다.
    """
    if _click_window_titlebar_close(win, title):
        log(f"[AIR-CLOSE] 환경설정 X 닫기: title='{title}'")
        return True

    clicked = _click_air_popup_button(
        win,
        target_texts=["닫기", "닫기(&c)", "close"],
        blocked_texts=["계속", "continue", "끝내기", "종료", "exit", "quit"],
    )
    if clicked:
        log(f"[AIR-CLOSE] 환경설정 닫기 버튼 클릭: '{clicked}' / title='{title}'")
        time.sleep(0.2)
        return True

    return False


def _close_crash_at_popup(timeout: float = 2.0, air_pids: set | None = None) -> bool:
    """
    환경설정 닫기 후 뜨는 'Crash at ...' 오류창은 닫기 버튼으로만 닫습니다.
    """
    item = _wait_for_air_exit_popup(
        lambda win, hwnd, title: (
            not _looks_like_final_air_exit_window(win)
            and _popup_has_crash_at_text(win, title)
        ),
        timeout=min(timeout, 0.8),
        air_pids=air_pids,
    )
    if not item:
        return _click_foreground_nonfinal_popup_x(timeout=timeout, label="Crash at/오류")

    win, hwnd, title = item
    if _click_window_titlebar_close(win, title):
        log(f"[AIR-CLOSE] 종료 순서 3/5 Crash at 오류창 X 닫기: title='{title}'")
        return True

    clicked = _click_air_popup_button(
        win,
        target_texts=["닫기", "닫기(&c)", "close"],
        blocked_texts=["계속", "continue", "끝내기", "종료", "exit", "quit", "ok", "확인"],
    )
    if clicked:
        log(f"[AIR-CLOSE] 종료 순서 3/5 Crash at 오류창 닫기 fallback: '{clicked}' / title='{title}'")
        time.sleep(0.2)
        return True

    log(f"[WARN] Crash at 오류창 닫기 실패: title='{title}'")
    return False


def _click_ok_popup_after_crash(timeout: float = 2.0, air_pids: set | None = None) -> bool:
    """
    Crash at 오류창을 닫은 다음 뜨는 OK/확인 팝업을 처리합니다.
    최종 종료 대화상자와 환경설정은 제외합니다.
    """
    item = _wait_for_air_exit_popup(
        lambda win, hwnd, title: (
            not _looks_like_final_air_exit_window(win)
            and not _is_air_settings_popup(win, title, hwnd=hwnd, air_pids=air_pids)
            and _popup_has_close_or_ok_button(win)
        ),
        timeout=timeout,
        air_pids=air_pids,
    )
    if not item:
        return False

    win, hwnd, title = item
    clicked = _click_air_popup_button(
        win,
        target_texts=["ok", "OK", "확인", "확인(&o)"],
        blocked_texts=["계속", "continue", "끝내기", "종료", "exit", "quit", "닫기", "close"],
    )
    if clicked:
        log(f"[AIR-CLOSE] 종료 순서 4/5 OK 팝업 처리: '{clicked}' / title='{title}'")
        time.sleep(0.2)
        return True

    try:
        bring_window_to_front(win)
        try:
            win.set_focus()
        except Exception:
            pass
        send_keys("{ENTER}")
        time.sleep(0.15)
        log(f"[AIR-CLOSE] 종료 순서 4/5 OK 팝업 Enter fallback: title='{title}'")
        return True
    except Exception as e:
        log(f"[WARN] OK 팝업 처리 실패: title='{title}' / {e}")
        return False


def drain_blocking_air_exit_popups(timeout: float = 1.5, air_pids: set | None = None, max_count: int = 5) -> int:
    """
    환경설정과 동시에 뜨는 timeout/오류/경고 팝업처럼 순서 진행을 막는 창을 먼저 제거합니다.
    최종 종료 대화상자는 절대 닫지 않고, 닫기/확인 계열만 사용합니다.
    """
    closed = 0
    start = time.time()

    while closed < max_count and time.time() - start < timeout:
        item = _wait_for_air_exit_popup(
            lambda win, hwnd, title: _is_blocking_nonfinal_air_exit_popup(
                win,
                title,
                hwnd=hwnd,
                air_pids=air_pids,
            ),
            timeout=0.4,
            air_pids=air_pids,
        )
        if not item:
            break

        win, hwnd, title = item

        popup_type = _classify_startup_popup(win, title, hwnd=hwnd, air_pids=air_pids)
        handled = False

        if popup_type == "notice":
            try:
                bring_window_to_front(win)
                try:
                    win.set_focus()
                except Exception:
                    pass
                send_keys("{ENTER}")
                time.sleep(0.12)
                log(f"[AIR-CLOSE] 막힘 알림 팝업 Enter 처리: title='{title}'")
                handled = True
            except Exception as e:
                log(f"[WARN] 막힘 알림 팝업 Enter 실패: title='{title}' / {e}")

        if not handled:
            clicked = _click_any_air_exit_popup_button_or_x(win, title)
            if clicked:
                log(f"[AIR-CLOSE] 막힘 팝업 닫기 처리: '{clicked}' / title='{title}'")
                handled = True

        if not handled and _click_window_titlebar_close(win, title):
            log(f"[AIR-CLOSE] 막힘 팝업 X 닫기 처리: title='{title}'")
            handled = True

        if not handled:
            break

        closed += 1
        time.sleep(0.12)

    return closed


def run_ordered_air_exit_sequence(air_pids: set | None = None, alert_already_handled: bool = False) -> bool:
    """
    AIR 종료 팝업을 사용자가 지정한 순서로 처리합니다.
    1) 메인 AIR X 클릭 후 뜨는 알림: Enter
    2) 환경설정: 닫기/X
    3) Crash at 오류창: 닫기
    4) 다음 OK 팝업: OK/Enter
    5) 마지막 종료 대화상자: Alt+Q
    """
    handled = False

    if alert_already_handled:
        handled = True
    else:
        alert_item = _wait_for_air_exit_popup(
            lambda win, hwnd, title: (
                not _looks_like_final_air_exit_window(win)
                and (
                    is_startup_notification_popup(win, title)
                    or "welcome to automatic inspection report" in title.lower()
                )
            ),
            timeout=0.8,
            air_pids=air_pids,
        )
        if alert_item:
            win, hwnd, title = alert_item
            try:
                bring_window_to_front(win)
                try:
                    win.set_focus()
                except Exception:
                    pass
                send_keys("{ENTER}")
                time.sleep(0.2)
                log(f"[AIR-CLOSE] 종료 순서 1/4 알림 Enter 입력: title='{title}'")
                handled = True
            except Exception as e:
                log(f"[WARN] 종료 순서 알림 Enter 실패: title='{title}' / {e}")

    setting_item = _find_air_settings_popup(timeout=2.0, air_pids=air_pids)
    if setting_item:
        win, hwnd, title = setting_item
        if _close_air_settings_popup(win, title):
            log(f"[AIR-CLOSE] 종료 순서 2/5 환경설정 닫기 완료: title='{title}'")
            handled = True
        else:
            log(f"[WARN] 종료 순서 환경설정 닫기 실패: title='{title}'")

    setting_item = _find_air_settings_popup(timeout=0.6, air_pids=air_pids)
    if setting_item:
        win, hwnd, title = setting_item
        if _close_air_settings_popup(win, title):
            log(f"[AIR-CLOSE] Alt+Q 전 환경설정 추가 닫기 완료: title='{title}'")
            handled = True

    if _has_air_settings_popup(timeout=0.3, air_pids=air_pids):
        log("[AIR-CLOSE] 환경설정이 아직 남아 있어 최종 Alt+Q를 보류합니다.")
        return handled

    crash_closed = _close_crash_at_popup(timeout=2.0, air_pids=air_pids)
    if crash_closed:
        handled = True
        if _click_ok_popup_after_crash(timeout=2.0, air_pids=air_pids):
            handled = True
    else:
        # Crash at 오류창은 발생하지 않을 수도 있습니다. 다른 경고가 끼면 닫고 계속 진행합니다.
        if _close_pending_air_timeout_warning(timeout=0.6, air_pids=air_pids):
            handled = True

    if drain_any_air_exit_popups_by_button_or_x(timeout=0.8, air_pids=air_pids):
        handled = True

    final_item = _wait_for_air_exit_popup(
        lambda win, hwnd, title: _looks_like_final_air_exit_window(win),
        timeout=3.0,
        air_pids=air_pids,
    )
    if final_item:
        win, hwnd, title = final_item
        if _close_final_air_exit_window(win, title):
            log(f"[AIR-CLOSE] 종료 순서 5/5 최종 대화상자 Alt+Q 완료: title='{title}'")
            handled = True
    else:
        # 최종창 제목/버튼 감지가 흔들릴 때를 대비해 기존 보강 스캔도 한 번 수행합니다.
        if close_final_air_exit_window_if_present(timeout=1.0):
            log("[AIR-CLOSE] 종료 순서 5/5 최종 대화상자 Alt+Q 보강 완료")
            handled = True
        elif drain_any_air_exit_popups_by_button_or_x(timeout=1.0, air_pids=air_pids):
            handled = True

    return handled


def _get_cached_air_window_for_close():
    cached = CONTROL_CACHE.get("air_window")
    if cached is None:
        return None

    try:
        title = cached.window_text().strip()
        if title and any(re.search(pattern, title) for pattern in AIR_TITLE_PATTERNS):
            return cached
    except Exception:
        CONTROL_CACHE.pop("air_window", None)

    return None


def _activate_hwnd(hwnd: int) -> bool:
    try:
        ctypes.windll.user32.ShowWindow(wintypes.HWND(int(hwnd)), 9)
        ctypes.windll.user32.SetForegroundWindow(wintypes.HWND(int(hwnd)))
        time.sleep(0.08)
        return True
    except Exception:
        return False


def _send_alt_f4_to_hwnd(hwnd: int, title: str = "") -> bool:
    if not hwnd or is_own_app_window(int(hwnd), title):
        return False

    if not _activate_hwnd(int(hwnd)):
        return False

    try:
        send_keys("%{F4}")
        time.sleep(0.25)
        log(f"[AIR-CLOSE] Alt+F4 입력: title='{title}'")
        return True
    except Exception as e:
        log(f"[WARN] Alt+F4 입력 실패: title='{title}' / {e}")
        return False


def _window_size_by_hwnd(hwnd: int) -> tuple[int, int] | None:
    try:
        rect = wintypes.RECT()
        if ctypes.windll.user32.GetWindowRect(wintypes.HWND(int(hwnd)), ctypes.byref(rect)):
            return int(rect.right - rect.left), int(rect.bottom - rect.top)
    except Exception:
        pass
    return None


def _looks_like_small_popup_hwnd(hwnd: int) -> bool:
    size = _window_size_by_hwnd(hwnd)
    if size is None:
        return False
    width, height = size
    return 80 <= width <= 1100 and 40 <= height <= 750


def _enum_visible_windows_native() -> list[tuple[int, str]]:
    windows = []

    try:
        enum_proc_type = ctypes.WINFUNCTYPE(ctypes.c_bool, wintypes.HWND, wintypes.LPARAM)

        def enum_proc(hwnd, lparam):
            try:
                hwnd_int = int(hwnd)
                if not ctypes.windll.user32.IsWindowVisible(wintypes.HWND(hwnd_int)):
                    return True

                title = _get_window_text_by_hwnd(hwnd_int)
                if title:
                    windows.append((hwnd_int, title))
            except Exception:
                pass
            return True

        ctypes.windll.user32.EnumWindows(enum_proc_type(enum_proc), 0)
    except Exception:
        pass

    return windows


def _native_title_is_air_main(title: str) -> bool:
    return bool(title) and any(re.search(pattern, title) for pattern in AIR_TITLE_PATTERNS)


def _native_title_is_air_related(title: str) -> bool:
    title_lower = (title or "").lower()
    keywords = [
        "ccs air",
        "opc air",
        "air system",
        "automatic inspection report",
        "newair",
        "검사기록조회",
        "검사이력",
    ]
    return any(keyword in title_lower for keyword in keywords) or _native_title_is_air_main(title)


def _native_window_is_final_exit(hwnd: int, title: str = "") -> bool:
    return _native_window_has_exit_button(hwnd) or (
        "welcome to automatic inspection report" in (title or "").lower()
        and _native_window_has_exit_button(hwnd)
    )


def _send_alt_q_to_hwnd(hwnd: int, title: str = "") -> bool:
    if not hwnd or is_own_app_window(int(hwnd), title):
        return False

    if not _activate_hwnd(int(hwnd)):
        return False

    try:
        send_keys("%q")
        time.sleep(0.25)
        log(f"[AIR-CLOSE] 마지막 종료 팝업 Alt+Q 입력: title='{title}'")
        return True
    except Exception as e:
        log(f"[WARN] Alt+Q 입력 실패: title='{title}' / {e}")
        return False


def _find_air_windows_for_alt_f4():
    windows = []
    for hwnd, title in _enum_visible_windows_native():
        try:
            if is_own_app_window(hwnd, title):
                continue
            if _native_title_is_air_main(title) or _native_title_is_air_related(title):
                windows.append((None, hwnd, title))
        except Exception:
            pass
    return windows


def _alt_f4_foreground_nonfinal_popup() -> bool:
    try:
        hwnd, title = get_foreground_window_info()
        title = (title or "").strip()
        if not hwnd or not title or is_own_app_window(int(hwnd), title):
            return False

        if _native_window_is_final_exit(int(hwnd), title):
            return _send_alt_q_to_hwnd(int(hwnd), title)

        if _native_title_is_air_main(title):
            return False

        if not _looks_like_small_popup_hwnd(int(hwnd)):
            return False

        return _send_alt_f4_to_hwnd(int(hwnd), title)
    except Exception:
        return False


def _close_native_final_exit_if_present() -> bool:
    try:
        hwnd, title = get_foreground_window_info()
        title = (title or "").strip()
        if hwnd and title and not is_own_app_window(int(hwnd), title):
            if _native_window_is_final_exit(int(hwnd), title):
                return _send_alt_q_to_hwnd(int(hwnd), title)
    except Exception:
        pass

    for hwnd, title in _enum_visible_windows_native():
        try:
            if is_own_app_window(hwnd, title):
                continue
            if _native_window_is_final_exit(hwnd, title):
                return _send_alt_q_to_hwnd(hwnd, title)
        except Exception:
            pass

    return False


def close_air_after_completion(timeout: int = 10) -> bool:
    """
    자동화 완료 후 AIR 종료를 단순 처리합니다.
    - AIR 메인창과 중간 팝업은 Alt+F4로 통일
    - 마지막 종료 팝업만 Alt+Q
    """
    log("\n[AIR-CLOSE] 실행 완료 후 AIR 창 종료 시작")

    start = time.time()
    acted = False

    while time.time() - start < timeout:
        log("[AIR-CLOSE] 종료 루프 진행 중")

        if _close_native_final_exit_if_present():
            log("[AIR-CLOSE] 마지막 종료 팝업 Alt+Q 처리 완료")
            return True

        if _alt_f4_foreground_nonfinal_popup():
            acted = True
            time.sleep(0.25)
            continue

        air_windows = _find_air_windows_for_alt_f4()

        if air_windows:
            for _win, hwnd, title in air_windows:
                try:
                    if _native_window_is_final_exit(hwnd, title):
                        if _send_alt_q_to_hwnd(hwnd, title):
                            log("[AIR-CLOSE] 마지막 종료 팝업 Alt+Q 처리 완료")
                            return True
                        continue

                    log(f"[AIR-CLOSE] AIR 창/팝업 Alt+F4 시도: {title}")
                    if _send_alt_f4_to_hwnd(hwnd, title):
                        acted = True
                except Exception as e:
                    log(f"[WARN] AIR Alt+F4 처리 실패: title='{title}' / {e}")

            time.sleep(0.25)
            continue

        if acted:
            if _close_native_final_exit_if_present():
                log("[AIR-CLOSE] 마지막 종료 팝업 Alt+Q 처리 완료")
                return True
            log("[AIR-CLOSE] AIR 종료 처리 완료로 판단")
            return True

        log("[AIR-CLOSE] 닫을 AIR 창이 없습니다.")
        return True

    log("[WARN] 제한 시간 안에 AIR 창 종료를 확인하지 못했습니다.")
    return False


def print_current_screen_summary(win):
    log("\n===== 현재 화면 요약 =====")
    log(f"창 제목: {win.window_text()}")
    log(f"창 좌표: {win.rectangle()}")

    log("\n[Button 목록]")
    for i, btn in enumerate(win.descendants(control_type="Button")):
        try:
            log(
                f"Button {i}: "
                f"title='{btn.window_text()}', "
                f"auto_id='{btn.automation_id()}', "
                f"rect={btn.rectangle()}"
            )
        except Exception:
            pass

    log("\n[주요 컨트롤 목록]")
    for i, ctrl in enumerate(win.descendants()):
        try:
            ctype = ctrl.element_info.control_type
            auto_id = ctrl.automation_id()
            title = ctrl.window_text()

            if ctype in ["Edit", "ComboBox", "Custom", "Pane", "Window", "Button"] or auto_id:
                log(
                    f"{i}: type='{ctype}', "
                    f"title='{title}', "
                    f"auto_id='{auto_id}', "
                    f"rect={ctrl.rectangle()}"
                )
        except Exception:
            pass

    log("=========================\n")


def find_child_by_auto_id(parent, auto_id: str):
    candidates = []

    for ctrl in parent.descendants():
        try:
            if ctrl.automation_id() == auto_id and is_real_visible_control(ctrl):
                candidates.append(ctrl)
        except Exception:
            pass

    if not candidates:
        return None

    candidates.sort(key=lambda c: (c.rectangle().top, c.rectangle().left))
    return candidates[0]


def get_child_by_auto_id(parent, auto_id: str, label: str):
    ctrl = find_child_by_auto_id(parent, auto_id)

    if ctrl is None:
        raise RuntimeError(f"{label} 컨트롤을 찾지 못했습니다. auto_id={auto_id}")

    return ctrl


def find_button_by_auto_id(parent, auto_id: str):
    try:
        btn = parent.child_window(auto_id=auto_id, control_type="Button").wrapper_object()
        if is_real_visible_control(btn):
            return btn
    except Exception:
        pass

    candidates = []

    for btn in parent.descendants(control_type="Button"):
        try:
            if btn.automation_id() == auto_id and is_real_visible_control(btn):
                candidates.append(btn)
        except Exception:
            pass

    if not candidates:
        return None

    candidates.sort(key=lambda c: (c.rectangle().top, c.rectangle().left))
    return candidates[0]


def find_button_by_title(parent, title: str):
    candidates = []

    for btn in parent.descendants(control_type="Button"):
        try:
            if btn.window_text().strip() == title and is_real_visible_control(btn):
                candidates.append(btn)
        except Exception:
            pass

    if not candidates:
        return None

    candidates.sort(key=lambda c: (c.rectangle().top, c.rectangle().left))
    return candidates[0]


def is_inspection_history_screen(win) -> bool:
    try:
        title = win.window_text()
    except Exception:
        title = ""

    result = INSPECTION_WINDOW_KEYWORD in title

    log(f"[CHECK] 검사이력 화면 여부: {result}")
    log(f" - 현재 창 제목: {title}")

    return result


def wait_until_inspection_history_screen(timeout=2):
    start = time.time()

    while time.time() - start < timeout:
        try:
            new_win = refresh_air_window()
            title = new_win.window_text()

            if INSPECTION_WINDOW_KEYWORD in title:
                log(f"[OK] 검사이력 화면 진입 확인: {title}")
                return new_win
        except Exception:
            pass

        time.sleep(0.05)

    return None


def click_inspection_history_menu(win):
    log("[STEP] 검사이력 메뉴 버튼 클릭 시도")

    btn = find_button_by_auto_id(win, INSPECTION_BUTTON_AUTO_ID)

    if btn is None:
        btn = find_button_by_title(win, INSPECTION_BUTTON_TITLE)

    if btn is None:
        print_current_screen_summary(win)
        raise RuntimeError("검사이력 메뉴 버튼을 찾지 못했습니다.")

    log("[OK] 검사이력 버튼 발견")
    log(f"title: {btn.window_text()}")
    log(f"automation_id: {btn.automation_id()}")
    log(f"rectangle: {btn.rectangle()}")

    bring_window_to_front(win)

    try:
        log("[TRY] 검사이력 메뉴 invoke")
        btn.invoke()
        time.sleep(0.08)

        new_win = wait_until_inspection_history_screen(timeout=1.5)
        if new_win is not None:
            log("[OK] 검사이력 메뉴 invoke 성공")
            return new_win
    except Exception as e:
        log(f"[WARN] 검사이력 메뉴 invoke 실패: {e}")

    click_control_fast(btn, "검사이력 메뉴")

    new_win = wait_until_inspection_history_screen(timeout=3)
    if new_win is not None:
        log("[OK] 검사이력 메뉴 실제 클릭 성공")
        return new_win

    print_current_screen_summary(win)
    raise RuntimeError("검사이력 메뉴 진입 실패")


def enter_inspection_history_strict(win):
    log("[STEP] 검사이력 화면 진입 시작")

    if is_inspection_history_screen(win):
        log("[OK] 이미 검사기록조회 화면입니다.")
        return win

    win = click_inspection_history_menu(win)

    if is_inspection_history_screen(win):
        log("[OK] 검사이력 화면 확인 완료")
        return win

    print_current_screen_summary(win)
    raise RuntimeError("검사이력 버튼 클릭 후에도 검사기록조회 화면이 확인되지 않았습니다.")


def get_inspection_form(win):
    """
    검사기록조회 폼(frmSetupList_1)을 찾습니다.

    속도 개선:
    - 한 번 찾은 form은 CONTROL_CACHE에 저장합니다.
    - 다음 부서 변경 때는 전체 descendants() 탐색을 하지 않고 캐시를 먼저 사용합니다.
    """
    cached = CONTROL_CACHE.get("inspection_form")

    if cached is not None:
        try:
            if cached.automation_id() == INSPECTION_FORM_AUTO_ID and is_real_visible_control(cached):
                return cached
        except Exception:
            CONTROL_CACHE.pop("inspection_form", None)

    for ctrl in win.descendants():
        try:
            if ctrl.automation_id() == INSPECTION_FORM_AUTO_ID:
                CONTROL_CACHE["inspection_form"] = ctrl
                log("[OK] 검사기록조회 폼 발견 및 캐시 저장: auto_id=frmSetupList_1")
                return ctrl
        except Exception:
            pass

    for ctrl in win.descendants():
        try:
            title = ctrl.window_text().strip()
            ctype = ctrl.element_info.control_type

            if title == INSPECTION_FORM_TITLE and ctype in ["Window", "Pane", "Custom"]:
                CONTROL_CACHE["inspection_form"] = ctrl
                log("[OK] 검사기록조회 폼 발견 및 캐시 저장: title=검사기록조회")
                return ctrl
        except Exception:
            pass

    print_current_screen_summary(win)
    raise RuntimeError("검사기록조회 폼(frmSetupList_1)을 찾지 못했습니다.")


def get_period_group(form):
    cached = CONTROL_CACHE.get("period_group")
    if cached is not None:
        try:
            if cached.automation_id() == PERIOD_GROUP_AUTO_ID and is_real_visible_control(cached):
                return cached
        except Exception:
            CONTROL_CACHE.pop("period_group", None)

    candidates = []

    for ctrl in form.descendants():
        try:
            if ctrl.automation_id() == PERIOD_GROUP_AUTO_ID and is_real_visible_control(ctrl):
                candidates.append(ctrl)
        except Exception:
            pass

    if not candidates:
        raise RuntimeError(f"기간설정 그룹박스({PERIOD_GROUP_AUTO_ID})를 찾지 못했습니다.")

    candidates.sort(key=lambda c: (c.rectangle().top, c.rectangle().left))
    group = candidates[0]
    CONTROL_CACHE["period_group"] = group

    log("[OK] 기간설정 그룹박스 발견")
    log(f" - title: {group.window_text()}")
    log(f" - auto_id: {group.automation_id()}")
    log(f" - rect: {group.rectangle()}")

    return group


def get_period_child_by_auto_id(form, auto_id: str, label: str):
    cache_key = f"period_child:{auto_id}"
    cached = CONTROL_CACHE.get(cache_key)
    if cached is not None:
        try:
            if cached.automation_id() == auto_id and is_real_visible_control(cached):
                return cached
        except Exception:
            CONTROL_CACHE.pop(cache_key, None)

    period_group = get_period_group(form)
    candidates = []

    for ctrl in period_group.descendants():
        try:
            if ctrl.automation_id() == auto_id and is_real_visible_control(ctrl):
                candidates.append(ctrl)
        except Exception:
            pass

    if not candidates:
        raise RuntimeError(f"{label} 컨트롤을 기간설정 그룹박스 안에서 찾지 못했습니다. auto_id={auto_id}")

    candidates.sort(key=lambda c: (c.rectangle().top, c.rectangle().left))
    ctrl = candidates[0]
    CONTROL_CACHE[cache_key] = ctrl

    log(f"[OK] {label} 기간설정 내부 컨트롤 발견")
    log(f" - title: {ctrl.window_text()}")
    log(f" - auto_id: {ctrl.automation_id()}")
    log(f" - class: {ctrl.class_name()}")
    log(f" - rect: {ctrl.rectangle()}")

    return ctrl


def get_condition_panel(form):
    cached = CONTROL_CACHE.get("condition_panel")
    if cached is not None and is_real_visible_control(cached):
        return cached

    candidates = []

    for ctrl in form.descendants():
        try:
            if ctrl.automation_id() == CONDITION_PANEL_AUTO_ID and is_real_visible_control(ctrl):
                candidates.append(ctrl)
        except Exception:
            pass

    if not candidates:
        raise RuntimeError(f"조건 영역 패널({CONDITION_PANEL_AUTO_ID})을 찾지 못했습니다.")

    candidates.sort(key=lambda c: (c.rectangle().top, c.rectangle().left))
    panel = candidates[0]
    CONTROL_CACHE["condition_panel"] = panel

    log("[OK] 조건 영역 panel1 발견")
    log(f" - rect: {panel.rectangle()}")

    return panel


# =========================================================
# 6. ComboBox 제어
# =========================================================

def send_message(hwnd, msg, wparam=0, lparam=0):
    return _SendMessageW(
        wintypes.HWND(int(hwnd)),
        wintypes.UINT(msg),
        WPARAM_T(int(wparam)),
        LPARAM_T(int(lparam)),
    )


def get_control_hwnd(ctrl) -> int:
    try:
        hwnd = int(ctrl.handle)
        if hwnd:
            return hwnd
    except Exception:
        pass

    try:
        hwnd = int(ctrl.element_info.handle)
        if hwnd:
            return hwnd
    except Exception:
        pass

    return 0


def get_combobox_items_by_win32(combo):
    hwnd = get_control_hwnd(combo)
    if not hwnd:
        return []

    count = send_message(hwnd, CB_GETCOUNT, 0, 0)

    if count < 0 or count > 1000:
        return []

    items = []

    for i in range(int(count)):
        text_len = send_message(hwnd, CB_GETLBTEXTLEN, i, 0)

        if text_len < 0 or text_len > 1000:
            items.append("")
            continue

        buf = ctypes.create_unicode_buffer(int(text_len) + 1)
        send_message(hwnd, CB_GETLBTEXT, i, ctypes.addressof(buf))
        items.append(buf.value)

    return items


def get_combobox_cur_sel(combo) -> int:
    hwnd = get_control_hwnd(combo)
    if not hwnd:
        return -1

    return int(send_message(hwnd, CB_GETCURSEL, 0, 0))


def notify_combobox_selection_changed(combo):
    hwnd = get_control_hwnd(combo)

    if not hwnd:
        return

    user32 = ctypes.windll.user32
    parent_hwnd = user32.GetParent(wintypes.HWND(hwnd))
    ctrl_id = user32.GetDlgCtrlID(wintypes.HWND(hwnd))

    if not parent_hwnd:
        return

    for notify_code in (CBN_SELCHANGE, CBN_SELENDOK):
        wparam = (notify_code << 16) | (ctrl_id & 0xFFFF)
        send_message(parent_hwnd, WM_COMMAND, wparam, hwnd)


def set_combobox_index_win32(combo, index: int, label: str) -> bool:
    hwnd = get_control_hwnd(combo)
    if not hwnd:
        return False

    log(f"[INFO] {label} CB_SETCURSEL 인덱스 선택 시도: index={index}")

    result = send_message(hwnd, CB_SETCURSEL, index, 0)
    
    notify_combobox_selection_changed(combo)
    time.sleep(0.05)
    

    cur_sel = get_combobox_cur_sel(combo)

    log(f"[INFO] {label} 선택 후 cur_sel={cur_sel}, target_index={index}, result={result}")

    return cur_sel == index


def select_combobox_index_by_keyboard(combo, index: int, label: str):
    rect = combo.rectangle()
    x = rect.right - 12
    y = int((rect.top + rect.bottom) / 2)

    log(f"[INFO] {label} 드롭다운 실제 클릭 후 키보드 인덱스 선택")
    log(f" - target_index={index}")
    log(f" - click point: ({x}, {y}), rect={rect}")

    real_mouse_click(x, y)
    

    send_keys("{HOME}")
    

    for _ in range(index):
        send_keys("{DOWN}")
        

    send_keys("{ENTER}")
    time.sleep(0.08)

def get_department_combo_fast(form):
    """
    부서명 ComboBox(cboCostCenter)를 빠르게 찾습니다.

    속도 개선:
    - 처음 1회만 form.descendants(control_type="ComboBox")로 탐색합니다.
    - 이후에는 CONTROL_CACHE["department_combo"]를 재사용합니다.
    - 기존처럼 ComboBox 항목 전체를 매번 읽지 않습니다.
    """
    cached = CONTROL_CACHE.get("department_combo")

    if cached is not None:
        try:
            if cached.automation_id() == DEPARTMENT_COMBO_AUTO_ID and is_real_visible_control(cached):
                return cached
        except Exception:
            CONTROL_CACHE.pop("department_combo", None)

    candidates = []

    for ctrl in form.descendants(control_type="ComboBox"):
        try:
            if ctrl.automation_id() == DEPARTMENT_COMBO_AUTO_ID and is_real_visible_control(ctrl):
                candidates.append(ctrl)
        except Exception:
            pass

    if not candidates:
        raise RuntimeError("부서명 ComboBox(cboCostCenter)를 찾지 못했습니다.")

    candidates.sort(key=lambda c: (c.rectangle().top, c.rectangle().left))
    combo = candidates[0]

    CONTROL_CACHE["department_combo"] = combo

    log("[OK] 부서명 ComboBox 발견 및 캐시 저장")
    log(f" - auto_id: {combo.automation_id()}")
    log(f" - rect: {combo.rectangle()}")

    return combo

def safe_select_department_by_index(form, department: str, target_index: int):
    """
    부서명 ComboBox를 인덱스로 빠르게 선택합니다.

    기존 방식:
    - 매번 모든 ComboBox 탐색
    - 매번 ComboBox item 전체 읽기
    - item 텍스트와 department 비교

    개선 방식:
    - 부서 ComboBox 캐시 사용
    - 현재 선택 index가 같으면 바로 skip
    - DEPARTMENT_JOBS의 combo_index를 신뢰하고 바로 CB_SETCURSEL 실행
    """
    combo = get_department_combo_fast(form)

    cur_sel = get_combobox_cur_sel(combo)

    if cur_sel == target_index:
        log(f"[SKIP] 부서명 이미 선택됨: {department} -> index {target_index}")
        return

    if set_combobox_index_win32(combo, target_index, f"부서명 {department}"):
        log(f"[OK] 부서명 선택 완료: {department} -> index {target_index}")
        return

    log(f"[WARN] Win32 인덱스 선택 실패, 키보드 방식으로 재시도: {department}")
    select_combobox_index_by_keyboard(combo, target_index, f"부서명 {department}")
    log(f"[OK] 부서명 키보드 선택 완료: {department} -> index {target_index}")


def safe_select_combobox_by_text(combo, value: str, label: str):
    value = str(value).strip()

    if not value:
        log(f"[SKIP] {label}: 값 없음")
        return

    items = get_combobox_items_by_win32(combo)

    log(f"[INFO] {label} ComboBox 항목: {items}")

    target_index = None

    for i, item in enumerate(items):
        if normalize_text(item) == normalize_text(value):
            target_index = i
            break

    if target_index is not None:
        if set_combobox_index_win32(combo, target_index, label):
            log(f"[OK] {label} 선택 완료: {value}")
            return

        select_combobox_index_by_keyboard(combo, target_index, label)
        log(f"[OK] {label} 키보드 선택 완료: {value}")
        return

    try:
        combo.set_focus()
        time.sleep(0.1)
        combo.click_input()
        time.sleep(0.1)

        send_keys("^a")
        time.sleep(0.05)
        send_keys("{BACKSPACE}")
        time.sleep(0.05)

        set_clipboard_text(value)
        send_keys("^v")
        time.sleep(0.1)
        send_keys("{ENTER}")
        time.sleep(0.2)

        log(f"[OK] {label} 텍스트 입력 방식 완료: {value}")
    except Exception as e:
        raise RuntimeError(f"{label} ComboBox 선택 실패: {e}")


def safe_select_period_time(form, auto_id: str, value: str, label: str):
    combo = get_period_child_by_auto_id(form, auto_id, label)
    safe_select_combobox_by_text(combo, value, label)


# =========================================================
# 7. 날짜 / 조건 입력
# =========================================================

def safe_set_text(control, value: str, field_name: str):
    if value is None or str(value).strip() == "":
        log(f"[SKIP] {field_name}: 입력값 없음")
        return

    value = str(value).strip()

    try:
        control.set_focus()
        time.sleep(0.1)

        try:
            control.set_edit_text(value)
            log(f"[OK] {field_name} set_edit_text 입력 완료: {value}")
            return
        except Exception as e:
            log(f"[WARN] {field_name} set_edit_text 실패, 클립보드 입력으로 재시도: {e}")

        control.click_input()
        time.sleep(0.1)

        send_keys("^a")
        time.sleep(0.05)
        send_keys("{BACKSPACE}")
        time.sleep(0.05)

        set_clipboard_text(value)
        send_keys("^v")
        time.sleep(0.1)
        send_keys("{TAB}")
        time.sleep(0.1)

        log(f"[OK] {field_name} 클립보드 입력 완료: {value}")

    except Exception as e:
        raise RuntimeError(f"{field_name} 입력 실패: {e}")


def set_datetime_picker_by_message(picker, year: str, month: str, day: str, label: str) -> bool:
    hwnd = get_control_hwnd(picker)
    if not hwnd:
        return False

    try:
        st = SYSTEMTIME(
            int(year),
            int(month),
            0,
            int(day),
            0,
            0,
            0,
            0,
        )
        result = _SendMessageW(
            wintypes.HWND(int(hwnd)),
            wintypes.UINT(DTM_SETSYSTEMTIME),
            WPARAM_T(GDT_VALID),
            LPARAM_T(ctypes.addressof(st)),
        )
        if int(result) != 0:
            log(f"[OK] {label} DTM_SETSYSTEMTIME 설정 완료: {year}-{month}-{day}")
            return True
        log(f"[WARN] {label} DTM_SETSYSTEMTIME 결과값 실패: {result}")
    except Exception as e:
        log(f"[WARN] {label} DTM_SETSYSTEMTIME 예외: {e}")

    return False


def set_datetime_picker_by_clipboard(picker, value: str, label: str) -> bool:
    try:
        picker.set_focus()
        time.sleep(0.05)
        picker.click_input()
        time.sleep(0.05)
        send_keys("^a")
        time.sleep(0.03)
        set_clipboard_text(value)
        send_keys("^v")
        time.sleep(0.05)
        send_keys("{TAB}")
        time.sleep(0.05)
        log(f"[OK] {label} 전체 날짜 클립보드 입력 완료: {value}")
        return True
    except Exception as e:
        log(f"[WARN] {label} 전체 날짜 클립보드 입력 실패: {e}")
        return False


def safe_set_datetime_picker(form, auto_id: str, value: str, label: str):
    """
    AIR DateTimePicker 날짜 입력 함수.

    중요:
    - DTM_SETSYSTEMTIME 같은 Win32 메시지 직접 전송은 사용하지 않음
    - 클립보드 전체 날짜 붙여넣기도 사용하지 않음
    - 기본 코드 방식처럼 실제 클릭 후 키보드로 yyyy -> MM -> dd 순서 입력
    """
    if value is None or str(value).strip() == "":
        log(f"[SKIP] {label}: 입력값 없음")
        return

    value = str(value).strip()
    year, month, day = parse_date_ymd(value)

    picker = get_period_child_by_auto_id(form, auto_id, label)

    log(f"[INFO] {label} DateTimePicker 입력 시도: {value}")
    log(f"[INFO] year={year}, month={month}, day={day}")

    rect_info = None
    try:
        rect_info = get_control_rect_tuple_safe(picker, label)
    except NameError:
        rect_info = None
    except Exception:
        rect_info = None

    if rect_info is not None:
        left, top, right, bottom, rect = rect_info
    else:
        rect = picker.rectangle()
        left = int(rect.left)
        top = int(rect.top)
        right = int(rect.right)
        bottom = int(rect.bottom)

    # 기본 코드와 동일하게 DateTimePicker 왼쪽 날짜 영역 클릭
    x = left + 20
    y = int((top + bottom) / 2)

    log(f"[INFO] {label} 실제 클릭 좌표: x={x}, y={y}, rect={rect}")

    if not is_inside_screen(x, y):
        raise RuntimeError(f"{label} DateTimePicker 클릭 좌표가 화면 밖입니다: ({x}, {y})")

    # 실제 마우스 클릭으로 DateTimePicker 포커스
    real_mouse_click(x, y)
    time.sleep(0.15)

    # 기본 방식: HOME으로 첫 필드 이동 후 연/월/일 순서 입력
    send_keys("{HOME}")
    time.sleep(0.05)

    send_keys(year)
    time.sleep(0.05)

    send_keys("{RIGHT}")
    time.sleep(0.05)

    send_keys(month)
    time.sleep(0.05)

    send_keys("{RIGHT}")
    time.sleep(0.05)

    send_keys(day)
    time.sleep(0.05)

    # 다음 컨트롤로 이동해서 입력 확정
    send_keys("{TAB}")
    time.sleep(0.15)

    log(f"[OK] {label} 입력 완료: {value}")


def input_conditions(
    win,
    department: str,
    department_index: int,
    start_date: str,
    end_date: str,
    start_time: str,
    end_time: str,
    product_no: str = "",
    machine_no: str = "",
    lot_no: str = "",
    barcode: str = "",
    tool_no: str = "",
    po_no: str = "",
    disposition: str = "",
    plant: str = "",
    set_period: bool = True,
    set_dates: bool | None = None,
    set_times: bool | None = None,
):
    log("[STEP] 검사이력 조건 입력 시작")

    try:
        bring_window_to_front(win)
        time.sleep(0.1)
    except Exception as e:
        log(f"[WARN] 조건 입력 전 AIR 창 전면화 실패: {e}")

    form = get_inspection_form(win)

    text_values = {
        "제품번호": product_no,
        "기계번호": machine_no,
        "Lot No": lot_no,
        "Barcode": barcode,
        "툴번호": tool_no,
        "PO No": po_no,
        "판정": disposition,
    }

    for label, value in text_values.items():
        auto_id = FIELD_AUTO_IDS[label]
        if value:
            ctrl = get_child_by_auto_id(form, auto_id, label)
            safe_set_text(ctrl, value, label)

    if plant:
        plant_combo = get_child_by_auto_id(form, PLANT_COMBO_AUTO_ID, "플랜트")
        safe_select_combobox_by_text(plant_combo, plant, "플랜트")

    # 부서명은 매번 변경해야 함
    safe_select_department_by_index(form, department, department_index)

    if set_dates is None:
        set_dates = set_period
    if set_times is None:
        set_times = set_period

    if set_dates:
        log("[STEP] 조회 날짜 조건을 설정합니다.")
        safe_set_datetime_picker(form, START_DATE_PICKER_AUTO_ID, start_date, "시작일")
        safe_set_datetime_picker(form, END_DATE_PICKER_AUTO_ID, end_date, "종료일")
    else:
        log("[SKIP] 조회 날짜 조건은 이미 설정되어 있으므로 생략합니다.")

    if set_times:
        log("[STEP] 조회 시간 조건을 설정합니다.")
        safe_select_period_time(form, START_TIME_COMBO_AUTO_ID, start_time, "시작시간")
        safe_select_period_time(form, END_TIME_COMBO_AUTO_ID, end_time, "종료시간")
    else:
        log("[SKIP] 조회 시간 조건은 고정값이 이미 설정되어 있으므로 생략합니다.")

    log("[OK] 검사이력 조건 입력 완료")
    return form


# =========================================================
# 8. 조회 버튼 / 엑셀 버튼
# =========================================================

def get_search_icon_button(form):
    cached = CONTROL_CACHE.get("search_button")
    if cached is not None and is_real_visible_control(cached):
        return cached

    panel = get_condition_panel(form)

    candidates = []

    for btn in panel.descendants(control_type="Button"):
        try:
            if btn.automation_id() == INSPECTION_SEARCH_BUTTON_AUTO_ID and is_real_visible_control(btn):
                candidates.append(btn)
        except Exception:
            pass

    if not candidates:
        raise RuntimeError("panel1 내부에서 기간설정 옆 조회 버튼(button1)을 찾지 못했습니다.")

    candidates.sort(key=lambda c: (c.rectangle().top, c.rectangle().left))
    btn = candidates[0]
    CONTROL_CACHE["search_button"] = btn

    log("[OK] 기간설정 옆 조회 아이콘 발견")
    log(f" - title: {btn.window_text()}")
    log(f" - auto_id: {btn.automation_id()}")
    log(f" - rect: {btn.rectangle()}")

    return btn


def get_excel_button(form):
    cached = CONTROL_CACHE.get("excel_button")
    if cached is not None and is_real_visible_control(cached):
        return cached

    panel = get_condition_panel(form)

    candidates = []

    for btn in panel.descendants(control_type="Button"):
        try:
            if btn.automation_id() == EXCEL_BUTTON_AUTO_ID and is_real_visible_control(btn):
                candidates.append(btn)
        except Exception:
            pass

    if not candidates:
        raise RuntimeError("panel1 내부에서 엑셀 내보내기 버튼(cmdExcel)을 찾지 못했습니다.")

    candidates.sort(key=lambda c: (c.rectangle().top, c.rectangle().left))
    btn = candidates[0]
    CONTROL_CACHE["excel_button"] = btn

    log("[OK] 기간설정 옆 엑셀 내보내기 버튼 발견")
    log(f" - title: {btn.window_text()}")
    log(f" - auto_id: {btn.automation_id()}")
    log(f" - rect: {btn.rectangle()}")

    return btn


def click_search(win, form=None):
    """
    조회 클릭 속도 개선 버전.
    - 매번 refresh_air_window()로 AIR를 다시 찾지 않음
    - 이미 잡아둔 form을 재사용
    - 조회 버튼은 1회만 클릭
    - 조회 완료 팝업이 없으므로 15초 고정 대기 후 엑셀 내보내기 단계로 진행
    """
    if form is None:
        form = get_inspection_form(win)

    search_button = get_search_icon_button(form)
    log("[INFO] 조회 버튼 단일 클릭 실행")
    click_control_fast(search_button, "기간설정 옆 조회 아이콘")

    log(f"[INFO] 조회 완료 팝업 없음 - 조회 결과 로딩 대기: {SEARCH_COMPLETE_TIMEOUT_SECONDS}초")
    time.sleep(SEARCH_COMPLETE_TIMEOUT_SECONDS)
    return True


# =========================================================
# 9. 저장 대화상자 처리 - 기존 성공 방식 복원 + 부서별 저장경로 적용
# =========================================================

def debug_dump_windows_for_save():
    """
    저장창 감지 실패 시 Python이 실제로 보고 있는 창 목록을 출력합니다.
    """
    log("\n===== 저장창 디버그: 현재 감지 가능한 창 목록 =====")

    for backend in ["uia", "win32"]:
        log(f"\n[{backend.upper()} windows]")
        try:
            for w in Desktop(backend=backend).windows():
                try:
                    title = w.window_text().strip()
                    rect = w.rectangle()
                    handle = w.handle

                    if not title:
                        continue

                    try:
                        class_name = w.class_name()
                    except Exception:
                        class_name = ""

                    log(f"hwnd={handle}, class='{class_name}', title='{title}', rect={rect}")
                except Exception:
                    pass
        except Exception as e:
            log(f"[WARN] {backend} 창 목록 출력 실패: {e}")

    log("=================================================\n")


def is_air_main_window_title(title: str) -> bool:
    title = str(title)

    return (
        "CCS AIR System" in title
        or "OPC AIR System" in title
        or "검사기록조회" in title
    )


def find_save_dialog(timeout=8):
    """
    Windows '다른 이름으로 저장' 창을 찾습니다.

    핵심:
    1. UIA와 Win32 backend를 모두 탐색합니다.
    2. AIR 메인창은 저장창 후보에서 제외합니다.
    3. 저장창을 못 찾았다고 해서 foreground 창을 저장창으로 간주하지 않습니다.
       이 부분이 기존 오류의 핵심 원인이었습니다.
    """
    title_patterns = [
        r".*다른 이름으로 저장.*",
        r".*Save As.*",
        r".*Save as.*",
        r".*파일 저장.*",
        r".*저장.*",
    ]

    start = time.time()

    while time.time() - start < timeout:
        for backend in ["uia", "win32"]:
            try:
                for w in Desktop(backend=backend).windows():
                    try:
                        title = w.window_text().strip()

                        if not title:
                            continue

                        try:
                            hwnd = int(w.handle)
                        except Exception:
                            hwnd = id(w)

                        if is_own_app_window(hwnd, title):
                            continue

                        if is_air_main_window_title(title):
                            continue

                        for pattern in title_patterns:
                            if re.search(pattern, title, re.IGNORECASE):
                                log(f"[OK] 저장 대화상자 감지: backend={backend}, title='{title}'")
                                log(f" - rect: {w.rectangle()}")
                                return w

                    except Exception:
                        pass
            except Exception:
                pass

        time.sleep(0.3)

    log("[WARN] 저장 대화상자를 찾지 못했습니다.")
    debug_dump_windows_for_save()
    return None


def bring_dialog_to_front(dialog):
    """
    저장 대화상자를 전면으로 가져옵니다.
    AIR 메인창이 아니라, 이미 찾은 저장 dialog 객체에 대해서만 실행합니다.
    """
    try:
        dialog.restore()
    except Exception:
        pass

    try:
        dialog.set_focus()
    except Exception:
        pass

    try:
        hwnd = int(dialog.handle)
        ctypes.windll.user32.ShowWindow(hwnd, 9)
        ctypes.windll.user32.SetForegroundWindow(hwnd)
    except Exception:
        pass

    time.sleep(0.3)


def find_filename_edit_in_save_dialog(dialog):
    """
    Windows Save As 창에서 파일 이름 입력 Edit를 찾습니다.

    우선순위:
    1. auto_id == '1001'인 Edit
    2. 아래쪽에 있는 넓은 Edit
    3. Win32 class_name='Edit' fallback
    """
    candidates = []

    # 1차: UIA 방식
    try:
        for edit in dialog.descendants(control_type="Edit"):
            try:
                if is_real_visible_control(edit):
                    candidates.append(edit)
            except Exception:
                pass
    except Exception:
        pass

    # 2차: Win32 방식
    if not candidates:
        try:
            for edit in dialog.descendants(class_name="Edit"):
                try:
                    if is_real_visible_control(edit):
                        candidates.append(edit)
                except Exception:
                    pass
        except Exception:
            pass

    if not candidates:
        log("[WARN] 저장창 안에서 Edit 컨트롤을 찾지 못했습니다.")
        return None

    log("\n===== 저장창 Edit 후보 =====")
    for i, edit in enumerate(candidates):
        try:
            log(
                f"[{i}] title='{edit.window_text()}', "
                f"auto_id='{edit.automation_id()}', "
                f"class='{edit.class_name()}', "
                f"rect={edit.rectangle()}"
            )
        except Exception:
            pass
    log("===========================\n")

    # 표준 Windows Save As 파일명 입력칸은 auto_id가 1001인 경우가 많음
    for edit in candidates:
        try:
            if edit.automation_id() == "1001":
                log("[OK] 파일명 입력칸 선택: auto_id=1001")
                return edit
        except Exception:
            pass

    # fallback: 넓고 아래쪽에 있는 Edit를 파일명 입력칸으로 판단
    candidates.sort(
        key=lambda e: (
            -(e.rectangle().width()),
            -e.rectangle().top,
        )
    )

    log("[OK] 파일명 입력칸 fallback 선택")
    log(f" - rect: {candidates[0].rectangle()}")

    return candidates[0]


def set_save_filename(dialog, save_path: str):
    """
    저장창의 파일 이름 입력칸에 전체 저장 경로를 입력합니다.
    기존 성공 코드처럼 Edit 컨트롤을 직접 잡아서 입력합니다.
    """
    save_path = str(Path(save_path).resolve())

    edit = find_filename_edit_in_save_dialog(dialog)

    if edit is not None:
        try:
            edit.set_focus()
            time.sleep(0.2)

            log("[STEP] 파일명 입력칸에 set_edit_text로 전체 경로 입력")
            edit.set_edit_text(save_path)
            time.sleep(0.3)
            return True

        except Exception as e:
            log(f"[WARN] set_edit_text 실패, 클립보드 입력으로 재시도: {e}")

            try:
                edit.click_input()
                time.sleep(0.2)

                send_keys("^a")
                time.sleep(0.1)

                set_clipboard_text(save_path)
                send_keys("^v")
                time.sleep(0.3)

                return True
            except Exception as e2:
                log(f"[WARN] Edit 클립보드 입력 실패: {e2}")

    # 최후 fallback: 저장창에 포커스를 두고 Ctrl+A / 붙여넣기
    try:
        log("[WARN] 파일명 Edit를 직접 제어하지 못해 dialog 포커스 기준으로 붙여넣기 시도")
        dialog.set_focus()
        time.sleep(0.2)

        send_keys("^a")
        time.sleep(0.1)

        set_clipboard_text(save_path)
        send_keys("^v")
        time.sleep(0.3)

        return True
    except Exception as e:
        log(f"[ERROR] 저장 경로 입력 최종 실패: {e}")
        return False


def find_save_button_in_dialog(dialog):
    """
    저장창 내부의 저장 버튼을 찾습니다.
    """
    save_titles = [
        "저장",
        "저장(&S)",
        "&저장",
        "Save",
        "&Save",
    ]

    candidates = []

    # UIA Button 후보
    try:
        for btn in dialog.descendants(control_type="Button"):
            try:
                if not is_real_visible_control(btn):
                    continue

                candidates.append(btn)
            except Exception:
                pass
    except Exception:
        pass

    # Win32 Button 후보
    if not candidates:
        try:
            for btn in dialog.descendants(class_name="Button"):
                try:
                    if not is_real_visible_control(btn):
                        continue

                    candidates.append(btn)
                except Exception:
                    pass
        except Exception:
            pass

    log("\n===== 저장창 Button 후보 =====")
    for i, btn in enumerate(candidates):
        try:
            log(
                f"[{i}] title='{btn.window_text()}', "
                f"auto_id='{btn.automation_id()}', "
                f"class='{btn.class_name()}', "
                f"rect={btn.rectangle()}"
            )
        except Exception:
            pass
    log("============================\n")

    for title in save_titles:
        for btn in candidates:
            try:
                if btn.window_text().strip() == title:
                    log(f"[OK] 저장 버튼 발견: '{title}'")
                    return btn
            except Exception:
                pass

    return None


def click_save_button_in_dialog(dialog):
    """
    저장창 내부 저장 버튼을 클릭합니다.
    실패 시 Alt+S, Enter 순서로 fallback 합니다.

    주의:
    Alt+S를 전역으로 바로 보내지 않고,
    먼저 실제 저장 dialog를 찾은 다음 dialog에 포커스를 둔 상태에서만 사용합니다.
    """
    btn = find_save_button_in_dialog(dialog)

    if btn is not None:
        try:
            click_control_by_real_mouse(btn, "저장 대화상자 저장 버튼")
            time.sleep(1.0)
            return True
        except Exception as e:
            log(f"[WARN] 저장 버튼 실제 클릭 실패: {e}")

    try:
        log("[STEP] 저장 버튼 클릭 실패 fallback: 저장창 포커스 후 Alt+S")
        dialog.set_focus()
        time.sleep(0.2)
        send_keys("%s")
        time.sleep(1.0)
        return True
    except Exception as e:
        log(f"[WARN] Alt+S 저장 실패: {e}")

    try:
        log("[STEP] Alt+S 실패 fallback: Enter")
        dialog.set_focus()
        time.sleep(0.2)
        send_keys("{ENTER}")
        time.sleep(1.0)
        return True
    except Exception as e:
        log(f"[ERROR] Enter 저장 실패: {e}")

    return False


def handle_overwrite_confirm(timeout=5):
    """
    같은 파일명이 있을 때 뜨는 덮어쓰기 확인창에서 예/Yes/확인을 누릅니다.
    """
    start = time.time()

    yes_titles = [
        "예",
        "예(&Y)",
        "&Yes",
        "Yes",
        "확인",
        "OK",
    ]

    while time.time() - start < timeout:
        for backend in ["uia", "win32"]:
            try:
                for w in Desktop(backend=backend).windows():
                    try:
                        title = w.window_text().strip()

                        if not title:
                            continue

                        try:
                            hwnd = int(w.handle)
                        except Exception:
                            hwnd = id(w)

                        if is_own_app_window(hwnd, title):
                            continue

                        if (
                            "확인" in title
                            or "Confirm" in title
                            or "덮어쓰기" in title
                            or "대체" in title
                            or "저장" in title
                            or "Save" in title
                        ):
                            for btn_title in yes_titles:
                                btn = None

                                try:
                                    btn = find_button_by_title(w, btn_title)
                                except Exception:
                                    pass

                                if btn is not None:
                                    click_control_by_real_mouse(btn, f"덮어쓰기 확인 버튼 '{btn_title}'")
                                    time.sleep(0.3)
                                    return True

                            try:
                                w.set_focus()
                                time.sleep(0.2)
                                send_keys("{ENTER}")
                                time.sleep(0.7)
                                return True
                            except Exception:
                                pass

                    except Exception:
                        pass
            except Exception:
                pass

        time.sleep(0.3)

    return False



def _wrap_window_by_handle(hwnd):
    """foreground hwnd를 UIA/Win32 wrapper로 감싸서 반환"""
    for backend in ["uia", "win32"]:
        try:
            return Desktop(backend=backend).window(handle=int(hwnd))
        except Exception:
            pass
    return None


def _window_has_keywords_shallow(win, keywords):
    try:
        title = win.window_text().strip()
        if any(k in title for k in keywords):
            return True
    except Exception:
        pass

    # 전체 데스크톱이 아니라 해당 작은 팝업 내부만 검사하므로 빠름
    try:
        for ctrl in win.descendants():
            try:
                text = ctrl.window_text().strip()
                if text and any(k in text for k in keywords):
                    return True
            except Exception:
                pass
    except Exception:
        pass

    return False


def _looks_like_small_air_popup(win):
    try:
        title = win.window_text().strip()
        rect = win.rectangle()
        is_small = rect.width() < 800 and rect.height() < 500
        title_ok = any(k in title for k in ["검사기록조회", "검사이력", "AIR", "알림", "확인", "Information"])
        return is_small and title_ok
    except Exception:
        return False


def handle_air_message_popup_fast(
    label: str,
    keywords,
    timeout: float = 1.2,
    accept_looks_popup: bool = True,
) -> bool:
    """
    조회 완료/저장 완료 같은 AIR 모달 팝업을 빠르게 닫습니다.
    기존 handle_air_save_complete_popup은 전체 창을 UIA/Win32로 길게 스캔해서 느렸으므로,
    foreground 작은 팝업을 먼저 확인하고 Enter로 닫습니다.
    """
    start = time.time()

    while time.time() - start < timeout:
        try:
            hwnd = ctypes.windll.user32.GetForegroundWindow()
            win = _wrap_window_by_handle(hwnd)

            if win is not None:
                try:
                    title = win.window_text().strip()
                except Exception:
                    title = ""

                if is_own_app_window(hwnd, title):
                    time.sleep(0.1)
                    continue

                has_keyword = _window_has_keywords_shallow(win, keywords)
                looks_popup = _looks_like_small_air_popup(win)

                if has_keyword or (accept_looks_popup and looks_popup):
                    log(f"[OK] {label} 팝업 빠른 감지: title='{title}'")

                    try:
                        win.set_focus()
                    except Exception:
                        pass

                    send_keys("{ENTER}")
                    time.sleep(0.2)
                    return True
        except Exception:
            pass

        time.sleep(0.1)

    log(f"[INFO] {label} 팝업 빠른 감지 없음")
    return False


def _close_message_popup_with_enter(win, label: str) -> bool:
    try:
        title = win.window_text().strip()
    except Exception:
        title = ""

    log(f"[OK] {label} 팝업 감지: title='{title}'")

    try:
        win.set_focus()
    except Exception:
        pass

    send_keys("{ENTER}")
    time.sleep(0.2)
    return True


def handle_air_search_complete_popup(timeout: float = SEARCH_COMPLETE_TIMEOUT_SECONDS) -> bool:
    """
    조회 완료 팝업 전용 처리.
    작은 AIR 팝업이라는 이유만으로 완료 처리하지 않고, 실제 완료 문구가 있을 때만 닫습니다.
    """
    start = time.time()
    next_full_scan = 0.0

    while time.time() - start < timeout:
        try:
            hwnd = ctypes.windll.user32.GetForegroundWindow()
            win = _wrap_window_by_handle(hwnd)

            if win is not None:
                try:
                    title = win.window_text().strip()
                except Exception:
                    title = ""

                if not is_own_app_window(hwnd, title) and _window_has_keywords_shallow(win, SEARCH_COMPLETE_KEYWORDS):
                    return _close_message_popup_with_enter(win, "조회 완료")
        except Exception:
            pass

        now = time.time()
        if now >= next_full_scan:
            next_full_scan = now + 0.35

            for backend in ["uia", "win32"]:
                try:
                    for w in Desktop(backend=backend).windows():
                        try:
                            title = w.window_text().strip()

                            if not title:
                                continue

                            try:
                                hwnd = int(w.handle)
                            except Exception:
                                hwnd = id(w)

                            if is_own_app_window(hwnd, title):
                                continue

                            if not _looks_like_small_air_popup(w):
                                continue

                            if _window_has_keywords_shallow(w, SEARCH_COMPLETE_KEYWORDS):
                                log(f"[INFO] 조회 완료 팝업 전체 탐색 감지: backend={backend}")
                                return _close_message_popup_with_enter(w, "조회 완료")
                        except Exception:
                            pass
                except Exception:
                    pass

        time.sleep(0.1)

    log("[INFO] 조회 완료 팝업 감지 없음")
    return False


def handle_save_dialog_if_present(save_path: str) -> bool:
    """
    저장 대화상자 처리.

    기존 성공 방식 기준:
    1. 저장창을 UIA/Win32 둘 다에서 실제로 찾음
    2. AIR 메인창은 저장창 후보에서 제외
    3. foreground fallback 사용하지 않음
    4. 파일명 Edit를 직접 찾아 전체 경로 입력
    5. 저장창 내부 저장 버튼 클릭
    6. 실패 시에만 Alt+S/Enter fallback
    """
    if not HANDLE_SAVE_DIALOG:
        return False

    save_path = str(Path(save_path).resolve())

    log("[INFO] 저장할 전체 경로:")
    log(f"       {save_path}")

    os.makedirs(os.path.dirname(save_path), exist_ok=True)

    dialog = find_save_dialog(timeout=SAVE_DIALOG_WAIT_SECONDS)

    if dialog is None:
        log("[WARN] 저장 대화상자가 감지되지 않아 저장 처리를 건너뜁니다.")
        log("[WARN] 이 경우 Python 권한, AIR 권한, 저장창 제목, UIA/Win32 인식 여부를 확인해야 합니다.")
        return False

    bring_dialog_to_front(dialog)

    if not set_save_filename(dialog, save_path):
        raise RuntimeError("저장 대화상자 파일명 입력칸에 저장 경로를 입력하지 못했습니다.")

    if not click_save_button_in_dialog(dialog):
        raise RuntimeError("저장 대화상자에서 저장 버튼 클릭 또는 Alt+S 저장에 실패했습니다.")

    handle_overwrite_confirm(timeout=0.6)

    if wait_until_excel_file_ready(save_path, timeout=45, label="저장된 엑셀 파일"):
        log("[OK] 엑셀 파일 저장 완료:")
        log(f"     {save_path}")
        return True

    log("[INFO] 저장 명령은 실행했지만 지정 경로에서 파일 생성을 아직 확인하지 못했습니다.")
    log(f"       예상 경로: {save_path}")
    return True


def click_excel_export(win, export_save_path: str, form=None) -> str:
    """
    엑셀 내보내기 속도 개선 버전.
    - refresh_air_window()/get_inspection_form() 중복 호출 제거
    - 저장창 대기 시간을 짧게 유지
    - 파일 저장 확인 후 저장 완료 팝업을 foreground 기준으로 즉시 Enter 처리
    """
    before_snapshot = snapshot_excel_files()

    if form is None:
        form = get_inspection_form(win)

    excel_button = get_excel_button(form)
    click_control_fast(excel_button, "기간설정 옆 엑셀 내보내기 버튼")

    # 저장창이 뜰 최소 시간만 확보
    time.sleep(POST_EXPORT_DIALOG_WAIT_SECONDS)

    handled = handle_save_dialog_if_present(export_save_path)

    if wait_until_excel_file_ready(export_save_path, timeout=45, label="엑셀 내보내기 파일"):
        log(f"[OK] 엑셀 내보내기 파일 확인: {export_save_path}")

        # 빠른 방식으로 먼저 닫고, 실패할 때만 기존 긴 스캔 fallback
        if not handle_air_message_popup_fast(
            label="저장 완료",
            keywords=["저장", "완료", "저장이 완료", "저장되었습니다"],
            timeout=FAST_POPUP_TIMEOUT_SECONDS,
        ):
            handle_air_save_complete_popup(timeout=0.8)

        return export_save_path

    new_file = wait_for_new_excel_file(before_snapshot, timeout=25)

    if new_file:
        final_path = copy_or_rename_export_file(new_file, export_save_path)
        if not wait_until_excel_file_ready(final_path, timeout=45, label="자동 생성 엑셀 파일"):
            raise RuntimeError(
                "자동 생성된 엑셀 파일이 아직 완전히 저장되지 않았거나 읽을 수 없습니다.\n"
                f"파일 경로: {final_path}"
            )
        log(f"[OK] 자동 생성된 엑셀 파일 확인: {new_file}")
        log(f"[OK] 테스트용 파일 경로: {final_path}")

        if not handle_air_message_popup_fast(
            label="저장 완료",
            keywords=["저장", "완료", "저장이 완료", "저장되었습니다"],
            timeout=FAST_POPUP_TIMEOUT_SECONDS,
        ):
            handle_air_save_complete_popup(timeout=0.8)

        return final_path

    if handled:
        handle_air_message_popup_fast(
            label="저장 완료",
            keywords=["저장", "완료", "저장이 완료", "저장되었습니다"],
            timeout=0.8,
        )
        raise RuntimeError(
            "저장 명령은 실행했지만 파일 생성 확인에 실패했습니다.\n"
            f"예상 경로: {export_save_path}"
        )

    handle_air_message_popup_fast(
        label="저장 완료",
        keywords=["저장", "완료", "저장이 완료", "저장되었습니다"],
        timeout=0.8,
    )

    raise RuntimeError(
        "엑셀 내보내기 파일을 확인하지 못했습니다.\n"
        f"예상 경로: {export_save_path}\n"
        "저장 대화상자 감지 또는 자동 저장 확인이 실패했습니다."
    )


def window_contains_any_text(win, keywords):
    """
    창 내부 텍스트에 특정 문구가 있는지 확인.
    예: '저장이 완료되었습니다'
    """
    try:
        title = win.window_text().strip()
        for keyword in keywords:
            if keyword in title:
                return True
    except Exception:
        pass

    try:
        for ctrl in win.descendants():
            try:
                text = ctrl.window_text().strip()
                if not text:
                    continue

                for keyword in keywords:
                    if keyword in text:
                        return True
            except Exception:
                pass
    except Exception:
        pass

    return False


def find_ok_button_in_dialog(dialog):
    """
    팝업 내부의 OK / 확인 버튼을 찾는다.
    """
    ok_keywords = [
        "OK",
        "Ok",
        "ok",
        "확인",
        "예",
        "Yes",
    ]

    # UIA Button 탐색
    try:
        for btn in dialog.descendants(control_type="Button"):
            try:
                if not is_real_visible_control(btn):
                    continue

                text = btn.window_text().strip()

                if any(k.lower() == text.lower() for k in ok_keywords):
                    return btn

                if any(k.lower() in text.lower() for k in ok_keywords):
                    return btn
            except Exception:
                pass
    except Exception:
        pass

    # Win32 Button 탐색 fallback
    try:
        for btn in dialog.descendants(class_name="Button"):
            try:
                if not is_real_visible_control(btn):
                    continue

                text = btn.window_text().strip()

                if any(k.lower() == text.lower() for k in ok_keywords):
                    return btn

                if any(k.lower() in text.lower() for k in ok_keywords):
                    return btn
            except Exception:
                pass
    except Exception:
        pass

    return None


def handle_air_save_complete_popup(timeout=10):
    """
    AIR 저장 완료 팝업 처리.

    저장 후 AIR에서 '저장이 완료되었습니다' 팝업이 뜨고,
    OK 버튼을 눌러야 다음 부서 조회가 가능하므로 반드시 처리해야 함.
    """
    complete_keywords = [
        "저장이 완료되었습니다",
        "저장 완료",
        "저장되었습니다",
        "완료되었습니다",
        "저장이 완료",
    ]

    popup_title_keywords = [
        "검사기록조회",
        "검사이력",
        "AIR",
        "알림",
        "확인",
        "Information",
    ]

    start = time.time()

    while time.time() - start < timeout:
        for backend in ["uia", "win32"]:
            try:
                for w in Desktop(backend=backend).windows():
                    try:
                        title = w.window_text().strip()
                        rect = w.rectangle()

                        if not title:
                            continue

                        try:
                            hwnd = int(w.handle)
                        except Exception:
                            hwnd = id(w)

                        if is_own_app_window(hwnd, title):
                            continue

                        # 1순위: 창 내부에 저장 완료 문구가 있는 경우
                        has_complete_text = window_contains_any_text(w, complete_keywords)

                        # 2순위: 검사기록조회 제목의 작은 팝업인 경우
                        # AIR 메인 화면보다 팝업은 보통 훨씬 작음
                        is_small_popup = rect.width() < 700 and rect.height() < 400
                        title_looks_like_popup = any(k in title for k in popup_title_keywords)

                        if not has_complete_text and not (is_small_popup and title_looks_like_popup):
                            continue

                        log("[OK] AIR 저장 완료 팝업 감지")
                        log(f" - backend: {backend}")
                        log(f" - title: {title}")
                        log(f" - rect: {rect}")

                        try:
                            w.set_focus()
                            time.sleep(0.2)
                        except Exception:
                            pass

                        ok_btn = find_ok_button_in_dialog(w)

                        if ok_btn is not None:
                            log("[STEP] 저장 완료 팝업 OK 버튼 클릭")
                            click_control_by_real_mouse(ok_btn, "저장 완료 팝업 OK 버튼")
                            time.sleep(0.3)
                            return True

                        # 버튼을 못 찾으면 Enter로 닫기
                        log("[STEP] OK 버튼을 못 찾아 Enter로 저장 완료 팝업 닫기")
                        try:
                            w.set_focus()
                            time.sleep(0.2)
                            send_keys("{ENTER}")
                            time.sleep(0.3)
                            return True
                        except Exception as e:
                            log(f"[WARN] Enter로 저장 완료 팝업 닫기 실패: {e}")

                    except Exception:
                        pass
            except Exception:
                pass

        time.sleep(0.3)

    log("[INFO] AIR 저장 완료 팝업이 감지되지 않았습니다.")
    return False


# =========================================================
# 10. 마스터 엑셀 업데이트
# =========================================================

def normalize_excel_text(value):
    """
    엑셀 셀 값을 비교용 문자열로 정규화합니다.
    None, 공백, 날짜/숫자 표현 차이로 인한 불필요한 중복판정 오류를 줄이기 위한 함수입니다.
    """
    if value is None:
        return ""

    # datetime/date는 ISO 형태로 통일
    try:
        import datetime
        if isinstance(value, (datetime.datetime, datetime.date)):
            return value.isoformat(sep=" ").strip()
    except Exception:
        pass

    text = str(value).strip()

    # 엑셀에서 숫자 1.0으로 읽히는 경우 1로 통일
    if re.fullmatch(r"-?\d+\.0", text):
        text = text[:-2]

    return text


def normalize_excel_header(value):
    """
    헤더명 매핑용 정규화.
    예: '툴번호 ▲'와 '툴번호'를 같은 컬럼으로 보기 위해 정렬표시 문자를 제거합니다.
    """
    text = normalize_excel_text(value)
    text = text.replace("▲", "").replace("▼", "")
    text = re.sub(r"\s+", "", text)
    return text.lower()


def get_effective_max_row(ws, max_col=None):
    """
    서식만 남아 있는 빈 행을 제외하고 실제 값이 있는 마지막 행을 찾습니다.
    """
    if max_col is None:
        max_col = ws.max_column

    for row in range(ws.max_row, 0, -1):
        for col in range(1, max_col + 1):
            if ws.cell(row=row, column=col).value not in (None, ""):
                return row

    return 1


def get_effective_max_col(ws, header_row=1):
    """
    헤더가 있는 마지막 열을 찾습니다.
    A열 헤더가 비어 있어도 기존 사용범위는 유지해야 하므로 ws.max_column을 기본으로 사용합니다.
    """
    max_col = ws.max_column

    # 너무 오른쪽에 서식만 있는 경우를 대비해, 헤더 또는 데이터가 있는 마지막 열을 보정
    for col in range(ws.max_column, 0, -1):
        has_value = False
        for row in range(1, min(ws.max_row, 50) + 1):
            if ws.cell(row=row, column=col).value not in (None, ""):
                has_value = True
                break
        if has_value:
            max_col = col
            break

    return max_col


def build_header_map(ws, header_row=1, max_col=None):
    """
    헤더명 -> 열 번호 매핑.
    A열처럼 헤더가 비어 있는 열은 제외합니다.
    """
    if max_col is None:
        max_col = get_effective_max_col(ws, header_row)

    result = {}

    for col in range(1, max_col + 1):
        header = ws.cell(row=header_row, column=col).value
        key = normalize_excel_header(header)

        if key:
            result[key] = col

    return result


def parse_master_recent_datetime(value):
    """
    마스터 엑셀의 '최근검사' 값을 datetime으로 변환합니다.
    - Excel 날짜 셀(datetime/date)
    - Excel serial number
    - 문자열 날짜/시간(2026-05-31 20:30:00, 2026/05/31, 2026.05.31, 오전/오후 포함)
    을 최대한 처리합니다.
    """
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel
            converted = from_excel(value)
            if isinstance(converted, datetime):
                return converted
            if isinstance(converted, date):
                return datetime.combine(converted, datetime.min.time())
        except Exception:
            pass

    text = str(value).strip()
    if not text:
        return None

    # 날짜 추출: 2026-05-31 / 2026.05.31 / 2026/05/31 / 26-05-31 대응
    date_match = re.search(r"(\d{2,4})[\-./년\s]+(\d{1,2})[\-./월\s]+(\d{1,2})", text)
    if not date_match:
        return None

    year = int(date_match.group(1))
    month = int(date_match.group(2))
    day = int(date_match.group(3))

    if year < 100:
        year += 2000

    hour = 0
    minute = 0
    second = 0

    time_match = re.search(r"(\d{1,2})\s*:\s*(\d{1,2})(?:\s*:\s*(\d{1,2}))?", text)
    if time_match:
        hour = int(time_match.group(1))
        minute = int(time_match.group(2))
        second = int(time_match.group(3) or 0)

        if "오후" in text and hour < 12:
            hour += 12
        if "오전" in text and hour == 12:
            hour = 0

    try:
        return datetime(year, month, day, hour, minute, second)
    except ValueError:
        return None


def get_master_global_latest_recent_query_plan():
    """
    AIR 조회 전에 반드시 마스터 엑셀을 먼저 확인합니다.

    요청 로직:
    1. 마스터 엑셀 List 시트를 먼저 확인
    2. 최근검사 열의 값을 수집
    3. 각 부서별 최근검사 최댓값을 계산
    4. 각 부서의 최신 최근검사 날짜를 해당 부서 AIR 조회 시작일 From으로 사용
    5. 조회 종료일 To는 항상 오늘 날짜 사용

    주의:
    - 마스터 엑셀 파일 자체를 실제로 정렬하지 않습니다.
    - 코드 내부에서만 부서별 최근검사 최댓값을 계산합니다.
    - 날짜 입력 방식은 기존 방식 그대로 사용합니다.
    """
    today = date.today().strftime("%Y-%m-%d")
    fallback_start = (date.today() - timedelta(days=QUERY_START_FALLBACK_DAYS)).strftime("%Y-%m-%d")

    log("\n" + "=" * 80)
    log("[MASTER-CHECK-START] AIR 실행 전 마스터 엑셀 먼저 확인 시작")
    log(f"[MASTER-CHECK] 설정값 USE_MASTER_LATEST_RECENT_DATE_FOR_FROM={USE_MASTER_LATEST_RECENT_DATE_FOR_FROM}")
    log(f"[MASTER-CHECK] 마스터 파일 경로: {MASTER_EXCEL_PATH}")
    log(f"[MASTER-CHECK] 대상 시트명: {MASTER_SHEET_NAME}")
    log(f"[MASTER-CHECK] 최근검사 기준 헤더명: {MASTER_RECENT_INSPECTION_HEADER}")
    log(f"[MASTER-CHECK] 조회 종료일 To 기준: 오늘 날짜 {today}")
    log("=" * 80)

    dept_latest = {
        job["department"]: {
            "dt": None,
            "row": None,
            "raw": None,
            "count": 0,
            "parsed_count": 0,
        }
        for job in DEPARTMENT_JOBS
    }
    source = "master_department_recent"

    if not USE_MASTER_LATEST_RECENT_DATE_FOR_FROM:
        source = "fallback_disabled"
        log(f"[WARN] 마스터 최근검사 기준 조회가 비활성화되어 부서별 fallback From={fallback_start} 사용")
    else:
        master_path = Path(MASTER_EXCEL_PATH)
        log(f"[MASTER-CHECK] 마스터 파일 존재 여부 확인: {master_path.exists()}")

        if not master_path.exists():
            raise FileNotFoundError(f"마스터 엑셀 파일을 찾지 못했습니다: {master_path}")

        log("[MASTER-CHECK] 마스터 엑셀 openpyxl 로드 시작")
        wb = None
        temp_copy_path = None

        wb, temp_copy_path = load_workbook_resilient(master_path, data_only=True)

        try:
            log(f"[MASTER-CHECK] 마스터 엑셀 로드 완료. 시트 목록: {wb.sheetnames}")

            if MASTER_SHEET_NAME in wb.sheetnames:
                ws = wb[MASTER_SHEET_NAME]
                log(f"[MASTER-CHECK] 지정 시트 사용: {ws.title}")
            else:
                ws = wb.active
                log(f"[WARN] '{MASTER_SHEET_NAME}' 시트를 찾지 못해 활성 시트 '{ws.title}'를 사용합니다.")

            max_col = get_effective_max_col(ws, header_row=1)
            last_row = get_effective_max_row(ws, max_col)
            header_map = build_header_map(ws, header_row=1, max_col=max_col)

            dept_col = header_map.get(normalize_excel_header("부서명"), MASTER_DEPARTMENT_COL)
            recent_col = header_map.get(normalize_excel_header(MASTER_RECENT_INSPECTION_HEADER))

            log("\n[MASTER-CHECK] 마스터 구조 확인 결과")
            log(f" - 사용 시트: {ws.title}")
            log(f" - 유효 마지막 행: {last_row}")
            log(f" - 유효 마지막 열: {max_col}")
            log(f" - 부서명 열: {dept_col}열")
            log(f" - 최근검사 열: {recent_col if recent_col is not None else 'NOT FOUND'}")

            if recent_col is None:
                raise RuntimeError(
                    f"마스터 엑셀에서 '{MASTER_RECENT_INSPECTION_HEADER}' 열을 찾지 못했습니다. "
                    "1행 헤더명을 확인하세요."
                )

            target_departments = {normalize_text(job["department"]): job["department"] for job in DEPARTMENT_JOBS}

            log("\n[MASTER-CHECK] 최근검사 열 스캔 시작")
            log(f" - 기준: {', '.join(job['department'] for job in DEPARTMENT_JOBS)} 부서별 데이터")
            log(" - 처리: 각 부서별 최근검사 최댓값을 조회 시작일로 사용")

            for row in range(2, last_row + 1):
                dept_value = ws.cell(row=row, column=dept_col).value
                dept_norm = normalize_text(dept_value)

                if dept_norm not in target_departments:
                    continue

                department = target_departments[dept_norm]
                dept_latest[department]["count"] += 1

                recent_value = ws.cell(row=row, column=recent_col).value
                recent_dt = parse_master_recent_datetime(recent_value)

                if recent_dt is None:
                    continue

                dept_latest[department]["parsed_count"] += 1

                if dept_latest[department]["dt"] is None or recent_dt > dept_latest[department]["dt"]:
                    dept_latest[department]["dt"] = recent_dt
                    dept_latest[department]["row"] = row
                    dept_latest[department]["raw"] = recent_value

            log("\n[MASTER-CHECK] 부서별 최근검사 최댓값")
            for job in DEPARTMENT_JOBS:
                department = job["department"]
                info = dept_latest[department]
                if info["dt"] is None:
                    log(
                        f" - {department}: 최근검사 유효값 없음 "
                        f"/ 부서 행 수={info['count']} / 날짜 파싱 성공={info['parsed_count']} "
                        f"/ fallback From={fallback_start} {DAILY_START_TIME}"
                    )
                else:
                    log(
                        f" - {department}: 최신 최근검사={info['dt']} "
                        f"/ 행={info['row']} / 원본값={info['raw']} "
                        f"/ 부서 행 수={info['count']} / 날짜 파싱 성공={info['parsed_count']}"
                    )

        finally:
            close_workbook_and_cleanup_temp(wb, temp_copy_path)
            log("[MASTER-CHECK] 마스터 엑셀 닫기 완료")

    query_plan = {}

    log("\n[MASTER-CHECK] AIR 부서별 조회 계획")
    for job in DEPARTMENT_JOBS:
        department = job["department"]
        info = dept_latest[department]
        latest_dt = info["dt"]
        start_date = latest_dt.strftime("%Y-%m-%d") if latest_dt else fallback_start
        if latest_dt:
            dept_source = source
        elif source == "fallback_disabled":
            dept_source = source
        else:
            dept_source = "fallback_no_department_recent"

        query_plan[department] = {
            "start_date": start_date,
            "end_date": today,
            "latest_recent": latest_dt,
            "source": dept_source,
        }

        log(
            f"[PLAN] {department}: From={start_date} {DAILY_START_TIME}, "
            f"To={today} {DAILY_END_TIME}, source={dept_source}"
        )

    log("[MASTER-CHECK-END] 마스터 기준 조회기간 산정 완료. 이제 AIR 실행/조회 단계로 이동합니다.")
    log("=" * 80 + "\n")
    return query_plan

def build_master_row_key(ws, row, start_col, end_col):
    """
    중복 판정 키 생성.
    A열 순번은 제외하고 B~마지막열 값으로 비교합니다.
    """
    return tuple(
        normalize_excel_text(ws.cell(row=row, column=col).value)
        for col in range(start_col, end_col + 1)
    )


def build_existing_keys(ws, compare_start_col, end_col):
    """
    마스터 전체 기존 데이터의 중복 키 집합을 생성합니다.
    부서명이 B열에 포함되므로 전체 기준으로 비교해도 부서별 중복 판정이 가능합니다.
    """
    last_row = get_effective_max_row(ws, end_col)
    keys = set()

    for row in range(2, last_row + 1):
        key = build_master_row_key(ws, row, compare_start_col, end_col)
        if any(key):
            keys.add(key)

    return keys

def remove_duplicate_rows_in_master(ws, compare_start_col, end_col):
    """
    마스터 시트 전체에서 중복 행을 최종 제거합니다.

    기준:
    - 1행은 헤더이므로 제외
    - A열 순번은 제외
    - B열부터 마지막 열까지 값이 모두 같으면 중복으로 판단
    - 먼저 나온 행은 유지
    - 나중에 나온 중복 행은 삭제
    """
    last_row = get_effective_max_row(ws, end_col)

    seen_keys = set()
    duplicate_rows = []

    for row in range(2, last_row + 1):
        key = build_master_row_key(
            ws=ws,
            row=row,
            start_col=compare_start_col,
            end_col=end_col,
        )

        # 완전히 빈 행은 중복 판단 대상에서 제외
        if not any(key):
            continue

        if key in seen_keys:
            duplicate_rows.append(row)
        else:
            seen_keys.add(key)

    if not duplicate_rows:
        log("[DUP-CLEAN] 마스터 전체 중복 행 없음")
        return 0

    # 아래에서 위로 삭제해야 행 번호 밀림 문제가 없음
    for row in reversed(duplicate_rows):
        ws.delete_rows(row, 1)

    log(f"[DUP-CLEAN] 마스터 전체 중복 행 삭제 완료: {len(duplicate_rows)}건")
    return len(duplicate_rows)

def find_department_last_row(ws, department, dept_col=MASTER_DEPARTMENT_COL, max_col=None):
    """
    마스터에서 특정 부서명의 마지막 행을 찾습니다.
    """
    if max_col is None:
        max_col = ws.max_column

    target = normalize_text(department)
    last_row = get_effective_max_row(ws, max_col)

    for row in range(last_row, 1, -1):
        value = ws.cell(row=row, column=dept_col).value
        if normalize_text(value) == target:
            return row

    return None


def copy_row_style(ws, src_row, dst_row, max_col):
    """
    기존 부서 마지막 행의 서식을 새 삽입 행에 복사합니다.
    값은 복사하지 않고 스타일만 복사합니다.
    """
    from copy import copy

    if src_row is None or src_row < 1:
        return

    try:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    except Exception:
        pass

    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)

        if src.has_style:
            dst._style = copy(src._style)

        if src.number_format:
            dst.number_format = src.number_format

        if src.alignment:
            dst.alignment = copy(src.alignment)

        if src.font:
            dst.font = copy(src.font)

        if src.fill:
            dst.fill = copy(src.fill)

        if src.border:
            dst.border = copy(src.border)

        if src.protection:
            dst.protection = copy(src.protection)


def reindex_department_numbers(ws, department, index_col=MASTER_INDEX_COL, dept_col=MASTER_DEPARTMENT_COL, max_col=None):
    """
    특정 부서 데이터의 A열 순번을 1부터 다시 부여합니다.
    Daily 데이터 삽입 후에도 부서별 순번이 끊기지 않도록 하기 위한 로직입니다.
    """
    if max_col is None:
        max_col = ws.max_column

    target = normalize_text(department)
    last_row = get_effective_max_row(ws, max_col)
    seq = 1

    for row in range(2, last_row + 1):
        dept_value = ws.cell(row=row, column=dept_col).value
        if normalize_text(dept_value) == target:
            ws.cell(row=row, column=index_col).value = seq
            seq += 1

    return seq - 1


def get_daily_department_from_filename(file_path):
    """
    saved_exports가 문자열 경로만 들어온 경우 파일명에서 부서명을 보조 추출합니다.
    """
    name = Path(file_path).name

    for job in DEPARTMENT_JOBS:
        dept = job["department"]
        if dept in name:
            return dept

    return ""


def normalize_saved_exports(saved_exports):
    """
    saved_exports를 [{'department': ..., 'path': ...}] 형태로 통일합니다.
    """
    normalized = []

    for item in saved_exports:
        if isinstance(item, dict):
            path = item.get("path") or item.get("file") or item.get("saved_path")
            department = item.get("department") or get_daily_department_from_filename(path or "")
        else:
            path = str(item)
            department = get_daily_department_from_filename(path)

        if not path:
            continue

        if not department:
            log(f"[WARN] Daily 파일에서 부서명을 판단하지 못해 건너뜁니다: {path}")
            continue

        normalized.append({"department": department, "path": str(path)})

    return normalized


def read_daily_rows_for_master(daily_path, department, master_headers, master_max_col):
    """
    Daily 엑셀 파일에서 마스터 컬럼 구조에 맞춰 행 데이터를 읽습니다.
    - Daily 헤더와 마스터 헤더를 이름 기준으로 매핑
    - A열 순번은 나중에 마스터에서 재부여하므로 비워둠
    - B열 부서명은 비어 있으면 파일 기준 department로 채움
    """
    daily_path = str(daily_path)

    if not os.path.exists(daily_path):
        raise FileNotFoundError(f"Daily 파일을 찾지 못했습니다: {daily_path}")

    wb = load_workbook_after_ready(
        daily_path,
        
        data_only=True,
        read_only=False,
        timeout=20,
        label=f"Daily 엑셀 파일({department})",
        require_writable=False,
    )

    if MASTER_SHEET_NAME in wb.sheetnames:
        ws = wb[MASTER_SHEET_NAME]
    else:
        ws = wb.active

    daily_max_col = get_effective_max_col(ws, header_row=1)
    daily_last_row = get_effective_max_row(ws, daily_max_col)
    daily_header_map = build_header_map(ws, header_row=1, max_col=daily_max_col)

    rows = []
    dept_key = normalize_excel_header("부서명")
    target_dept_norm = normalize_text(department)

    for src_row in range(2, daily_last_row + 1):
        # 빈 행 제외
        has_value = False
        for col in range(1, daily_max_col + 1):
            if ws.cell(row=src_row, column=col).value not in (None, ""):
                has_value = True
                break
        if not has_value:
            continue

        new_values = [None] * master_max_col

        for master_col in range(1, master_max_col + 1):
            # A열 순번은 마스터에서 재인덱싱
            if master_col == MASTER_INDEX_COL:
                new_values[master_col - 1] = None
                continue

            master_header = master_headers.get(master_col, "")
            header_key = normalize_excel_header(master_header)

            if not header_key:
                continue

            daily_col = daily_header_map.get(header_key)
            if daily_col:
                new_values[master_col - 1] = ws.cell(row=src_row, column=daily_col).value

        # 부서명 컬럼 보정
        daily_dept_value = None
        daily_dept_col = daily_header_map.get(dept_key)
        if daily_dept_col:
            daily_dept_value = ws.cell(row=src_row, column=daily_dept_col).value

        if normalize_excel_text(daily_dept_value):
            # Daily 파일 안에 부서명이 있는데 현재 처리 부서와 다르면 스킵
            if normalize_text(daily_dept_value) != target_dept_norm:
                continue
        else:
            daily_dept_value = department

        new_values[MASTER_DEPARTMENT_COL - 1] = daily_dept_value
        rows.append(new_values)

    wb.close()
    return rows


def insert_department_rows_to_master(ws, department, rows_to_add, master_max_col):
    """
    중복 제거 후 Daily 데이터를 마스터에 삽입합니다.
    APPEND_NEW_ROWS_TO_BOTTOM=True이면 마스터 전체 마지막 행 아래에 추가하고,
    False이면 해당 부서 마지막 행 바로 아래에 삽입합니다.
    삽입 후 해당 부서 A열을 1부터 재인덱싱합니다.
    """
    if not rows_to_add:
        return {"department": department, "read": 0, "inserted": 0, "duplicates": 0, "last_count": 0}

    existing_keys = build_existing_keys(
        ws,
        compare_start_col=DUPLICATE_COMPARE_START_COL,
        end_col=master_max_col,
    )

    unique_rows = []
    duplicate_count = 0

    for values in rows_to_add:
        key = tuple(normalize_excel_text(values[col - 1]) for col in range(DUPLICATE_COMPARE_START_COL, master_max_col + 1))

        if not any(key):
            continue

        if key in existing_keys:
            duplicate_count += 1
            continue

        existing_keys.add(key)
        unique_rows.append(values)

    if not unique_rows:
        last_count = reindex_department_numbers(ws, department, max_col=master_max_col)
        return {
            "department": department,
            "read": len(rows_to_add),
            "inserted": 0,
            "duplicates": duplicate_count,
            "last_count": last_count,
        }

    dept_last_row = find_department_last_row(ws, department, max_col=master_max_col)

    if APPEND_NEW_ROWS_TO_BOTTOM:
        # 요청사항 반영: 공정 순서와 무관하게 마스터 전체 마지막 행 아래에 추가
        insert_at = get_effective_max_row(ws, master_max_col) + 1
        template_row = dept_last_row if dept_last_row is not None else insert_at - 1
        log(f"[STEP] {department}: 마스터 전체 마지막 행 아래에 추가합니다.")
    else:
        # 기존 방식: 해당 부서의 마지막 행 바로 아래에 삽입하여 부서 블록을 유지
        if dept_last_row is None:
            insert_at = get_effective_max_row(ws, master_max_col) + 1
            template_row = insert_at - 1
            log(f"[WARN] 마스터에서 '{department}' 기존 행을 찾지 못했습니다. 마지막 행 아래에 추가합니다.")
        else:
            insert_at = dept_last_row + 1
            template_row = dept_last_row

    log(f"[STEP] {department}: 삽입 위치={insert_at}행, 삽입 대상={len(unique_rows)}건, 중복 제외={duplicate_count}건")

    ws.insert_rows(insert_at, amount=len(unique_rows))

    for offset, values in enumerate(unique_rows):
        dst_row = insert_at + offset
        copy_row_style(ws, template_row, dst_row, master_max_col)

        for col in range(1, master_max_col + 1):
            ws.cell(row=dst_row, column=col).value = values[col - 1]

    last_count = reindex_department_numbers(ws, department, max_col=master_max_col)

    return {
        "department": department,
        "read": len(rows_to_add),
        "inserted": len(unique_rows),
        "duplicates": duplicate_count,
        "last_count": last_count,
    }


def update_master_excel_with_daily_exports(saved_exports):
    """
    AIR에서 부서별로 내보낸 Daily 엑셀 파일을 마스터 파일에 반영합니다.

    처리 내용:
    1. Daily 파일을 부서명 기준으로 마스터 List 시트에 매핑
    2. A열 제외 B~마지막열 전체 값 기준 중복 제거
    3. 신규 데이터는 설정에 따라 마스터 전체 마지막 행 또는 해당 부서 마지막 행 아래에 삽입
    4. 삽입 후 각 부서 A열 순번 재인덱싱
    5. 백업 생성 후 마스터 파일에 저장
    """
    if not UPDATE_MASTER_AFTER_EXPORT:
        log("[SKIP] UPDATE_MASTER_AFTER_EXPORT=False 이므로 마스터 업데이트를 생략합니다.")
        return []

    master_path = Path(MASTER_EXCEL_PATH)

    if not master_path.exists():
        raise FileNotFoundError(f"마스터 엑셀 파일을 찾지 못했습니다: {master_path}")

    exports = normalize_saved_exports(saved_exports)

    if not exports:
        log("[WARN] 마스터에 반영할 Daily 파일이 없습니다.")
        return []

    log("\n" + "=" * 70)
    log("[START] 마스터 엑셀 Daily 데이터 삽입 시작")
    log(f" - 마스터 파일: {master_path}")
    log(f" - 대상 파일 수: {len(exports)}")
    log("=" * 70)

    if CREATE_MASTER_BACKUP:
        master_backup_dir = get_master_backup_dir(master_path)
        master_backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = master_backup_dir / f"{master_path.stem}_backup_{time.strftime('%Y%m%d_%H%M%S')}{master_path.suffix}"
        shutil.copy2(master_path, backup_path)
        log(f"[OK] 마스터 백업 생성: {backup_path}")

    wb = load_workbook_after_ready(
        master_path,
        data_only=False,
        read_only=False,
        timeout=45,
        label="마스터 엑셀 파일",
        require_writable=True,
    )

    if MASTER_SHEET_NAME in wb.sheetnames:
        ws = wb[MASTER_SHEET_NAME]
    else:
        ws = wb.active
        log(f"[WARN] '{MASTER_SHEET_NAME}' 시트를 찾지 못해 활성 시트 '{ws.title}'를 사용합니다.")

    master_max_col = get_effective_max_col(ws, header_row=1)
    master_headers = {
        col: ws.cell(row=1, column=col).value
        for col in range(1, master_max_col + 1)
    }

    summary = []

    for export in exports:
        department = export["department"]
        daily_path = export["path"]

        log("\n" + "-" * 60)
        log(f"[DEPT] {department}")
        log(f"[FILE] {daily_path}")

        daily_rows = read_daily_rows_for_master(
            daily_path=daily_path,
            department=department,
            master_headers=master_headers,
            master_max_col=master_max_col,
        )

        result = insert_department_rows_to_master(
            ws=ws,
            department=department,
            rows_to_add=daily_rows,
            master_max_col=master_max_col,
        )

        summary.append(result)

        log(
            f"[RESULT] {department}: Daily읽음={result['read']}건, "
            f"신규삽입={result['inserted']}건, 중복제외={result['duplicates']}건, "
            f"재인덱싱후총건수={result['last_count']}건"
        )

    # ------------------------------------------------------------
    # 최종 안전장치: 모든 Daily 업데이트 완료 후 마스터 전체 중복 제거
    # ------------------------------------------------------------
    final_duplicate_deleted = remove_duplicate_rows_in_master(
        ws=ws,
        compare_start_col=DUPLICATE_COMPARE_START_COL,
        end_col=master_max_col,
    )

    # 중복 삭제 후 A열 부서별 순번 다시 정리
    for job in DEPARTMENT_JOBS:
        reindex_department_numbers(
            ws=ws,
            department=job["department"],
            max_col=master_max_col,
        )

    log(f"[DUP-CLEAN] 최종 중복 삭제 수: {final_duplicate_deleted}건")

    try:
        wb.save(master_path)
    except PermissionError:
        raise PermissionError(
            "마스터 엑셀 파일 저장에 실패했습니다. "
            "파일이 Excel에서 열려 있으면 닫은 뒤 다시 실행하세요. "
            f"대상 파일: {master_path}"
        )
    finally:
        wb.close()

    log("\n" + "=" * 70)
    log("[OK] 마스터 엑셀 업데이트 완료")
    log(f" - 저장 파일: {master_path}")
    log("=" * 70)

    return summary


# =========================================================
# 11. 부서별 AIR 내보내기 실행
# =========================================================

def run_air_export_for_department(
    department: str,
    department_index: int,
    start_date: str,
    end_date: str,
    start_time: str,
    end_time: str,
    set_period: bool = True,
    set_dates: bool | None = None,
    set_times: bool | None = None,
    win=None,
) -> str:
    export_path = make_export_path(department)

    log("\n" + "=" * 70)
    log(f"[START] AIR 조회/내보내기 시작: {department}")
    log(f" - 부서 인덱스: {department_index}")
    log(f" - 기간: {start_date} {start_time} ~ {end_date} {end_time}")
    log(f" - 기간/시간 설정 여부: {set_period}")
    log(f" - 날짜 설정 여부: {set_dates if set_dates is not None else set_period}")
    log(f" - 시간 설정 여부: {set_times if set_times is not None else set_period}")
    log(f" - 저장 경로: {export_path}")
    log("=" * 70)

    if win is None:
        win = refresh_air_window()
        win = enter_inspection_history_strict(win)

    form = input_conditions(
        win=win,
        department=department,
        department_index=department_index,
        start_date=start_date,
        end_date=end_date,
        start_time=start_time,
        end_time=end_time,
        set_period=set_period,
        set_dates=set_dates,
        set_times=set_times,
    )

    click_search(win, form=form)

    saved_path = click_excel_export(win, export_path, form=form)

    try:
        log("[STEP] 다음 부서 입력 전 자동으로 열린 Daily 엑셀 파일 닫기")
        close_only_exported_excel_workbooks(
            [{"department": department, "path": saved_path}],
            timeout=8,
            wait_for_excel_appearance=True,
        )
    except Exception as e:
        log(f"[WARN] 내보낸 Daily 엑셀 파일 즉시 닫기 실패: {e}")

    try:
        bring_window_to_front(win)
        time.sleep(0.1)
    except Exception as e:
        log(f"[WARN] 내보내기 후 AIR 창 전면화 실패: {e}")

    log(f"[OK] AIR 내보내기 완료: {saved_path}")

    return saved_path


# =========================================================
# 12. 메인 실행
# =========================================================


def build_pipeline_steps():
    return (
        ["마스터 최근검사 확인"]
        + [f"{job['department']} 조회/내보내기" for job in DEPARTMENT_JOBS]
        + [
            "자동으로 열린 Excel 닫기",
            "마스터 업데이트",
            "중복 제거 및 저장",
            "AIR 종료",
        ]
    )


PIPELINE_STEPS = build_pipeline_steps()


def execute_automation_pipeline(progress_callback=None, air_ready: bool = False):
    global RUN_EXPORT_DIR
    global STOP_REQUESTED

    STOP_REQUESTED = False
    CONTROL_CACHE.clear()

    def update_progress(step_name: str, status: str, detail: str = ""):
        if progress_callback:
            progress_callback(step_name, status, detail)

    log("===== AIR 검사이력 부서별 엑셀 내보내기 + 마스터 업데이트 =====")
    log(f"CODE_VERSION: {CODE_VERSION}")

    Path(EXPORT_DIR).mkdir(parents=True, exist_ok=True)

    # 이번 실행에서 생성되는 Daily 파일 3개만 따로 관리하기 위한 실행 폴더를 먼저 생성합니다.
    # 예: AIR_Daily_Export\20260601_1
    RUN_EXPORT_DIR = create_run_export_dir()

    log("\n[INFO] 이번 실행 Daily 엑셀 저장 폴더")
    log(f"오늘 내보낸 파일 저장 폴더- {RUN_EXPORT_DIR}")

    # AIR/Excel에서 생성한 xlsx의 pageSetup collated 속성 때문에
    # openpyxl load_workbook()이 실패하는 문제를 방지합니다.
    # 마스터 엑셀 확인도 load_workbook()을 사용하므로 AIR 실행 전 먼저 적용합니다.
    patch_openpyxl_print_page_setup_collated()

    start_time = DAILY_START_TIME
    end_time = DAILY_END_TIME

    ensure_not_stopped()
    update_progress("마스터 최근검사 확인", "in_progress", "마스터 최근검사 확인 진행 중")

    # AIR를 조회하기 전에 마스터 엑셀을 먼저 확인하여
    # 각 부서별 최근검사 최댓값의 날짜를 부서별 From으로 사용합니다.
    query_plan = get_master_global_latest_recent_query_plan()

    query_basis_lines = []
    for job in DEPARTMENT_JOBS:
        department = job["department"]
        plan = query_plan.get(department, {})
        query_basis_lines.append(f"{department}: {plan.get('start_date', '')} {start_time}")

    if query_basis_lines:
        update_progress("__QUERY_BASIS__", "info", "\n".join(query_basis_lines))

    update_progress("마스터 최근검사 확인", "completed", "마스터 최근검사 확인 완료")

    log("\n[INFO] AIR 부서별 조회 조건")
    for job in DEPARTMENT_JOBS:
        department = job["department"]
        plan = query_plan[department]
        log(f" - {department}: From={plan['start_date']} {start_time} / To={plan['end_date']} {end_time}")

    log(f" - 기본 저장 폴더: {EXPORT_DIR}")
    log(f" - 이번 실행 저장 폴더: {RUN_EXPORT_DIR}")
    log(f" - 마스터 파일: {MASTER_EXCEL_PATH}")

    ensure_not_stopped()
    if air_ready:
        log("[INFO] AIR 실행/로그인 단계는 이미 완료되어 재실행을 생략합니다.")
    else:
        launch_air_if_needed()
    air_win = refresh_air_window()
    air_win = enter_inspection_history_strict(air_win)

    saved_exports = []

    for idx, job in enumerate(DEPARTMENT_JOBS):
        department = job["department"]
        department_index = job["combo_index"]
        plan = query_plan[department]

        # 부서별 최근검사 기준일이 다를 수 있으므로 날짜는 매번 설정합니다.
        # 시간은 항상 08:30~20:30 고정이므로 첫 부서에서만 설정합니다.
        set_dates = True
        set_times = idx == 0

        step_name = f"{department} 조회/내보내기"
        update_progress(step_name, "in_progress", f"{department} 조회/내보내기 진행 중")

        try:
            ensure_not_stopped()
            saved_path = run_air_export_for_department(
                department=department,
                department_index=department_index,
                start_date=plan["start_date"],
                end_date=plan["end_date"],
                start_time=start_time,
                end_time=end_time,
                set_period=True,
                set_dates=set_dates,
                set_times=set_times,
                win=air_win,
            )

            saved_exports.append({
                "department": department,
                "path": saved_path,
                "from": plan["start_date"],
                "to": plan["end_date"],
            })
            update_progress(step_name, "completed", f"{department} 조회/내보내기 완료")

        except Exception as e:
            log(f"[ERROR] {department} 내보내기 실패")
            log(str(e))
            update_progress(step_name, "error", f"{department} 조회/내보내기 실패")
            CONTROL_CACHE.clear()
            try:
                air_win = refresh_air_window()
                air_win = enter_inspection_history_strict(air_win)
                log("[INFO] 다음 부서 처리를 위해 AIR 검사이력 화면을 다시 확인했습니다.")
            except Exception as recover_error:
                log("[ERROR] AIR 검사이력 화면 복구 실패 - 남은 부서 처리를 중단합니다.")
                log(str(recover_error))
                break
            continue

    if saved_exports:
        log("\n[INFO] AIR 부서별 엑셀 내보내기 완료")
        for export in saved_exports:
            log(f" - {export['department']}: {export['path']}")

        ensure_not_stopped()
        update_progress("자동으로 열린 Excel 닫기", "in_progress", "자동으로 열린 Excel 닫기 진행 중")

        # 저장 완료 팝업 OK 후 자동으로 열린 Daily 엑셀 창 때문에
        # openpyxl이 파일을 읽거나 마스터를 업데이트할 때 문제가 생길 수 있습니다.
        # 단, 사용자가 작업 중인 다른 엑셀 파일은 절대 닫지 않고, 이번에 내보낸 파일만 닫습니다.
        closed_ok = close_only_exported_excel_workbooks(saved_exports, timeout=20)

        if not closed_ok:
            update_progress("자동으로 열린 Excel 닫기", "error", "자동으로 열린 Excel 닫기 실패")
            raise RuntimeError(
                "내보낸 엑셀 파일이 아직 Excel에서 열려 있거나 잠금 상태입니다. "
                "자동으로 열린 Daily 엑셀 파일을 닫은 뒤 다시 실행하세요."
            )

        update_progress("자동으로 열린 Excel 닫기", "completed", "자동으로 열린 Excel 닫기 완료")

        patch_openpyxl_print_page_setup_collated()

        try:
            ensure_not_stopped()
            update_progress("마스터 업데이트", "in_progress", "마스터 업데이트 진행 중")
            update_progress("중복 제거 및 저장", "in_progress", "중복 제거 및 저장 진행 중")
            update_master_excel_with_daily_exports(saved_exports)
            update_progress("마스터 업데이트", "completed", "마스터 업데이트 완료")
            update_progress("중복 제거 및 저장", "completed", "중복 제거 및 저장 완료")
        except Exception as e:
            log("[ERROR] 마스터 엑셀 업데이트 실패")
            log(str(e))
            update_progress("마스터 업데이트", "error", "마스터 업데이트 실패")
            update_progress("중복 제거 및 저장", "error", "중복 제거 및 저장 실패")
            raise

        ensure_not_stopped()
        update_progress("AIR 종료", "in_progress", "AIR 종료 진행 중")
        if close_air_after_completion(timeout=10):
            update_progress("AIR 종료", "completed", "AIR 종료 완료")
        else:
            update_progress("AIR 종료", "error", "AIR 종료 확인 실패")
    else:
        log("[WARN] 저장된 Daily 파일이 없어 마스터 엑셀 업데이트를 생략합니다.")

    return saved_exports


def main():
    execute_automation_pipeline()


class AirAutomationApp(ctk.CTk if ctk else object):
    CLEAN_KEYWORDS = [
        "[OK]",
        "[WARN]",
        "[ERROR]",
        "[START]",
        "[STEP]",
        "[RESULT]",
        "[DUP-CLEAN]",
        "[MASTER-CHECK-RESULT]",
    ]

    def __init__(self, auto_mode: bool = False):
        super().__init__()

        self.auto_mode = auto_mode

        self.title("AIR 검사이력 자동 내보내기 + 마스터 업데이트")
        self.geometry("1180x960")
        self.minsize(1024, 840)

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.font_family = "맑은 고딕"
        self.font_title = ctk.CTkFont(family=self.font_family, size=24, weight="bold")
        self.font_section = ctk.CTkFont(family=self.font_family, size=16, weight="bold")
        self.font_body = ctk.CTkFont(family=self.font_family, size=12)
        self.font_body_bold = ctk.CTkFont(family=self.font_family, size=13, weight="bold")

        self.raw_log_visible = False
        self.worker_thread = None
        self._closing = False
        self.export_dir_var = ctk.StringVar(value="")
        self.master_file_var = ctk.StringVar(value="")
        self.paths_initialized = False
        self.user_id_var = ctk.StringVar(value="")
        self.password_var = ctk.StringVar(value="")
        self.save_credentials_var = ctk.BooleanVar(value=True)
        self._pending_uid = ""
        self._pending_pw = ""

        self.step_states = {step: "대기" for step in PIPELINE_STEPS}
        self.step_labels = {}
        self.department_vars = {}

        self._build_layout()
        self._reset_progress_ui()
        self._load_paths_on_startup()
        self._load_saved_credentials()
        self.protocol("WM_DELETE_WINDOW", self._on_window_close)

        if auto_mode:
            # 자동 실행 모드: 시작 후 2초 뒤 자동으로 로그인 + 실행
            self._safe_after(2000, self._auto_start)

    def _build_layout(self):
        self.main_container = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.main_container.pack(fill="both", expand=True, padx=16, pady=16)
        self.main_container.grid_columnconfigure(0, weight=1)

        self.header_frame = ctk.CTkFrame(self.main_container, corner_radius=10)
        self.header_frame.grid(row=0, column=0, sticky="ew", padx=0, pady=(0, 10))
        self.header_frame.grid_columnconfigure(0, weight=1)

        self.title_label = ctk.CTkLabel(
            self.header_frame,
            text="AIR 검사이력 자동 내보내기 + 마스터 업데이트",
            font=self.font_title,
            anchor="w",
        )
        self.title_label.grid(row=0, column=0, sticky="w", padx=14, pady=(12, 2))

        self.version_label = ctk.CTkLabel(
            self.header_frame,
            text=f"Version: {GUI_VERSION}",
            font=self.font_body,
            anchor="w",
        )
        self.version_label.grid(row=1, column=0, sticky="w", padx=14, pady=(0, 12))

        self.action_frame = ctk.CTkFrame(self.main_container, corner_radius=10)
        self.action_frame.grid(row=1, column=0, sticky="ew", padx=0, pady=(0, 10))

        self.action_button_row = ctk.CTkFrame(self.action_frame, fg_color="transparent")
        self.action_button_row.grid(row=0, column=0, sticky="w", padx=14, pady=(14, 10))

        self.run_button = ctk.CTkButton(
            self.action_button_row,
            text="실행",
            width=120,
            fg_color="#3b82f6",
            hover_color="#2563eb",
            text_color="#ffffff",
            font=self.font_body_bold,
            command=self._on_run_clicked,
        )
        self.run_button.grid(row=0, column=0, padx=(0, 8), pady=0)

        self.stop_button = ctk.CTkButton(
            self.action_button_row,
            text="중지",
            width=90,
            fg_color="#ff727e",
            hover_color="#fd6d75",
            text_color="#ffffff",
            font=self.font_body_bold,
            command=self._on_stop_clicked,
            state="disabled",
        )
        self.stop_button.grid(row=0, column=1, padx=(0, 8), pady=0)

        self.open_export_folder_button = ctk.CTkButton(
            self.action_button_row,
            text="오늘 내보낸 파일 저장 폴더 열기",
            width=130,
            fg_color="#60a5fa",
            hover_color="#3b82f6",
            text_color="#ffffff",
            font=self.font_body_bold,
            command=self._open_export_folder,
        )
        self.open_export_folder_button.grid(row=0, column=2, padx=(0, 8), pady=0)

        self.open_master_file_button = ctk.CTkButton(
            self.action_button_row,
            text="마스터 파일 열기",
            width=160,
            fg_color="#60a5fa",
            hover_color="#3b82f6",
            text_color="#ffffff",
            font=self.font_body_bold,
            command=self._open_master_file,
        )
        self.open_master_file_button.grid(row=0, column=3, padx=(0, 0), pady=0)

        self.current_status_label = ctk.CTkLabel(
            self.action_frame,
            text="현재 상태: 대기 중",
            font=self.font_body_bold,
            anchor="center",
            corner_radius=8,
            padx=10,
            pady=4,
            fg_color="#7CADFC",
            text_color="#ffffff",
        )
        self.current_status_label.grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 6))

        self.path_input_frame = ctk.CTkFrame(self.action_frame, fg_color="transparent")
        self.path_input_frame.grid(row=2, column=0, sticky="ew", padx=14, pady=(0, 10))
        self.path_input_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            self.path_input_frame,
            text="오늘 내보낸 파일 저장 위치",
            font=self.font_body,
            anchor="w",
        ).grid(row=0, column=0, sticky="w", pady=(0, 4))

        self.export_dir_entry = ctk.CTkEntry(
            self.path_input_frame,
            textvariable=self.export_dir_var,
            font=self.font_body,
        )
        self.export_dir_entry.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        self.export_dir_entry.bind("<FocusOut>", self._commit_path_inputs)
        self.export_dir_entry.bind("<Return>", self._commit_path_inputs)

        ctk.CTkLabel(
            self.path_input_frame,
            text="마스터 파일 경로",
            font=self.font_body,
            anchor="w",
        ).grid(row=2, column=0, sticky="w", pady=(0, 4))

        self.master_file_entry = ctk.CTkEntry(
            self.path_input_frame,
            textvariable=self.master_file_var,
            font=self.font_body,
        )
        self.master_file_entry.grid(row=3, column=0, sticky="ew", pady=(0, 8))
        self.master_file_entry.bind("<FocusOut>", self._commit_path_inputs)
        self.master_file_entry.bind("<Return>", self._commit_path_inputs)

        self.path_hint_label = ctk.CTkLabel(
            self.path_input_frame,
            text="경로를 입력한 뒤 Enter를 누르거나 다른 칸으로 이동하면 저장됩니다. 저장되어야 실행할 수 있습니다.",
            font=self.font_body,
            anchor="w",
            justify="left",
            text_color="#ff9e9e",
        )
        self.path_hint_label.grid(row=4, column=0, sticky="w", pady=(2, 0))

        self._update_path_hint_visibility()

        self.action_frame.grid_columnconfigure(0, weight=1)

        # ── AIR 로그인 설정 프레임 (row=2) ─────────────────────────────────
        self.login_frame = ctk.CTkFrame(self.main_container, corner_radius=10)
        self.login_frame.grid(row=2, column=0, sticky="ew", padx=0, pady=(0, 10))
        self.login_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            self.login_frame,
            text="AIR 로그인 설정",
            font=self.font_section,
            anchor="w",
        ).grid(row=0, column=0, sticky="w", padx=14, pady=(12, 6))

        self.login_input_row = ctk.CTkFrame(self.login_frame, fg_color="transparent")
        self.login_input_row.grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 4))
        self.login_input_row.grid_columnconfigure(1, weight=1)
        self.login_input_row.grid_columnconfigure(3, weight=1)

        ctk.CTkLabel(
            self.login_input_row, text="아이디", font=self.font_body, anchor="w", width=60,
        ).grid(row=0, column=0, sticky="w", padx=(0, 6))

        self.user_id_entry = ctk.CTkEntry(
            self.login_input_row, textvariable=self.user_id_var, font=self.font_body, width=180,
        )
        self.user_id_entry.grid(row=0, column=1, sticky="ew", padx=(0, 16))

        ctk.CTkLabel(
            self.login_input_row, text="비밀번호", font=self.font_body, anchor="w", width=60,
        ).grid(row=0, column=2, sticky="w", padx=(0, 6))

        self.password_entry = ctk.CTkEntry(
            self.login_input_row, textvariable=self.password_var, font=self.font_body, show="●", width=180,
        )
        self.password_entry.grid(row=0, column=3, sticky="ew", padx=(0, 0))

        self.login_option_row = ctk.CTkFrame(self.login_frame, fg_color="transparent")
        self.login_option_row.grid(row=2, column=0, sticky="w", padx=14, pady=(4, 4))

        self.save_credentials_check = ctk.CTkCheckBox(
            self.login_option_row,
            text="자격증명 저장 (다음 실행 시 자동 사용)",
            variable=self.save_credentials_var,
            font=self.font_body,
            text_color="#ffffff",
            border_color="#93c5fd",
            fg_color="#3b82f6",
            hover_color="#2563eb",
            checkmark_color="#ffffff",
        )
        self.save_credentials_check.grid(row=0, column=0, padx=(0, 20))

        self.login_status_label = ctk.CTkLabel(
            self.login_frame,
            text="아이디/비밀번호를 입력하면 실행 버튼 클릭 시 AIR 자동 로그인 후 실행됩니다.",
            font=self.font_body,
            anchor="w",
            text_color="#93c5fd",
        )
        self.login_status_label.grid(row=3, column=0, sticky="w", padx=14, pady=(0, 10))

        self.condition_summary_frame = ctk.CTkFrame(self.main_container, corner_radius=10)
        self.condition_summary_frame.grid(row=3, column=0, sticky="ew", padx=0, pady=(0, 10))
        self.condition_summary_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            self.condition_summary_frame,
            text="조회 조건 요약",
            font=self.font_section,
            anchor="w",
        ).grid(row=0, column=0, sticky="w", padx=14, pady=(12, 6))

        self.start_time_label = ctk.CTkLabel(
            self.condition_summary_frame,
            text=f"조회 시작 시간: {DAILY_START_TIME}",
            font=self.font_body,
            anchor="w",
        )
        self.start_time_label.grid(row=1, column=0, sticky="w", padx=14, pady=2)

        self.end_time_label = ctk.CTkLabel(
            self.condition_summary_frame,
            text=f"조회 종료 시간: {DAILY_END_TIME}",
            font=self.font_body,
            anchor="w",
        )
        self.end_time_label.grid(row=2, column=0, sticky="w", padx=14, pady=2)

        self.query_basis_label = ctk.CTkLabel(
            self.condition_summary_frame,
            text="부서별 조회 시작일: 실행 후 자동 계산",
            font=self.font_body,
            anchor="w",
            justify="left",
        )
        self.query_basis_label.grid(row=3, column=0, sticky="w", padx=14, pady=(2, 8))

        self.department_check_area = ctk.CTkFrame(self.condition_summary_frame, fg_color="transparent")
        self.department_check_area.grid(row=4, column=0, sticky="w", padx=14, pady=(0, 12))

        for idx, job in enumerate(DEPARTMENT_JOBS):
            department = job["department"]
            var = ctk.BooleanVar(value=False)
            chk = ctk.CTkCheckBox(
                self.department_check_area,
                text=department,
                variable=var,
                font=self.font_body,
                text_color="#ffffff",
                border_color="#93c5fd",
                fg_color="#3b82f6",
                hover_color="#2563eb",
                checkmark_color="#ffffff",
                state="disabled",
            )
            chk.grid(row=0, column=idx, padx=(0, 20), pady=2, sticky="w")
            self.department_vars[department] = var

        self.progress_frame = ctk.CTkFrame(self.main_container, corner_radius=10)
        self.progress_frame.grid(row=4, column=0, sticky="ew", padx=0, pady=(0, 10))
        self.progress_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            self.progress_frame,
            text="진행 상태",
            font=self.font_section,
            anchor="w",
        ).grid(row=0, column=0, sticky="w", padx=14, pady=(12, 6))

        self.total_progress_label = ctk.CTkLabel(
            self.progress_frame,
            text="전체 진행률: 0%",
            font=self.font_body,
            anchor="w",
        )
        self.total_progress_label.grid(row=1, column=0, sticky="w", padx=14, pady=(0, 4))

        self.total_progress_bar = ctk.CTkProgressBar(self.progress_frame)
        self.total_progress_bar.grid(row=2, column=0, sticky="ew", padx=14, pady=(0, 10))
        self.total_progress_bar.set(0.0)

        self.step_status_list = ctk.CTkFrame(self.progress_frame)
        self.step_status_list.grid(row=3, column=0, sticky="ew", padx=14, pady=(0, 12))
        self.step_status_list.grid_columnconfigure(0, weight=1)

        for idx, step in enumerate(PIPELINE_STEPS, start=1):
            label = ctk.CTkLabel(
                self.step_status_list,
                text=f"{idx}. {step:<20} 대기",
                font=self.font_body,
                anchor="w",
                justify="left",
            )
            label.grid(row=idx - 1, column=0, sticky="ew", pady=1)
            self.step_labels[step] = label

        self.clean_log_frame = ctk.CTkFrame(self.main_container, corner_radius=10)
        self.clean_log_frame.grid(row=5, column=0, sticky="ew", padx=0, pady=(0, 10))
        self.clean_log_frame.grid_columnconfigure(0, weight=1)
        self.clean_log_frame.grid_rowconfigure(1, weight=1)

        self.clean_log_title = ctk.CTkLabel(
            self.clean_log_frame,
            text="로그",
            font=self.font_section,
            anchor="w",
        )
        self.clean_log_title.grid(row=0, column=0, sticky="w", padx=14, pady=(12, 6))

        self.clean_log_textbox = ctk.CTkTextbox(self.clean_log_frame, height=240, wrap="word", font=self.font_body)
        self.clean_log_textbox.grid(row=1, column=0, sticky="nsew", padx=14, pady=(0, 12))

        self.raw_log_frame = ctk.CTkFrame(self.main_container, corner_radius=10)
        self.raw_log_frame.grid(row=6, column=0, sticky="ew", padx=0, pady=(0, 0))
        self.raw_log_frame.grid_columnconfigure(0, weight=1)
        self.raw_log_frame.grid_rowconfigure(1, weight=1)

        self.raw_log_toggle_button = ctk.CTkButton(
            self.raw_log_frame,
            text="Raw Log 보기 ⏷",
            font=self.font_body_bold,
            command=self._toggle_raw_log,
        )
        self.raw_log_toggle_button.grid(row=0, column=0, sticky="ew", padx=14, pady=(12, 10))

        self.raw_log_textbox = ctk.CTkTextbox(self.raw_log_frame, height=180, wrap="none", font=self.font_body)
        self.raw_log_textbox.grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 12))
        self.raw_log_textbox.grid_remove()

    def _set_status_text(self, text: str):
        if getattr(self, "_closing", False):
            return
        color_map = {
            "대기 중": ("#374151", "#ffffff"),
            "실행 중": ("#1d4ed8", "#ffffff"),
            "중지 요청됨": ("#b45309", "#ffffff"),
            "중지됨": ("#6b7280", "#ffffff"),
            "완료": ("#047857", "#ffffff"),
            "오류": ("#b91c1c", "#ffffff"),
        }
        fg, tc = color_map.get(text, ("#64799B", "#e5e7eb"))
        try:
            self.current_status_label.configure(text=f"현재 상태: {text}", fg_color=fg, text_color=tc)
        except Exception:
            pass

    def _reset_progress_ui(self):
        if getattr(self, "_closing", False):
            return
        self.step_states = {step: "대기" for step in PIPELINE_STEPS}

        for idx, step in enumerate(PIPELINE_STEPS, start=1):
            self.step_labels[step].configure(text=f"{idx}. {step:<20} 대기")

        for department, var in self.department_vars.items():
            var.set(False)

        self.query_basis_label.configure(text="부서별 조회 시작일: 실행 후 자동 계산")

        self.total_progress_bar.set(0.0)
        self.total_progress_label.configure(text="전체 진행률: 0%")

    def _safe_after(self, delay_ms: int, callback, *args):
        if getattr(self, "_closing", False):
            return None

        def run_if_open():
            if getattr(self, "_closing", False):
                return
            try:
                callback(*args)
            except Exception:
                pass

        try:
            return self.after(delay_ms, run_if_open)
        except Exception:
            return None

    def _on_window_close(self):
        self._closing = True
        request_stop()
        try:
            unregister_log_listener(self._on_log_line)
        except Exception:
            pass
        try:
            self.destroy()
        except Exception:
            pass

    def _on_progress(self, step_name: str, status: str, detail: str = ""):
        self._safe_after(0, self._apply_progress, step_name, status, detail)

    def _apply_progress(self, step_name: str, status: str, detail: str = ""):
        if getattr(self, "_closing", False):
            return
        if step_name == "__QUERY_BASIS__":
            self.query_basis_label.configure(
                text=f"부서별 조회 시작일:\n{detail}\n(마스터 List 시트의 부서별 최근검사 최신 날짜)"
            )
            return

        status_text_map = {
            "pending": "대기",
            "in_progress": "진행 중",
            "completed": "완료",
            "error": "오류",
        }

        text_status = status_text_map.get(status, status)
        self.step_states[step_name] = text_status

        if step_name in self.step_labels:
            idx = PIPELINE_STEPS.index(step_name) + 1
            self.step_labels[step_name].configure(text=f"{idx}. {step_name:<20} {text_status}")

        for department, var in self.department_vars.items():
            if step_name == f"{department} 조회/내보내기":
                var.set(status == "completed")

        completed_count = sum(1 for s in self.step_states.values() if s == "완료")
        ratio = completed_count / max(1, len(PIPELINE_STEPS))
        self.total_progress_bar.set(ratio)
        self.total_progress_label.configure(text=f"전체 진행률: {int(ratio * 100)}%")

        if detail:
            self._append_clean_log(detail)

    def _append_raw_log(self, message: str):
        if getattr(self, "_closing", False):
            return
        try:
            self.raw_log_textbox.insert("end", f"{message}\n")
            self.raw_log_textbox.see("end")
        except Exception:
            pass

    def _append_clean_log(self, message: str):
        if getattr(self, "_closing", False):
            return
        stamp = datetime.now().strftime("%H:%M:%S")
        try:
            self.clean_log_textbox.insert("end", f"{stamp}  {message}\n")
            self.clean_log_textbox.see("end")
        except Exception:
            pass

    def _is_clean_log_target(self, message: str) -> bool:
        if any(keyword in message for keyword in self.CLEAN_KEYWORDS):
            return True

        summary_words = ["실행", "조회", "내보내기", "저장", "업데이트", "중복", "완료", "실패", "오류", "중지"]
        return any(word in message for word in summary_words)

    def _on_log_line(self, message: str):
        self._safe_after(0, self._apply_log_line, message)

    def _apply_log_line(self, message: str):
        if getattr(self, "_closing", False):
            return
        self._append_raw_log(message)

        if self._is_clean_log_target(message):
            self._append_clean_log(message)

    def _toggle_raw_log(self):
        self.raw_log_visible = not self.raw_log_visible

        if self.raw_log_visible:
            self.raw_log_textbox.grid()
            self.raw_log_toggle_button.configure(text="Raw Log 숨기기 ⏶")
        else:
            self.raw_log_textbox.grid_remove()
            self.raw_log_toggle_button.configure(text="Raw Log 보기 ⏷")

    def _refresh_path_labels(self):
        if not getattr(self, "paths_initialized", False):
            self._update_path_hint_visibility()
            return

        if hasattr(self, "export_dir_var"):
            self.export_dir_var.set(str(EXPORT_DIR or ""))
        if hasattr(self, "master_file_var"):
            self.master_file_var.set(str(MASTER_EXCEL_PATH or ""))
        self._update_path_hint_visibility()

    def _update_path_hint_visibility(self):
        if not hasattr(self, "path_hint_label"):
            return

        export_text = str(self.export_dir_var.get() if hasattr(self, "export_dir_var") else "").strip()
        master_text = str(self.master_file_var.get() if hasattr(self, "master_file_var") else "").strip()

        if export_text and master_text:
            self.path_hint_label.grid_remove()
        else:
            self.path_hint_label.grid()

    def _commit_path_inputs(self, event=None):
        export_dir = normalize_user_path(self.export_dir_var.get() if hasattr(self, "export_dir_var") else "")
        master_path = normalize_user_path(self.master_file_var.get() if hasattr(self, "master_file_var") else "")

        if not export_dir.strip() or not master_path.strip():
            self.run_button.configure(state="disabled")
            self._update_path_hint_visibility()
            return

        if self._apply_path_settings(export_dir, master_path, require_master_exists=False):
            self.run_button.configure(state="normal")
            self._update_path_hint_visibility()

    def _apply_path_settings(self, export_dir: str, master_path: str, require_master_exists: bool = True) -> bool:
        global EXPORT_DIR
        global MASTER_EXCEL_PATH

        export_dir = normalize_user_path(export_dir)
        master_path = normalize_user_path(master_path)

        if not export_dir:
            messagebox.showwarning("경로 입력 필요", "저장 폴더 경로를 입력해 주세요.")
            return False

        if not master_path:
            messagebox.showwarning("경로 입력 필요", "마스터 파일 경로를 입력해 주세요.")
            return False

        try:
            Path(export_dir).mkdir(parents=True, exist_ok=True)
        except Exception as e:
            messagebox.showerror("저장 폴더 오류", f"저장 폴더를 준비할 수 없습니다.\n{export_dir}\n\n{e}")
            return False

        if require_master_exists and not Path(master_path).is_file():
            messagebox.showerror("마스터 파일 오류", f"선택한 마스터 파일을 찾을 수 없습니다.\n{master_path}")
            return False

        EXPORT_DIR = export_dir
        MASTER_EXCEL_PATH = master_path
        self.paths_initialized = True

        if not self._persist_paths():
            return False

        self._refresh_path_labels()
        self.run_button.configure(state="normal")
        self._update_path_hint_visibility()
        try:
            self._append_clean_log("사용자 경로 설정이 저장되었습니다.")
        except Exception:
            pass

        return True

    def _configure_paths_button_clicked(self):
        # 기존 경로 설정 버튼/창은 제거되었습니다.
        self._set_status_text("경로를 상단 입력칸에 먼저 저장해 주세요")
        try:
            self.export_dir_entry.focus_set()
        except Exception:
            pass

    def _configure_paths_interactive(self, require_master_exists: bool = True) -> bool:
        return self._commit_path_inputs()

    def _is_path_config_ready(self) -> bool:
        export_dir = normalize_user_path(self.export_dir_var.get() if hasattr(self, "export_dir_var") else "")
        master_path = normalize_user_path(self.master_file_var.get() if hasattr(self, "master_file_var") else "")
        return bool(export_dir and master_path)

    def _persist_paths(self):
        try:
            save_user_path_settings(EXPORT_DIR, MASTER_EXCEL_PATH)
        except Exception as e:
            messagebox.showerror("설정 저장 실패", f"경로 설정 저장에 실패했습니다.\n{e}")
            return False
        return True

    def _open_export_folder(self):
        target = RUN_EXPORT_DIR if RUN_EXPORT_DIR else EXPORT_DIR

        if not target:
            messagebox.showwarning("경로 없음", "열 수 있는 실행 폴더가 없습니다.")
            return

        try:
            Path(target).mkdir(parents=True, exist_ok=True)
        except Exception as e:
            messagebox.showerror("폴더 열기 실패", f"저장 폴더를 준비할 수 없습니다.\n{target}\n\n{e}")
            return

        os.startfile(target)

    def _open_master_file(self):
        if not Path(MASTER_EXCEL_PATH).is_file():
            messagebox.showwarning("파일 없음", f"마스터 파일을 찾지 못했습니다.\n{MASTER_EXCEL_PATH}")
            return

        os.startfile(MASTER_EXCEL_PATH)

    def _ensure_paths_ready(self) -> bool:
        export_dir = normalize_user_path(self.export_dir_var.get() if hasattr(self, "export_dir_var") else "")
        master_path = normalize_user_path(self.master_file_var.get() if hasattr(self, "master_file_var") else "")

        if not export_dir or not master_path:
            self._set_status_text("경로 설정 필요")
            self._append_clean_log("실행 전에 저장 폴더와 마스터 파일 경로를 먼저 입력하고 저장해 주세요.")
            try:
                self.export_dir_entry.focus_set()
            except Exception:
                pass
            return False

        if not Path(master_path).is_file():
            messagebox.showerror("마스터 파일 오류", f"마스터 파일을 찾을 수 없습니다.\n{master_path}")
            return False

        if not self._apply_path_settings(export_dir, master_path, require_master_exists=True):
            return False

        self._refresh_path_labels()
        return True

    def _load_paths_on_startup(self):
        settings = load_user_path_settings()
        saved_export = normalize_user_path(remap_path_to_current_windows_user(settings.get("export_dir", "")))
        saved_master = normalize_user_path(normalize_master_excel_path(settings.get("master_excel_path", "")) if settings.get("master_excel_path", "") else "")

        global EXPORT_DIR
        global MASTER_EXCEL_PATH
        EXPORT_DIR = saved_export or EXPORT_DIR or DEFAULT_EXPORT_DIR
        MASTER_EXCEL_PATH = saved_master or MASTER_EXCEL_PATH or DEFAULT_MASTER_EXCEL_PATH

        self.paths_initialized = bool(saved_export and saved_master)
        self.export_dir_var.set(saved_export)
        self.master_file_var.set(saved_master)

        self._refresh_path_labels()

        if self._is_path_config_ready():
            self.run_button.configure(state="normal")
            self._set_status_text("대기 중")
        else:
            self.run_button.configure(state="disabled")
            self._set_status_text("경로 설정 필요")
            self._append_clean_log("실행 전에 저장 폴더와 마스터 파일 경로를 먼저 입력해 주세요.")
            try:
                self._safe_after(100, self.export_dir_entry.focus_set)
            except Exception:
                pass

    def _on_run_clicked(self):
        if self.worker_thread and self.worker_thread.is_alive():
            return

        if not self._ensure_paths_ready():
            self._set_status_text("대기 중")
            return

        # 실행 시점의 자격증명을 스냅샷으로 보관합니다.
        self._pending_uid = self.user_id_var.get().strip() if hasattr(self, "user_id_var") else ""
        self._pending_pw = self.password_var.get() if hasattr(self, "password_var") else ""

        # 자격증명 저장 체크박스가 체크된 경우 저장
        if self._pending_uid and self._pending_pw and getattr(self, "save_credentials_var", None) and self.save_credentials_var.get():
            try:
                save_air_credentials(self._pending_uid, self._pending_pw)
            except Exception:
                pass

        self.clean_log_textbox.delete("1.0", "end")
        self.raw_log_textbox.delete("1.0", "end")
        self._reset_progress_ui()

        self._set_status_text("실행 중")

        self.run_button.configure(state="disabled")
        self.stop_button.configure(state="normal")

        register_log_listener(self._on_log_line)

        self.worker_thread = threading.Thread(target=self._run_worker, daemon=True)
        self.worker_thread.start()

    def _on_stop_clicked(self):
        request_stop()
        self._set_status_text("중지 요청됨")
        self._append_clean_log("중지 요청됨")

    # ── 로그인 관련 메서드 ──────────────────────────────────────────────────

    def _load_saved_credentials(self):
        """저장된 자격증명을 로드해서 ID/PW 필드에 채웁니다."""
        try:
            uid, pw = load_air_credentials()
            if uid:
                self.user_id_var.set(uid)
            if pw:
                self.password_var.set(pw)
            if uid:
                self.login_status_label.configure(
                    text="저장된 계정이 반영되었습니다. 실행 버튼을 누르면 AIR 자동 로그인 후 실행됩니다.",
                    text_color="#34d399",
                )
        except Exception:
            pass

    def _auto_start(self):
        """자동 실행 모드: 저장된 계정으로 AIR 로그인 후 파이프라인 실행"""
        uid, pw = load_air_credentials()
        if not uid:
            self._append_clean_log("[ERROR] 저장된 계정이 없습니다. 먼저 GUI에서 로그인 정보를 저장해 주세요.")
            self.login_status_label.configure(
                text="저장된 계정 없음 - 아이디/비밀번호를 입력하고 자격증명 저장 후 다시 등록하세요.",
                text_color="#f87171",
            )
            return

        self.user_id_var.set(uid)
        self.password_var.set(pw)
        self._append_clean_log(f"[INFO] 자동 실행 모드 - 계정: {uid}")
        # 저장된 자격증명을 pending으로 설정한 뒤 실행버튼 로직 그대로 사용
        self._pending_uid = uid
        self._pending_pw = pw
        self._on_run_clicked()

    def _run_worker(self):
        try:
            uid = getattr(self, "_pending_uid", "")
            pw = getattr(self, "_pending_pw", "")

            # 1단계: AIR 실행 (이미 실행 중이면 스킵)
            launch_air_if_needed()

            # 2단계: AIR 로그인 (자격증명이 입력된 경우만)
            if uid and pw:
                self._safe_after(0, lambda: self.login_status_label.configure(
                    text="AIR 로그인 진행 중...", text_color="#fbbf24"
                ))
                air_auto_login(uid, pw)
                self._safe_after(0, lambda: self.login_status_label.configure(
                    text="AIR 로그인 완료", text_color="#34d399"
                ))
            else:
                log("[INFO] 자격증명 미입력 - AIR가 이미 로그인된 상태로 가정하고 진행합니다.")

            # 3단계: 기존 자동화 파이프라인 실행
            execute_automation_pipeline(progress_callback=self._on_progress, air_ready=True)

            self._safe_after(0, self._set_status_text, "완료")
            self._safe_after(0, self._append_clean_log, "전체 작업 완료")

        except UserStopRequested:
            self._safe_after(0, self._set_status_text, "중지됨")
            self._safe_after(0, self._append_clean_log, "사용자 요청으로 작업 중지")

        except Exception as e:
            self._safe_after(0, self._set_status_text, "오류")
            self._safe_after(0, self._append_clean_log, f"오류 요약: {e}")
            self._safe_after(0, self._append_raw_log, traceback.format_exc())

        finally:
            unregister_log_listener(self._on_log_line)
            self._safe_after(0, lambda: self.run_button.configure(state="normal"))
            self._safe_after(0, lambda: self.stop_button.configure(state="disabled"))


def launch_gui(auto_mode: bool = False):
    app = AirAutomationApp(auto_mode=auto_mode)
    app.mainloop()


if __name__ == "__main__":
    _auto_mode = "--auto" in sys.argv

    if ctk is None:
        print("[WARN] customtkinter를 찾지 못해 CLI 모드로 실행합니다. (pip install customtkinter)")
        try:
            if _auto_mode:
                uid, pw = load_air_credentials()
                if uid:
                    launch_air_if_needed()
                    air_auto_login(uid, pw)
                else:
                    print("[ERROR] 저장된 계정이 없습니다. GUI에서 먼저 자격증명을 저장해 주세요.")
            main()
        except Exception as e:
            print("\n[ERROR] 자동화 중 오류 발생")
            print(e)
            print("\n오류 직전 로그를 확인하세요.")
        finally:
            if not _auto_mode:
                input("\n종료하려면 Enter를 누르세요...")
    else:
        launch_gui(auto_mode=_auto_mode)
